import os
import math
import json
import logging
import platform
from dataclasses import dataclass
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import re
import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import customtkinter as ctk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.patches import Rectangle

def ask_yes_no_cancel(title, message):
    dialog = ctk.CTkToplevel()
    dialog.title(title)
    dialog.geometry("300x140")       # ширина x высота
    dialog.resizable(False, False)
    dialog.grab_set()                # делаем модальным
    dialog.focus_force()

    # Текст
    ctk.CTkLabel(
        dialog, text=message,
        font=("Segoe UI", 12),        # размер шрифта
        wraplength=280,
        justify="center"
    ).pack(pady=15)

    result = {"value": None}

    def set_result(value):
        result["value"] = value
        dialog.destroy()

    # Кнопки
    btn_frame = ctk.CTkFrame(dialog)
    btn_frame.pack(pady=10)

    ctk.CTkButton(btn_frame, text="Да", width=70, command=lambda: set_result(True)).pack(side="left", padx=5)
    ctk.CTkButton(btn_frame, text="Нет", width=70, command=lambda: set_result(False)).pack(side="left", padx=5)
    ctk.CTkButton(btn_frame, text="Отмена", width=70, command=lambda: set_result(None)).pack(side="left", padx=5)

    dialog.wait_window()
    return result["value"]

# =============================
# Константы и логирование
# =============================
MAX_PARTS = 200
MATERIAL_DB = "mat_data.xlsx"  # опционально, если используете внешний файл
CONFIG_FILE = "gridmaster_config.ini"  # JSON-конфиг
THEME_KEY = "ui_theme"          # "dark" | "light"
COLOR_THEME = "dark-blue"       # палитра customtkinter

logging.basicConfig(
    filename='gridmaster_log.log',
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# =============================
# Данные проекта
# =============================
from dataclasses import dataclass

@dataclass
class ProjectPart:
    position: str
    length: float
    width: float
    quantity: int
    angle: str
    bumper: float
    work: float
    radius: float
    perimeter: float = 0.0
    bumper_length: float = 0.0
    radius_part: float = 1.0 
    subtract_bumper: bool = True 

# =============================
# Основной интерфейс
# =============================
class GridMasterInterface:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Grid Master V.4 ©Viiptag")
        self.root.geometry("1200x800")
        self._tv_to_part: dict[str, ProjectPart] = {}
        self._part_editor_win = None
        self._part_editor_opening = False

        # Данные
        self.project_parts: List[ProjectPart] = []
        self.mat_data: List[Dict] = []
        self.material_prices: List[Tuple[str, float, str]] = []  # (наименование, цена, ед.)
        self.frame_weights: List[Tuple[str, float]] = []  # (тип, кг/м)
        self.angle_weights: List[Tuple[str, float]] = []  # (тип, кг/м)
        self.bumper_weights: List[Tuple[str, float]] = []  # (тип, кг/м)

        self.current_project_path: str = ""
        self.last_export_dir: Optional[str] = None
        self._last_article: Optional[str] = None
        self._cutting_results = []    # сюда положим результат раскроя для предпросмотра

        # UI переменные
        self.material_combos: Dict[str, ttk.Combobox] = {
            "Обрамление": None,
            "Перфоуголок": None,
            "Отбойник": None,
        }
        

        # Инициализация UI
        self.setup_styles()
        self.setup_main_frames()
        self.setup_project_tab()
        self.setup_tables_tab()
        self.setup_status_bar()
        self.load_config()
        # поле "Название проекта" всегда пустое при запуске
        try:
            self.project_name_entry.delete(0, "end")
        except Exception:
            pass

        # Загрузка данных
        self.load_all_data()
        self.refresh_all_tables()
        self._refresh_article_combobox()
        #self.load_last_project()

        # События и завершение
        self.setup_shortcuts()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
    def _set_project_name_from_file(self, file_path: str):
        import os
        try:
            base = os.path.splitext(os.path.basename(file_path))[0]
            self.project_name_entry.delete(0, "end")
            self.project_name_entry.insert(0, base)
            # если хочешь запоминать имя в конфиге — оставь строку ниже
            self.save_config()
        except Exception:
            pass

        
    def _init_default_values(self):
        """
        НЕ пересоздает self.material_entries.
        Только подставляет дефолтные значения в уже созданные Entry,
        если они пустые.
        """
        try:
            entries = getattr(self, "material_entries", None)
            if not entries:
                return  # словарь еще не инициализирован UI — ничего не делаем

            defaults = {"frame": "0.79", "angle": "2.0", "bumper": "0.79", "work": "30.0"}
            for key, entry in entries.items():
                try:
                    cur = (entry.get() or "").strip()
                except Exception:
                    continue
                if cur == "":
                    entry.insert(0, defaults.get(key, "0.0"))

            if hasattr(self, "zinc_entry"):
                if not (self.zinc_entry.get() or "").strip():
                    self.zinc_entry.insert(0, "0.6")
            if hasattr(self, "frame_coef_entry"):
                if not (self.frame_coef_entry.get() or "").strip():
                    self.frame_coef_entry.insert(0, "1.05")
        except Exception as e:
            logging.error(f"_init_default_values error: {e}")

        
        

    # =============================
    # UI
    # =============================
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", rowheight=25)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        style.configure("Cost.Treeview", font=("Segoe UI", 12))
        style.configure("Cost.Treeview.Heading", font=("Segoe UI", 12, "bold"))
        style.configure("Weight.Treeview", font=("Segoe UI", 12))
        style.configure("Weight.Treeview.Heading", font=("Segoe UI", 12, "bold"))


    def setup_main_frames(self):
        # Верхняя панель с переключателем темы и названием проекта
        topbar = ctk.CTkFrame(self.root)
        topbar.pack(fill='x', padx=5, pady=(5,0))

        ctk.CTkLabel(topbar, text="Тема:").pack(side='left', padx=(8,4), pady=6)

        def _on_theme_toggle():
            new_mode = "light" if self.theme_switch.get() else "dark"
            self._apply_theme_and_save(new_mode)

        # switch: ON=light, OFF=dark (по умолчанию OFF для dark)
        self.theme_switch = ctk.CTkSwitch(
            topbar,
            text="Светлая",
            command=_on_theme_toggle
        )
        # выставим положение по текущей теме
        self.theme_switch.select() if getattr(self, "ui_theme", "dark") == "light" else self.theme_switch.deselect()
        self.theme_switch.pack(side='left', padx=(0,10), pady=6)
        
        # Название проекта в верхней панели
        ctk.CTkLabel(topbar, text="Название проекта:").pack(side='left', padx=(20,5), pady=6)
        self.project_name_entry = ctk.CTkEntry(topbar, width=180, placeholder_text="Название проекта")
        self.project_name_entry.pack(side='left', padx=(0,10), pady=6)
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=5, pady=5)
        self.project_tab = ttk.Frame(self.notebook)
        self.tables_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.project_tab, text="Проект")
        self.notebook.add(self.tables_tab, text="Таблицы")


    def setup_project_tab(self):
        self.setup_parameters_frame()
        self.setup_add_part_frame()
        self.project_stats = {
            "Кол-во матов": tk.StringVar(value="0"),
            "Себестоимость €/кг": tk.StringVar(value="0.00"),
            "Цена проекта €": tk.StringVar(value="0.00"),
            "Общий черный вес": tk.StringVar(value="0.00"),
        }

        # --- общий контейнер для таблицы деталей + блока результатов ---
        middle_frame = ctk.CTkFrame(self.project_tab)
        middle_frame.pack(fill='both', expand=True, padx=5, pady=5)

        middle_frame.grid_rowconfigure(0, weight=7, minsize=280)  # таблица деталей — приоритет и минимум
        middle_frame.grid_rowconfigure(1, weight=1, minsize=100)  # результаты — ужимаются первыми
        middle_frame.grid_columnconfigure(0, weight=1)

        # Таблица деталей в верхней части
        self.setup_parts_table(parent=middle_frame)

        # Блок результатов в нижней части
        self.setup_results_frame(parent=middle_frame)

    def setup_parameters_frame(self):
        # Общий контейнер секции
        params_frame = ctk.CTkFrame(self.project_tab)
        params_frame.pack(fill='x', padx=8, pady=(6, 8))

        # --- Основная строка: ЛЕВО — ЦЕНТР — ПРАВО ---
        columns_frame = ctk.CTkFrame(params_frame)
        columns_frame.pack(fill='x')

        # ---- ЛЕВАЯ КОЛОНКА (параметры мата)
        left_col = ctk.CTkFrame(columns_frame)
        left_col.pack(side='left', fill='both', expand=True, padx=(0, 10))

        # 🔹 Заголовок блока
        ctk.CTkLabel(left_col, text="Параметры матов", font=("Segoe UI", 12, "bold"))\
            .grid(row=0, column=0, columnspan=4, sticky='w', padx=6, pady=(2, 6))

        # Все элементы сдвигаем на +1 вниз
        ctk.CTkLabel(left_col, text="Артикул мата:").grid(row=1, column=0, sticky='e', padx=5, pady=4)
        self.article_cb = ctk.CTkComboBox(left_col, width=140, values=[])
        self.article_cb.grid(row=1, column=1, sticky='w', padx=5, pady=4)

        ctk.CTkLabel(left_col, text="Кол-во матов:").grid(row=2, column=0, sticky='e', padx=5, pady=4)
        self.mats_count_entry = ctk.CTkEntry(left_col, width=90)
        self.mats_count_entry.insert(0, "0")
        self.mats_count_entry.grid(row=2, column=1, sticky='w', padx=5, pady=4)
        
        
        def _update_mats_count_effective(*_):
            s = (self.mats_count_entry.get() or "").strip().replace(",", ".")
            try:
                base = float(s)
            except Exception:
                base = 0.0
            eff = round(base * 1.15, 1)
            self.mats_count_effective_var.set(f"с запасом: {eff}")

        

        self.manual_mats_var = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(left_col, text="Ввести вручную", variable=self.manual_mats_var,
                        command=self._toggle_manual_mats).grid(row=2, column=2, sticky='w', padx=(8, 0), pady=4)

        ctk.CTkLabel(left_col, text="Длина мата (мм):").grid(row=3, column=0, sticky='e', padx=5, pady=4)
        self.mat_length_entry = ctk.CTkEntry(left_col, width=100); self.mat_length_entry.insert(0, "6100")
        self.mat_length_entry.grid(row=3, column=1, sticky='w', padx=5, pady=4)

        ctk.CTkLabel(left_col, text="Ширина мата (мм):").grid(row=3, column=2, sticky='e', padx=5, pady=4)
        self.mat_width_entry = ctk.CTkEntry(left_col, width=100); self.mat_width_entry.insert(0, "1000")
        self.mat_width_entry.grid(row=3, column=3, sticky='w', padx=5, pady=4)

        ctk.CTkLabel(left_col, text="Режим расчета:").grid(row=4, column=0, sticky='e', padx=5, pady=(6, 4))
        self.calc_mode = tk.StringVar(value="По длине")
        ctk.CTkSegmentedButton(left_col, values=["По длине", "По площади"], variable=self.calc_mode)\
            .grid(row=4, column=1, columnspan=3, sticky='w', padx=5, pady=(6, 4))

        ttk.Separator(left_col, orient='horizontal').grid(row=5, column=0, columnspan=4, sticky='ew', pady=(8, 6))
        left_btns = ctk.CTkFrame(left_col)
        left_btns.grid(row=6, column=0, columnspan=4, sticky='w', padx=5)
        
        ctk.CTkButton(left_btns, text="Рассчитать количество", command=self.cutting_calculate_mats).pack(side='left',padx=(0, 10))
        ctk.CTkButton(left_btns, text="Предпросмотр раскроя", command=self.cutting_preview).pack(side='left')

        # ---- ЦЕНТРАЛЬНАЯ КОЛОНКА (КНОПКИ СВЕРХУ В СТОЛБИК)
        middle_col = ctk.CTkFrame(columns_frame)
        middle_col.pack(side='left', fill='both', expand=True, padx=(0, 10))
        ctk.CTkLabel(middle_col, text="Действия с проектом", font=("Segoe UI", 12, "bold"))\
            .pack(anchor='w', padx=6, pady=(2, 6))

        ctk.CTkButton(middle_col, text="Открыть проект",   command=self.open_project_dialog).pack(fill='x', padx=8, pady=4)
        ctk.CTkButton(middle_col, text="Сохранить проект", command=self.save_project).pack(fill='x', padx=8, pady=4)
        ctk.CTkButton(middle_col, text="Удалить проект",   command=self.clear_project).pack(fill='x', padx=8, pady=4)   
        ctk.CTkButton(middle_col, text="Экспорт в Excel",  command=self.export_to_excel).pack(fill='x', padx=8, pady=4)
        self.update_btn = ctk.CTkButton(
            middle_col,
            text="Обновить расчет",
            command=self.recalculate_project,
            height=44,                         # повыше
            font=("Segoe UI", 13, "bold"),     # жирнее
            fg_color=("#2E8B57", "#2E8B50"),   # ярче (светлая / темная темы)
            hover_color=("#3CB371", "#3CB365"),
            text_color=("white", "white"),
            corner_radius=10,
            border_width=1,
            border_color=("#165EA3", "#0F4E8A"),
        )
        self.update_btn.pack(fill='x', padx=8, pady=(8, 6))  # побольше отступ сверху

        # ---- ПРАВАЯ КОЛОНКА (Материалы и цены) — ширина как у левой
        right_col = ctk.CTkFrame(columns_frame)
        right_col.pack(side='left', fill='both', expand=True)

        materials_frame = ctk.CTkFrame(right_col)
        materials_frame.pack(fill='x', pady=(0, 10))
        ctk.CTkLabel(materials_frame, text="Материалы и цены", font=("Segoe UI", 12, "bold")).pack(anchor='w')

        materials = [
            {"name": "Обрамление", "key": "frame",  "unit": "€/кг",  "default": "0.79"},
            {"name": "Перфоуголок","key": "angle",  "unit": "€/шт",  "default": "2.0"},
            {"name": "Отбойник",   "key": "bumper", "unit": "€/кг",  "default": "0.79"},
            {"name": "Работа",     "key": "work",   "unit": "€/час", "default": "35.0"},
        ]
        self.material_entries = {}
        for mat in materials:
            row = ctk.CTkFrame(materials_frame)
            row.pack(fill='x', pady=2)
            ctk.CTkLabel(row, text=mat["name"] + ":", width=90).pack(side='left')
            if mat["name"] != "Работа":
                combo = ctk.CTkComboBox(row, width=120)
                combo.pack(side='left', padx=5)
                self.material_combos[mat["name"]] = combo
            entry = ctk.CTkEntry(row, width=80); entry.insert(0, mat["default"])
            entry.pack(side='left')
            self.material_entries[mat["key"]] = entry
            ctk.CTkLabel(row, text=mat["unit"]).pack(side='left', padx=5)

        # Коэффициенты
        coeff_frame = ctk.CTkFrame(right_col)
        coeff_frame.pack(fill='x')
        
        # Первая строка коэффициентов
        row1 = ctk.CTkFrame(coeff_frame)
        row1.pack(fill='x', pady=2)
        
        ctk.CTkLabel(row1, text="Коэф. работы:").pack(side='left', padx=(0, 5))
        self.work_coeff = ctk.CTkEntry(row1, width=70); self.work_coeff.insert(0, "1.0")
        self.work_coeff.pack(side='left', padx=(0, 15))
        
        ctk.CTkLabel(row1, text="Цинк (€/кг):").pack(side='left', padx=(0, 5))
        self.zinc_entry = ctk.CTkEntry(row1, width=70); self.zinc_entry.insert(0, "0.6")
        self.zinc_entry.pack(side='left')

        # Вторая строка коэффициентов
        row2 = ctk.CTkFrame(coeff_frame)
        row2.pack(fill='x', pady=2)
        
        ctk.CTkLabel(row2, text="Коэф. обрамления:").pack(side='left', padx=(0, 5))
        self.frame_coef_entry = ctk.CTkEntry(row2, width=70); self.frame_coef_entry.insert(0, "1.05")
        self.frame_coef_entry.pack(side='left', padx=(0, 15))
        
        # Чекбокс "Вычитать отбойник" теперь здесь
        self.subtract_bumper_var = tk.BooleanVar(value=True)
        

    def setup_add_part_frame(self):
        add_frame = ctk.CTkFrame(self.project_tab)
        add_frame.pack(fill='x', padx=5, pady=5)

        # Заголовок
        header_row = ctk.CTkFrame(add_frame)
        header_row.pack(fill='x', pady=(0, 5))
        
        ctk.CTkLabel(header_row, text="Добавить деталь", font=("Segoe UI", 13, "bold")).pack(side='left')

        # Поля ввода - используем grid для точного позиционирования
        input_frame = ctk.CTkFrame(add_frame)
        input_frame.pack(fill='x', pady=(0, 5))
        
        fields = ["Позиция", "Длина (мм)", "Ширина (мм)", "Кол-во", "Перфоугол", "Отбойник", "Работа (ч)", "Радиус (мм)"]
        self.part_entries = {}

        for i, field in enumerate(fields):
            ctk.CTkLabel(input_frame, text=field).grid(row=0, column=i, padx=2, pady=2)
            key = field.split()[0].lower()

            if key == "радиус":
                radius_entry = ctk.CTkEntry(input_frame, width=80)
                radius_entry.grid(row=1, column=i, padx=(2,0), pady=2)
                self.part_entries[key] = radius_entry

                ctk.CTkLabel(input_frame, text="Доля окружн.").grid(row=0, column=i+1, padx=2, pady=2)
                radius_part_cb = ctk.CTkComboBox(input_frame, values=["1.0", "0.75", "0.5", "0.25"], width=80)
                radius_part_cb.set("1.0")
                radius_part_cb.grid(row=1, column=i+1, padx=(0,2), pady=2)
                self.part_entries["доля радиуса"] = radius_part_cb
                
                # Добавляем чекбокс "Вычитать отбойник" после доли окружности
                ctk.CTkLabel(input_frame, text="Вычит. отбойник").grid(row=0, column=i+2, padx=2, pady=2)
                self.subtract_bumper_checkbox = ctk.CTkCheckBox(input_frame, text="", width=20)
                self.subtract_bumper_checkbox.select()  # По умолчанию включено
                self.subtract_bumper_checkbox.grid(row=1, column=i+2, padx=2, pady=2)
            else:
                e = ctk.CTkEntry(input_frame, width=110)
                e.grid(row=1, column=i, padx=2, pady=2)
                self.part_entries[key] = e

        # Кнопки
        btn_frame = ctk.CTkFrame(add_frame)
        btn_frame.pack(fill='x')
        
        ctk.CTkButton(btn_frame, text="Добавить", command=self.add_part).pack(side='left', padx=(0, 5))
        ctk.CTkButton(btn_frame, text="Очистить", command=self.clear_part_fields).pack(side='left', padx=(0, 5))
        ctk.CTkButton(btn_frame, text="Импорт из Excel", command=self.import_parts_from_excel).pack(side='left')
        
    def _toggle_manual_mats(self):
        """Включает/выключает ручное редактирование поля 'Кол-во матов'."""
        if getattr(self, "manual_mats_var", None) and self.manual_mats_var.get():
            self.mats_count_entry.configure(state="normal")
        else:
            self.mats_count_entry.configure(state="readonly")


    

    def setup_parts_table(self, parent=None):
        if parent is None:
            parent = self.project_tab

        table_frame = ctk.CTkFrame(parent)
        # таблица деталей должна быть в верхней строке контейнера middle_frame
        table_frame.grid(row=0, column=0, sticky='nsew')

        # внутри table_frame настроим растяжение
        table_frame.grid_rowconfigure(0, weight=1)   # строка с Treeview
        table_frame.grid_columnconfigure(0, weight=1)

        columns = ["Позиция", "Длина", "Ширина", "Кол-во", "Площадь",
                   "Обрамление", "Перфоугол", "Отбойник", "Работа", "Радиус", "Длина отбойника"]

        self.parts_table = ttk.Treeview(table_frame, columns=columns, show="headings", height=6, selectmode='browse')
        for col in columns:
            self.parts_table.heading(col, text=col)
            self.parts_table.column(col, width=90, anchor='center', stretch=True)

        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.parts_table.yview)
        xscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.parts_table.xview)
        self.parts_table.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        self.parts_table.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        parent.grid_rowconfigure(0, minsize=280)

        # бинды/обработчики как были
        self.parts_table.unbind("<Double-1>")
        self.parts_table.bind("<Double-1>", self.on_parts_table_dblclick)
        
        self._init_parts_context_menu()

    def _get_mat_dimensions_mm(self) -> tuple[int, int]:
        """Берем длину/ширину мата из полей параметров проекта.
        Если полей нет — дефолт 6100×1000 мм."""
        L = 6100
        W = 1000
        try:
            if hasattr(self, "mat_length_entry"):
                L = int(float((self.mat_length_entry.get() or "6100").replace(",", ".")))
            if hasattr(self, "mat_width_entry"):
                W = int(float((self.mat_width_entry.get() or "1000").replace(",", ".")))
        except Exception:
            pass
        return max(L, 0), max(W, 0)

    def _has_manual_mats_input(self) -> bool:
        """True, если пользователь ввел число в «Кол-во матов» (приоритет ручного ввода)."""
        s = (self.mats_count_entry.get() or "").strip().replace(",", ".")
        if not s:
            return False
        import re
        ok = re.fullmatch(r'\d+([.,]\d{0,1})?', s) is not None
        if not ok:
            return False
        try:
            return float(s.replace(",", ".")) > 0
        except Exception:
            return False
        
    def on_parts_table_dblclick(self, event=None):
        # если уже идет открытие — не даем запуститься второй раз
        if getattr(self, "_part_editor_opening", False):
            if self._part_editor_win and self._part_editor_win.winfo_exists():
                try:
                    self._part_editor_win.lift(); self._part_editor_win.focus_force()
                except Exception:
                    pass
            return "break"  # ← ВОЗВРАТ ДОЛЖЕН БЫТЬ БЕЗУСЛОВНЫМ

        # если окно уже открыто — просто поднимем
        if self._part_editor_win and self._part_editor_win.winfo_exists():
            try:
                self._part_editor_win.lift(); self._part_editor_win.focus_force()
            except Exception:
                pass
            return "break"

        self._part_editor_opening = True
        try:
            row_iid = self.parts_table.identify_row(event.y) if event else None
            if not row_iid:
                sel = self.parts_table.selection()
                row_iid = sel[0] if sel else None
            if not row_iid or row_iid not in self._tv_to_part:
                messagebox.showwarning("Внимание", "Выберите строку для редактирования")
                return "break"

            self._open_part_editor(self._tv_to_part[row_iid], row_iid)
        finally:
            # если окно так и не создалось — опустим флаг
            if not (self._part_editor_win and self._part_editor_win.winfo_exists()):
                self._part_editor_opening = False

        return "break"


    def _open_part_editor(self, part: ProjectPart, row_iid: str):
        # Если окно уже открыто — просто поднять
        if self._part_editor_win and self._part_editor_win.winfo_exists():
            try:
                self._part_editor_win.lift(); self._part_editor_win.focus_force()
            except Exception:
                pass
            return

        win = ctk.CTkToplevel(self.root)
        self._part_editor_win = win
        win.title(f"Редактирование: {part.position}")
        win.transient(self.root)
        win.grab_set()

        # Геометрия/поведение
        win.geometry("760x520+120+120")
        win.minsize(700, 460)
        win.resizable(True, True)
        win.grid_columnconfigure(1, weight=1, minsize=320)

        def _on_close():
            try:
                win.grab_release()
            except Exception:
                pass
            self._part_editor_opening = False
            self._part_editor_win = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)

        # ---- 1) Подготовка значений к показу
        labels = [
            ("Позиция", part.position),
            ("Длина (мм)", part.length),
            ("Ширина (мм)", part.width),
            ("Кол-во", part.quantity),
            ("Перфоугол", part.angle),
            ("Отбойник", part.bumper),
            ("Работа (ч)", part.work),
            ("Радиус (мм)", part.radius),
        ]
        int_mm_fields = {"Длина (мм)", "Ширина (мм)", "Радиус (мм)", "Отбойник", "Кол-во"}
        normalized = []
        for lbl, val in labels:
            if lbl in int_mm_fields:
                try:
                    val = int(round(float(val or 0)))
                except Exception:
                    val = 0
            elif lbl == "Работа (ч)":
                try:
                    val = f"{float(val or 0):.1f}"
                except Exception:
                    val = "0.0"
            normalized.append((lbl, val))
        labels = normalized

        # ---- 2) Поля ввода
        entries: Dict[str, ctk.CTkEntry] = {}
        for i, (lbl, val) in enumerate(labels):
            ctk.CTkLabel(win, text=lbl).grid(row=i, column=0, padx=8, pady=3, sticky='e')
            e = ctk.CTkEntry(win)
            e.insert(0, str(val))
            e.grid(row=i, column=1, padx=8, pady=3, sticky='we')
            entries[lbl] = e

        # Валидация для целочисленных миллиметров/количества
        vcmd_digits = (win.register(lambda s: (s.isdigit() or s == "")), "%P")
        for lbl, entry in entries.items():
            if lbl in int_mm_fields:
                entry.configure(validate="key", validatecommand=vcmd_digits)

        # ---- 3) Доля окружности
        row_base = len(labels)
        ctk.CTkLabel(win, text="Доля окружности").grid(row=row_base, column=0, padx=8, pady=(8, 4), sticky="e")
        radius_part_cb = ctk.CTkComboBox(win, values=["1.0", "0.75", "0.5", "0.25"], width=100)
        try:
            radius_part_cb.set(str(float(getattr(part, "radius_part", 1.0))))
        except Exception:
            radius_part_cb.set("1.0")
        radius_part_cb.grid(row=row_base, column=1, padx=8, pady=(8, 4), sticky="w")
        entries["Доля окружности"] = radius_part_cb  # храним вместе с остальными
        
        row_base += 1
        subtract_bumper_var = tk.BooleanVar(value=getattr(part, "subtract_bumper", True))
        subtract_bumper_cb = ctk.CTkCheckBox(win, text="Вычитать отбойник", variable=subtract_bumper_var)
        subtract_bumper_cb.grid(row=row_base, column=0, columnspan=2, padx=8, pady=(4, 8), sticky="w")
        entries["Вычитать отбойник"] = subtract_bumper_var

        # ---- 4) Кнопки
        btn_frame = ctk.CTkFrame(win)
        btn_frame.grid(row=row_base+1, column=0, columnspan=2, pady=10, sticky='e')

        ctk.CTkButton(
            btn_frame,
            text="Сохранить",
            command=lambda: self._save_part_changes(win, entries, part, row_iid)
        ).pack(side='left', padx=5)
        ctk.CTkButton(btn_frame, text="Отмена", command=_on_close).pack(side='left', padx=5)

        # Клавиши
        win.bind("<Return>", lambda e: self._save_part_changes(win, entries, part, row_iid))
        win.bind("<Escape>", lambda e: _on_close())

        # Фокус
        entries["Позиция"].focus_set()


    def _save_part_changes(self, win, entries, part, row_iid):
        """Сохраняет изменения в объект 'part' и МГНОВЕННО обновляет выбранную строку Treeview."""
        try:
            # --- безопасные парсеры
            def _to_int_mm(s: str) -> int:
                try:
                    return int(round(float((s or "0").replace(",", "."))))
                except Exception:
                    return 0

            def _to_float1(s: str) -> float:
                try:
                    return round(float((s or "0").replace(",", ".")), 1)
                except Exception:
                    return 0.0

            # --- читаем UI
            subtract_bumper = entries["Вычитать отбойник"].get()
            part.position = entries["Позиция"].get().strip()
            part.length   = _to_int_mm(entries["Длина (мм)"].get())
            part.width    = _to_int_mm(entries["Ширина (мм)"].get())
            part.quantity = _to_int_mm(entries["Кол-во"].get() or "1")
            part.angle    = entries["Перфоугол"].get().strip()
            part.bumper   = _to_int_mm(entries["Отбойник"].get())
            part.work     = _to_float1(entries["Работа (ч)"].get())
            part.radius   = _to_int_mm(entries["Радиус (мм)"].get())
            part.subtract_bumper = subtract_bumper  # Сохраняем состояние чекбокса

            # доля окружности (из ComboBox)
            rp_s = (entries["Доля окружности"].get() or "1.0").replace(",", ".")
            try:
                part.radius_part = float(rp_s)
            except Exception:
                part.radius_part = 1.0

            # --- пересчет производных полей
            # длина отбойника: прямая часть + дуга от радиуса*доля
            try:
                bumper_val = float(part.bumper)
            except Exception:
                bumper_val = 0.0                

            # эта функция у тебя есть; возвращает (linear_mm, total_bumper_mm)
            linear_mm, total_bumper_mm = self.calculate_bumper_lengths(
                bumper_val, part.radius, part.radius_part
            )
            part.bumper_length = round(total_bumper_mm / 1000, 2)  # м
            
            

            # периметр с учетом индивидуального вычитания отбойника
            per_mm = (2 * part.width + part.length) if part.width < 1000 else (2 * part.width)
            if part.bumper and part.subtract_bumper:  # Используем индивидуальное значение
                per_mm -= linear_mm
            part.perimeter = max(0, round(per_mm / 1000, 2))

            # площадь детали (м²) — для отображения
            area_m2 = (part.length * part.width) / 1_000_000

            # --- МГНОВЕННОЕ обновление текущей строки, без полной перерисовки таблицы
            if self.parts_table.exists(row_iid):
                self.parts_table.item(row_iid, values=(
                    part.position,
                    f"{part.length:.0f}",
                    f"{part.width:.0f}",
                    part.quantity,
                    f"{area_m2:.3f}".replace('.', ','),
                    f"{part.perimeter:.2f}".replace('.', ','),
                    part.angle,
                    part.bumper,
                    f"{part.work:.1f}".replace('.', ','),
                    f"{part.radius:.0f}" if part.radius else "",
                    f"{part.bumper_length:.2f}"
                ))
                # мэппинг iid->part сохраняем (part — тот же объект)
                self._tv_to_part[row_iid] = part

            # --- пересчет итогов проекта
            self.recalculate_project()

            # Закрыть окно редактора
            try:
                win.grab_release()
            except Exception:
                pass
            self._part_editor_opening = False
            self._part_editor_win = None
            win.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Некорректные данные: {e}")

    def _calc_mat_length_used(self, rects) -> float:
        """Сколько длины реально занято на мате: max(x + length)."""
        if not rects:
            return 0.0
        return max((d["x"] + d["length"]) for d in rects)

    def _calc_mat_area_used(self, rects) -> float:
        """Сумма площадей деталей на мате."""
        return sum((d["length"] * d["width"]) for d in rects)
        
    def _mats_count_for_export(self) -> float:
        """Количество матов для Excel: БЕЗ +15%, 1 знак после запятой."""
        # 1) напрямую из поля
        try:
            s = (self.mats_count_entry.get() or "").strip().replace(",", ".")
            if s != "":
                return round(float(s), 1)
        except Exception:
            pass
        # 2) иначе — как считает функция выше (тоже без +15%)
        try:
            return round(float(getattr(self, "calculate_mats_count", lambda: 0.0)()), 1)
        except Exception:
            return 0.0
    def _fraction_from_ratio(self, ratio: float) -> float:
        """
        Применяет критерий:
          - ratio >= 0.9  -> 1.0
          - ratio < 0.1   -> 0.0
          - иначе округление до 0.1
        """
        if ratio >= 0.9:
            return 1.0
        if ratio < 0.1:
            return 0.0
        return round(ratio, 1)
                

        def save_changes():
            try:
                part.position = entries["Позиция"].get().strip()
                part.length   = _to_int_mm(entries["Длина (мм)"].get())
                part.width    = _to_int_mm(entries["Ширина (мм)"].get())
                part.quantity = _to_int_mm(entries["Кол-во"].get() or "1")
                part.angle    = entries["Перфоугол"].get().strip()
                part.bumper   = _to_int_mm(entries["Отбойник"].get())
                part.work     = _to_float1(entries["Работа (ч)"].get())
                part.radius   = _to_int_mm(entries["Радиус (мм)"].get())

                # доля окружности
                rp_s = (entries["Доля окружности"].get() or "1.0").replace(",", ".")
                part.radius_part = float(rp_s)

                # Пересчеты
                try:
                    bumper_val = float(part.bumper)
                except Exception:
                    bumper_val = 0.0

                linear_mm, total_bumper_mm = self.calculate_bumper_lengths(bumper_val, part.radius, part.radius_part)
                part.bumper_length = round(total_bumper_mm / 1000, 2)

                per_mm = (2 * part.width + part.length) if part.width < 1000 else (2 * part.width)
                if part.bumper and self.subtract_bumper_var.get():
                    per_mm -= linear_mm
                part.perimeter = max(0, round(per_mm / 1000, 2))

                # Обновляем строку в таблице
                area = (part.length * part.width) / 1_000_000
                self.parts_table.item(row_iid, values=(
                    part.position,
                    f"{part.length:.0f}", f"{part.width:.0f}", part.quantity,
                    f"{area:.3f}".replace('.', ','), f"{part.perimeter:.2f}".replace('.', ','),
                    part.angle, part.bumper,
                    f"{part.work:.1f}".replace('.', ','), f"{part.radius:.0f}" if part.radius else "",
                    f"{part.bumper_length:.2f}" if part.bumper_length > 0 else "0.00"
                ))

                self.recalculate_project()
                _on_close()
            except Exception as e:
                messagebox.showerror("Ошибка", f"Некорректные данные: {e}")

        # ---- 5) Кнопки — НА СЛЕДУЮЩЕЙ строке ----
        row_buttons = row_base + 1
        btns = ttk.Frame(win)
        btns.grid(row=row_buttons, column=0, columnspan=2, padx=8, pady=(10, 10), sticky="e")
        ttk.Button(btns, text="Сохранить", command=save_changes).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btns, text="Отмена", command=_on_close).grid(row=0, column=1)
        win.bind("<Return>", lambda e: save_changes())
        win.bind("<Escape>", lambda e: _on_close())

    def remove_selected_part(self):
        sel = self.parts_table.selection()
        if not sel:
            return
        iid = sel[0]
        part = self._tv_to_part.get(iid)
        if not part:
            return
        # удалить из списка
        try:
            self.project_parts.remove(part)
        except Exception:
            pass
        # удалить из дерева и маппинга
        self.parts_table.delete(iid)
        self._tv_to_part.pop(iid, None)
        self.recalculate_project()

    def _init_parts_context_menu(self):
        self._parts_menu = tk.Menu(self.root, tearoff=0)
        self._parts_menu.add_command(label="Редактировать", command=lambda: self.on_parts_table_dblclick())
        self._parts_menu.add_command(label="Удалить", command=self.remove_selected_part)

        def _show_menu(event):
            iid = self.parts_table.identify_row(event.y)
            if iid:
                self.parts_table.selection_set(iid)
                try:
                    self._parts_menu.tk_popup(event.x_root, event.y_root)
                finally:
                    self._parts_menu.grab_release()

        self.parts_table.bind("<Button-3>", _show_menu)  # правый клик
       

    def setup_results_frame(self, parent=None):
        if parent is None:
            parent = self.project_tab

        results_frame = ctk.CTkFrame(parent)
        results_frame.grid(row=1, column=0, sticky='nsew')

        # колонки охотно сжимаются; таблице деталей выше мы уже задали приоритет
        for col in (0, 1, 2):
            results_frame.grid_columnconfigure(col, weight=1, uniform="res", minsize=80)
        results_frame.grid_rowconfigure(0, weight=1, minsize=80)

        # --- СЕБЕСТОИМОСТЬ ---
        cost_frame = ctk.CTkFrame(results_frame)
        cost_frame.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        ctk.CTkLabel(cost_frame, text="Себестоимость", font=("Segoe UI", 13, "bold")).pack(anchor='w', padx=6, pady=(4, 2))
        self.cost_table = ttk.Treeview(cost_frame, columns=("comp", "value"), show="headings", height=5)
        self.cost_table.heading("comp", text="Компонент")
        self.cost_table.heading("value", text="Стоимость")
        self.cost_table.pack(fill='both', expand=True, padx=6, pady=6)

        # --- ВЕС ---
        weight_frame = ctk.CTkFrame(results_frame)
        weight_frame.grid(row=0, column=1, sticky='nsew', padx=5, pady=5)
        ctk.CTkLabel(weight_frame, text="Вес", font=("Segoe UI", 13, "bold")).pack(anchor='w', padx=6, pady=(4, 2))
        self.weight_table = ttk.Treeview(weight_frame, columns=("comp", "value"), show="headings", height=5)
        self.weight_table.heading("comp", text="Материал")
        self.weight_table.heading("value", text="Вес (кг)")
        self.weight_table.pack(fill='both', expand=True, padx=6, pady=6)

        # --- СТАТИСТИКА (со скроллом, чтобы ничего не обрезалось) ---
        stats_outer = ctk.CTkFrame(results_frame)
        stats_outer.grid(row=0, column=2, sticky='nsew', padx=5, pady=5)
        stats_outer.grid_rowconfigure(1, weight=1)  # зона со скроллом тянется
        stats_outer.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(stats_outer, text="Статистика", font=("Segoe UI", 13, "bold"))\
            .grid(row=0, column=0, sticky='w', padx=6, pady=(4, 2))

        # прокручиваемая область
        stats_panel = ctk.CTkScrollableFrame(stats_outer)
        stats_panel.grid(row=1, column=0, sticky='nsew', padx=1, pady=(0, 1))

        def stat_card(parent, title, var):
            f = ctk.CTkFrame(parent)
            f.pack(fill='x', padx=6, pady=4)
            ctk.CTkLabel(f, text=title, font=("Segoe UI", 11)).pack(anchor='w', padx=6, pady=(2, 0))
            ctk.CTkLabel(f, textvariable=var, font=("Segoe UI", 16, "bold")).pack(anchor='w', padx=6, pady=(0, 4))
            return f

        for key in ("Кол-во матов", "Себестоимость €/кг", "Цена проекта €", "Общий черный вес"):
            stat_card(stats_panel, key, self.project_stats[key])

    

    def setup_tables_tab(self):
        tab = ttk.Notebook(self.tables_tab)   # оставляем ttk.Notebook
        tab.pack(fill='both', expand=True, padx=5, pady=5)

        # Вкладка "Данные матов"
        frame_mats = ctk.CTkFrame(tab)
        tab.add(frame_mats, text="Данные матов")

        # Убрали "Полоса"
        columns = ["Артикул", "Ячейка", "Цена мата", "Вес мата", "Цена м²", "Вес м²", "Длина мата"]
        self.mat_table = ttk.Treeview(frame_mats, columns=columns, show="headings", height=10)
        for c in columns:
            self.mat_table.heading(c, text=c)
            self.mat_table.column(c, width=110, anchor='center', stretch=True)

        yscroll = ttk.Scrollbar(frame_mats, orient="vertical", command=self.mat_table.yview)
        xscroll = ttk.Scrollbar(frame_mats, orient="horizontal", command=self.mat_table.xview)
        self.mat_table.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.mat_table.grid(row=0, column=0, sticky='nsew')
        self.mat_table.bind('<Double-1>', self.edit_mat_cell)
        yscroll.grid(row=0, column=1, sticky='ns')
        xscroll.grid(row=1, column=0, sticky='ew')
        frame_mats.grid_rowconfigure(0, weight=1)
        frame_mats.grid_columnconfigure(0, weight=1)

        # Остальные таблицы:
        self._setup_material_table(tab, "Обрамление", ["Размер", "Вес 1м"])
        self._setup_material_table(tab, "Перфоуголок", ["Размер", "Вес 1м"])
        self._setup_material_table(tab, "Отбойник", ["Размер", "Вес 1м"])

        # Панель кнопок импорта/сохранения
        button_frame = ctk.CTkFrame(self.tables_tab)
        button_frame.pack(side='bottom', fill='x', pady=8)
        ctk.CTkButton(button_frame, text="Импорт данных матов", command=self.import_mat_data).pack(side='left', padx=5)
        ctk.CTkButton(button_frame, text="Импорт весов обрамления", command=lambda: self.import_weight_data("Обрамление")).pack(side='left', padx=5)
        ctk.CTkButton(button_frame, text="Импорт весов перфоуголка", command=lambda: self.import_weight_data("Перфоуголок")).pack(side='left', padx=5)
        ctk.CTkButton(button_frame, text="Импорт весов отбойника", command=lambda: self.import_weight_data("Отбойник")).pack(side='left', padx=5)

        



    def _setup_material_table(self, notebook: ttk.Notebook, title: str, columns: List[str]):
        frame = ttk.Frame(notebook)
        notebook.add(frame, text=title)
        table = ttk.Treeview(frame, columns=columns, show="headings", height=10)
        for c in columns:
            table.heading(c, text=c)
            table.column(c, width=110, anchor='center', stretch=True)
        ys = ttk.Scrollbar(frame, orient='vertical', command=table.yview)
        xs = ttk.Scrollbar(frame, orient='horizontal', command=table.xview)
        table.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        table.grid(row=0, column=0, sticky='nsew')
        ys.grid(row=0, column=1, sticky='ns')
        xs.grid(row=1, column=0, sticky='ew')
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        setattr(self, f"{title}_table", table)

        # ⬇️ ВАЖНО: включаем инлайн-редактирование для таблицы «Цены»
        if title == "Цены":
            table.bind('<Double-1>', lambda e, t=table: self.edit_price_cell(e, t))
    
        
    def edit_mat_cell(self, event):
        region = self.mat_table.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = self.mat_table.identify_column(event.x)
        if column != "#4":  # Только редактируем "Цена мата"
            return

        row_id = self.mat_table.identify_row(event.y)
        if not row_id:
            return

        x, y, w, h = self.mat_table.bbox(row_id, column)
        current_values = list(self.mat_table.item(row_id, "values"))
        col_index = int(column[1:]) - 1  # 3 для "Цена мата"
        current_value = current_values[col_index]

        entry = ttk.Entry(self.mat_table)
        entry.insert(0, str(current_value))
        entry.select_range(0, tk.END)
        entry.focus()
        entry.place(x=x, y=y, width=w, height=h)

        def commit():
            try:
                new_price = float(str(entry.get()).replace(",", "."))
                new_price_m2 = round(new_price / 6.1, 2)
            except ValueError:
                messagebox.showerror("Ошибка", "Введите корректное число")
                return

            # Обновляем видимые данные
            current_values[3] = f"{new_price:.2f}"
            current_values[5] = f"{new_price_m2:.2f}"
            self.mat_table.item(row_id, values=current_values)
            entry.destroy()

            # Обновляем данные в self.mat_data
            art = current_values[0]
            for m in self.mat_data:
                if m.get("Артикул") == art:
                    m["Цена мата"] = new_price
                    m["Цена м²"] = new_price_m2
                    break
            self._refresh_article_combobox()

            self.save_config()
            self.status_var.set("Цена мата и цена м² обновлены")

        entry.bind("<Return>", lambda e: commit())
        entry.bind("<FocusOut>", lambda e: entry.destroy())

    def _cut_build_dense_columns(self, items, mat_width, kerf=5):
        remaining = [dict(d) for d in items]
        remaining.sort(key=lambda d: (d["length"], d["width"]), reverse=True)
        columns = []
        for part in remaining:
            best_col, best_delta = None, None
            for col in columns:
                need_h = (col["height_used"] == 0) and part["width"] or (kerf + part["width"])
                if col["height_used"] + need_h > mat_width:
                    continue
                new_wx = max(col["width_x"], part["length"])
                delta = new_wx - col["width_x"]
                if (best_delta is None) or (delta < best_delta):
                    best_delta, best_col = delta, col
            if best_col is None:
                columns.append({"width_x": part["length"], "height_used": part["width"], "items": [part]})
            else:
                best_col["width_x"] = max(best_col["width_x"], part["length"])
                best_col["items"].append(part)
                best_col["height_used"] += (kerf + part["width"])
        for col in columns:
            col["items"].sort(key=lambda d: d["width"], reverse=True)
        return columns

    def _cut_place_columns_into_mats(self, columns, mat_length, mat_width, kerf=5):
        cols = sorted(columns, key=lambda c: c["width_x"], reverse=True)
        mats = []
        while cols:
            x = 0
            bin_res = []
            i = 0
            while i < len(cols):
                col = cols[i]
                need_x = col["width_x"] if x == 0 else (kerf + col["width_x"])
                if x + need_x <= mat_length:
                    y = 0
                    for j, d in enumerate(col["items"]):
                        if j > 0:
                            y += kerf
                        bin_res.append({
                            "position": d["position"],
                            "length": d["length"],
                            "width": d["width"],
                            "x": x if need_x == col["width_x"] else x + kerf,
                            "y": y
                        })
                        y += d["width"]
                    x += need_x
                    cols.pop(i)
                    continue
                i += 1
            if not bin_res and cols:
                col = cols.pop(0)
                y = 0
                for j, d in enumerate(col["items"]):
                    if j > 0:
                        y += kerf
                    bin_res.append({"position": d["position"], "length": d["length"], "width": d["width"], "x": 0, "y": y})
                    y += d["width"]
            mats.append(bin_res)
        return mats

    def _current_parts_as_items(self):
        """Разворачиваем детали проекта в список прямоугольников (повторяем по quantity)."""
        items = []
        for p in self.project_parts:
            try:
                l = int(p.length); w = int(p.width); q = int(p.quantity)
            except Exception:
                continue
            for _ in range(max(q, 0)):
                items.append({"position": p.position, "length": l, "width": w})
        return items
        
    def cutting_calculate_mats(self):
        # Явный приоритет ручного ввода
        if getattr(self, "manual_mats_var", None) and self.manual_mats_var.get():
            messagebox.showinfo("Ручной режим", "Включен ручной ввод количества матов.")
            return

        L, W = self._get_mat_dimensions_mm()
        if L <= 0 or W <= 0:
            messagebox.showwarning("Параметры мата", "Неверные размеры мата.")
            return

        all_items = self._current_parts_as_items()
        items = [d for d in all_items if d["length"] <= L and d["width"] <= W]
        skipped = len(all_items) - len(items)
        if not items:
            messagebox.showwarning("Нет подходящих деталей", "Ни одна деталь не помещается в заданный мат.")
            return

        columns = self._cut_build_dense_columns(items, W, kerf=5)
        mats = self._cut_place_columns_into_mats(columns, L, W, kerf=5)
        self._cutting_results = mats

        # режим: По длине / По площади
        mode = (self.calc_mode.get() if hasattr(self, "calc_mode") else "По длине")

        fractions = []
        if mode == "По площади":
            total_area = float(L) * float(W)
            for rects in mats:
                used_area = self._calc_mat_area_used(rects)
                ratio = (used_area / total_area) if total_area > 0 else 0.0
                fractions.append(self._fraction_from_ratio(ratio))
        else:
            for rects in mats:
                used_len = self._calc_mat_length_used(rects)
                ratio = (used_len / float(L)) if L > 0 else 0.0
                fractions.append(self._fraction_from_ratio(ratio))

        total_frac = round(sum(fractions), 1)

        # записать и пересчитать
        self.mats_count_entry.configure(state="normal")  # временно открыть для записи
        self.mats_count_entry.delete(0, "end")
        self.mats_count_entry.insert(0, f"{total_frac:.1f}")
        self._toggle_manual_mats()  # вернуть состояние

        self.recalculate_project()

        msg = f"Листов (раскладок): {len(mats)}\nРежим: {mode}\nСуммарно матов: {total_frac:.1f}"
        if skipped > 0:
            msg += f"\nПредупреждение: {skipped} дет. не поместились по размерам и исключены."
        messagebox.showinfo("Раскрой завершен", msg)
        


    def cutting_preview(self):
        if not self._cutting_results:
            messagebox.showwarning("Нет данных", "Сначала выполните расчет раскроя.")
            return

        mats = self._cutting_results
        L, W = self._get_mat_dimensions_mm()

        win = tk.Toplevel(self.root)
        win.title("Предпросмотр раскроя")
        win.geometry("1000x720")
        win.grab_set()

        # состояние просмотра
        state = {"index": 0}

        canvas_frame = ctk.CTkFrame(win)
        canvas_frame.pack(fill="both", expand=True, padx=8, pady=8)

        btns = ctk.CTkFrame(win)
        btns.pack(fill="x", padx=8, pady=(0,8))

        fig, ax = plt.subplots(figsize=(11.7, 8.27))  # A4-ish
        canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill="both", expand=True)

        def draw(idx: int):
            ax.clear()
            ax.set_xlim(0, L); ax.set_ylim(0, W)
            ax.set_aspect('equal', adjustable='box')
            ax.axis('off')
            ax.add_patch(Rectangle((0,0), L, W, linewidth=2, edgecolor='black', facecolor='none'))
            ax.set_title(f"Мат #{idx+1} / {len(mats)}", fontsize=12)
            for d in mats[idx]:
                ax.add_patch(Rectangle((d["x"], d["y"]), d["length"], d["width"], linewidth=1, edgecolor='black', facecolor='none'))
                ax.text(d["x"] + d["length"]/2, d["y"] + d["width"]/2,
                        f'{d["position"]}\n{d["length"]}×{d["width"]}',
                        ha='center', va='center', fontsize=8)
            fig.tight_layout(pad=0.5)
            canvas.draw_idle()

        def prev_():
            state["index"] = (state["index"] - 1) % len(mats)
            draw(state["index"])

        def next_():
            state["index"] = (state["index"] + 1) % len(mats)
            draw(state["index"])

        ctk.CTkButton(btns, text="←", width=80, command=prev_).pack(side="left", padx=4, pady=4)
        ctk.CTkButton(btns, text="→", width=80, command=next_).pack(side="left", padx=4, pady=4)

        draw(0)


    def update_material_prices(self, table: ttk.Treeview):
        """Перечитывает цены из таблицы «Цены» в self.material_prices"""
        rows = []
        for iid in table.get_children():
            values = table.item(iid, "values")
            if not values:
                continue
            name = str(values[0])
            try:
                price = float(str(values[1]).replace(",", "."))
            except Exception:
                price = 0.0
            unit = str(values[2]) if len(values) > 2 else ""
            rows.append((name, price, unit))
        if rows:
            self.material_prices = rows


    def setup_status_bar(self):
        self.status_var = tk.StringVar(value="Готов к работе")
        status_bar = ctk.CTkLabel(self.root, textvariable=self.status_var, anchor='w')
        status_bar.pack(side='bottom', fill='x')

    def setup_shortcuts(self):
        pass  # при необходимости горячие клавиши
        
    def _apply_theme_and_save(self, mode: str):
        try:
            ctk.set_appearance_mode(mode)
            self.ui_theme = mode
            self.save_config()
            # подсказка в статус-бар
            if hasattr(self, "status_var"):
                self.status_var.set(f"Тема: {mode}")
        except Exception as e:
            logging.error(f"Theme switch error: {e}")
            messagebox.showerror("Ошибка", f"Не удалось применить тему: {e}")

    # =============================
    # Загрузка/сохранение данных и таблиц
    # =============================
    def get_default_material_prices(self):
        return [("обрамление", 0.79, "евро/кг"), ("перфоугол", 2.0, "евро/кг"), ("отбойник", 0.79, "евро/кг"), ("работа", 35.0, "евро/ч")]

    def get_default_frame_weights(self):
        return [("30*2", 0.7), ("30*3", 0.5), ("4*3", 1.95), ("25*3", 0.9), ("35*3", 0.95), ("40*5", 1.5), ("60*5", 1.56)]

    def get_default_angle_weights(self):
        return [("30*3", 0.7), ("25*3", 0.5), ("50*5", 1.95), ("40*3", 0.9), ("30*4", 0.95), ("85*5", 1.5), ("40*5", 1.56), ("20*3", 0.4)]

    def get_default_bumper_weights(self):
        return [("50*4", 1.58), ("40*3", 0.948), ("25*5", 0.9875), ("30*2.5", 0.5925), ("50*5", 1.975), ("100*5", 3.95), ("90*4", 2.844), ("70*5", 2.765), ("60*3", 1.422)]

    def load_default_values(self):
        self.mat_data = []
        self.material_prices = self.get_default_material_prices()
        self.frame_weights = self.get_default_frame_weights()
        self.angle_weights = self.get_default_angle_weights()
        self.bumper_weights = self.get_default_bumper_weights()
        self.populate_material_combos()
        self.refresh_all_tables()

    def load_all_data(self):
        try:
            default_docs = os.path.expanduser("~/Documents")
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)                    
                    self.mat_data = cfg.get("mat_data", [])
                    self.material_prices = cfg.get("material_prices", self.get_default_material_prices())
                    self.frame_weights = cfg.get("frame_weights", self.get_default_frame_weights())
                    self.angle_weights = cfg.get("angle_weights", self.get_default_angle_weights())
                    self.bumper_weights = cfg.get("bumper_weights", self.get_default_bumper_weights())
                    self._last_article = cfg.get("last_article")
                    self.last_export_dir = cfg.get("last_export_dir", default_docs)
                    self.ui_theme = cfg.get(THEME_KEY, "dark")
                    try:
                        ctk.set_appearance_mode(self.ui_theme)
                    except Exception:
                        ctk.set_appearance_mode("dark")
                        self.ui_theme = "dark"
             # >>> НОРМАЛИЗАЦИЯ ПОСЛЕ ЗАГРУЗКИ КОНФИГА
            for m in self.mat_data:
                try:
                    price = float(m.get("Цена мата", 0) or 0)
                    weight = float(m.get("Вес мата", 0) or 0)
                except Exception:
                    price = 0.0; weight = 0.0
                try:
                    m["Длина мата"] = float(m.get("Длина мата", 6100) or 6100)
                except Exception:
                    m["Длина мата"] = 6100.0
                if price > 0:
                    m["Цена м²"] = round(price / 6.1, 2)
                if weight > 0:
                    m["Вес м²"]  = round(weight / 6.1, 3)
                self.save_config()    
            # <<< НОРМАЛИЗАЦИЯ       
            # если не было в конфиге — подгрузим дефолт/файлы
            if not self.material_prices:
                self.material_prices = self.get_default_material_prices()
            if not self.frame_weights:
                self.frame_weights = self.get_default_frame_weights()
            if not self.angle_weights:
                self.angle_weights = self.get_default_angle_weights()
            if not self.bumper_weights:
                self.bumper_weights = self.get_default_bumper_weights()
            self.populate_material_combos()
        except Exception as e:
            logging.error(f"Ошибка загрузки данных: {e}")
            self.load_default_values()

    def save_config(self):
        try:
            cfg = {
                "mat_data": self.mat_data,
                "material_prices": {
                    "frame": self.material_entries["frame"].get() if "frame" in self.material_entries else "",
                    "angle": self.material_entries["angle"].get() if "angle" in self.material_entries else "",
                    "bumper": self.material_entries["bumper"].get() if "bumper" in self.material_entries else "",
                    "work": self.material_entries["work"].get() if "work" in self.material_entries else ""
                },
              
                "frame_weights": self.frame_weights,
                "angle_weights": self.angle_weights,
                "bumper_weights": self.bumper_weights,
                "last_project": self.current_project_path,
                "last_article": self.article_cb.get() if hasattr(self, 'article_cb') else self._last_article,
                "last_export_dir": self.last_export_dir or os.path.expanduser('~/Documents'),
                THEME_KEY: getattr(self, "ui_theme", "dark"),
            }
            
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.error(f"Ошибка сохранения конфига: {e}")
            
    def refresh_mat_table(self):
        """Перерисовать таблицу 'Данные матов' в точном соответствии с текущими колонками Treeview."""
        try:
            self.mat_table.delete(*self.mat_table.get_children())
        except Exception:
            return

        # Текущий порядок и состав колонок берем из самого Treeview
        cols = list(self.mat_table["columns"])

        for m in getattr(self, "mat_data", []):
            row = []
            for c in cols:
                v = m.get(c, "")

                # Немного форматирования чисел под человекочитаемый вид
                if c in ("Цена мата", "Цена м²"):
                    try:
                        v = float(str(v).replace(",", "."))
                        v = f"{v:.2f}".replace(".", ",")
                    except Exception:
                        pass
                elif c in ("Вес мата", "Вес м²"):
                    try:
                        v = float(str(v).replace(",", "."))
                        v = f"{v:.3f}".replace(".", ",")
                    except Exception:
                        pass
                elif c == "Длина мата":
                    try:
                        v = int(float(str(v).replace(",", ".")))
                    except Exception:
                        pass

                row.append(v)

            self.mat_table.insert("", "end", values=tuple(row))



    def refresh_all_tables(self):
        self.refresh_mat_table()
       
        self.refresh_material_table("Обрамление")
        self.refresh_material_table("Перфоуголок")
        self.refresh_material_table("Отбойник")

     

    def refresh_material_table(self, name: str):
        table = getattr(self, f"{name}_table", None)
        if not table:
            return
        table.delete(*table.get_children())
        data: List[Tuple] = []
        if name == "Цены":
            data = self.material_prices
        elif name == "Обрамление":
            data = self.frame_weights
        elif name == "Перфоуголок":
            data = self.angle_weights
        elif name == "Отбойник":
            data = self.bumper_weights
        for row in data:
            table.insert("", "end", values=row)

    def populate_material_combos(self):
        try:
            if self.material_combos.get("Обрамление"):
                self.material_combos["Обрамление"].configure(values=[str(x[0]) for x in self.frame_weights])
                if self.frame_weights:
                    self.material_combos["Обрамление"].set(self.frame_weights[0][0])
            if self.material_combos.get("Перфоуголок"):
                self.material_combos["Перфоуголок"].configure(values=[str(x[0]) for x in self.angle_weights])
                if self.angle_weights:
                    self.material_combos["Перфоуголок"].set(self.angle_weights[0][0])
            if self.material_combos.get("Отбойник"):
                self.material_combos["Отбойник"].configure(values=[str(x[0]) for x in self.bumper_weights])
                if self.bumper_weights:
                    self.material_combos["Отбойник"].set(self.bumper_weights[0][0])
        except Exception as e:
            logging.error(f"populate_material_combos: {e}")

    def _refresh_article_combobox(self, prefer: str = None):
        try:
            def _pos(x):
                try:
                    return float(x) > 0
                except Exception:
                    return False

            # Все артикулы из справочника (не пустые)
            all_articles = []
            for m in self.mat_data:
                art = str(m.get("Артикул", "")).strip()
                if art:
                    all_articles.append(art)

            # Мягкий фильтр: берем те, где есть хоть какая-то числовая информация > 0
            filtered = []
            for m in self.mat_data:
                art = str(m.get("Артикул", "")).strip()
                if not art:
                    continue
                vals = [m.get("Цена мата", 0), m.get("Вес мата", 0), m.get("Цена м²", 0), m.get("Вес м²", 0)]
                if any(_pos(v) for v in vals) or _pos(m.get("Вес м²", 0)):
                    filtered.append(art)

            # Чем показывать: отфильтрованные или, если их нет, все
            articles = sorted(set(filtered)) if filtered else sorted(set(all_articles))

            self.article_cb.configure(values=[str(a) for a in articles])
            target = prefer or self._last_article

            if articles:
                if target in articles:
                    self.article_cb.set(target)
                else:
                    # если ничего не выбрано — ставим первый
                    if not self.article_cb.get():
                        self.article_cb.set(articles[0])
            else:
                self.article_cb.set("")
        except Exception as e:
            logging.error(f"_refresh_article_combobox error: {e}")
            try:
                self.article_cb.configure(values=[])
                self.article_cb.set("")
            except Exception:
                pass


    # =============================
    # Импорт/сохранение справочников
    # =============================
    def import_mat_data(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")],
            title="Выберите файл с данными матов")
        if not file_path:
            return
         
        try:
            df = pd.read_excel(file_path)
            df = df.fillna('')
            self.mat_data = []

            def _f(v, default=0.0):
                s = str(v).strip()
                if s == "":
                    return default
                try:
                    return float(s.replace(",", "."))
                except Exception:
                    return default

            for _, row in df.iterrows():
                price_m2_cell  = row.get('Цена м²', row.get('Цена м2', ''))
                weight_m2_cell = row.get('Вес м²',  row.get('Вес м2',  ''))

                self.mat_data.append({
                    "Артикул":     str(row.get('Артикул', '')).strip(),
                    "Ячейка":      row.get('Ячейка', ''),
                    "Полоса":      row.get('Полоса', ''),
                    "Цена мата":   _f(row.get('Цена мата', 0), 0.0),
                    "Вес мата":    _f(row.get('Вес мата', 0), 0.0),
                    "Цена м²":     _f(price_m2_cell, 0.0),   # ← берем ТОЛЬКО из файла, ничего не пересчитываем
                    "Вес м²":      _f(weight_m2_cell, 0.0),  # ← берем ТОЛЬКО из файла
                    "Длина мата":  _f(row.get('Длина мата', 6100), 6100.0),
                })

            # НИКАКИХ «нормализаций» и делений на 6.1 далее
            self.refresh_mat_table()
            self._refresh_article_combobox()
            self._last_article = self.article_cb.get()
            self.save_config()
            messagebox.showinfo("Успех", "Данные матов успешно импортированы")
        except Exception as e:
            logging.error(f"import_mat_data: {e}")
            messagebox.showerror("Ошибка", f"Не удалось импортировать данные матов:\n{e}")


    

    def import_weight_data(self, name: str):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")], title=f"Выберите файл с весами {name}")
        if not file_path:
            return
            
        
        try:
            df = pd.read_excel(file_path).fillna('')
            lst: List[Tuple[str, float]] = []
            for _, r in df.iterrows():
                try:
                    lst.append((str(r[0]), float(str(r[1]).replace(',', '.'))))
                except Exception:
                    continue
            if name == "Обрамление":
                self.frame_weights = lst
            elif name == "Перфоуголок":
                self.angle_weights = lst
            elif name == "Отбойник":
                self.bumper_weights = lst
            self.refresh_material_table(name)
            self.populate_material_combos()
            self.save_config()
            messagebox.showinfo("Успех", f"Веса '{name}' импортированы")
        except Exception as e:
            logging.error(f"import_weight_data: {e}")
            messagebox.showerror("Ошибка", f"Не удалось импортировать веса '{name}':\n{e}")

    

    # =============================
    # Операции с деталями
    # =============================

    def add_part(self):
        try:
            position = self.part_entries['позиция'].get().strip()
            if not position:
                messagebox.showerror("Ошибка", "Поле «Позиция» обязательно для заполнения.")
                self.position_entry.focus_set()
                return
            for part in self.project_parts:
                if part.position == position:
                    messagebox.showwarning("Дубликат позиции", f"Деталь с позицией «{position}» уже существует.")
                    return

            length = int(self.part_entries['длина'].get() or 0)
            width = int(self.part_entries['ширина'].get() or 0)
            quantity = int(self.part_entries['кол-во'].get() or 1)
            angle = self.part_entries['перфоугол'].get().strip()
            bumper = int(self.part_entries['отбойник'].get() or 0)
            work = float(self.part_entries['работа'].get() or 0)
            radius = int(self.part_entries['радиус'].get() or 0)
            subtract_bumper = self.subtract_bumper_checkbox.get()  # Получаем состояние чекбокса

            # Получаем долю окружности
            try:
                radius_part_str = self.part_entries.get("доля радиуса")
                radius_part = float(radius_part_str.get().strip() or "1.0")
            except Exception:
                messagebox.showerror("Ошибка", "Неверное значение доли окружности.")
                return

            # Вычисляем длину отбойника
            linear_mm, total_bumper_mm = self.calculate_bumper_lengths(float(bumper), radius, radius_part)
            bumper_length_m = round(total_bumper_mm / 1000, 2)

            # Периметр с учетом вычитания отбойника
            perimeter_mm = (2 * width + length) if width < 1000 else (2 * width)
            if bumper and subtract_bumper:  # Используем индивидуальное значение
                perimeter_mm -= linear_mm
            perimeter_m = max(0, round(perimeter_mm / 1000, 2))

            new_part = ProjectPart(
                position, length, width, quantity, angle,
                bumper, work, radius, perimeter_m, bumper_length_m,
                radius_part, subtract_bumper  # Добавляем состояние чекбокса
            )
            self.project_parts.append(new_part)
            self.update_parts_table()
            self.recalculate_project()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Некорректные данные детали: {e}")



    def update_parts_table(self):
        # Полная перерисовка с надежными iid и маппингом
        self._tv_to_part.clear()
        self.parts_table.delete(*self.parts_table.get_children())

        for idx, p in enumerate(self.project_parts):
            iid = f"part-{idx}"  # стабильный iid
            area = (p.length * p.width) / 1_000_000
            self.parts_table.insert(
                "",
                "end",
                iid=iid,
                values=(
                    p.position,
                    f"{p.length:.0f}",
                    f"{p.width:.0f}",
                    p.quantity,
                    f"{area:.3f}".replace('.', ','),
                    f"{p.perimeter:.2f}".replace('.', ','),
                    p.angle,
                    p.bumper,
                    f"{p.work:.1f}".replace('.', ','),
                    f"{p.radius:.0f}" if p.radius else "",
                    f"{p.bumper_length:.2f}" if p.bumper_length > 0 else "0.00",
                ),
            )
            self._tv_to_part[iid] = p


    def edit_selected_part(self, event=None):
        messagebox.showinfo("Инфо", "Редактор деталей не реализован в этой сборке. Удалите и добавьте деталь заново.")

    def clear_part_fields(self):
        for e in self.part_entries.values():
            if isinstance(e, (ctk.CTkEntry, ctk.CTkComboBox)):
                if isinstance(e, ctk.CTkEntry):
                    e.delete(0, 'end')
                elif isinstance(e, ctk.CTkComboBox):
                    e.set("")
        # Сбрасываем чекбокс к значению по умолчанию
        self.subtract_bumper_checkbox.select()

    def clear_project(self):
        if messagebox.askyesno("Подтверждение", "Очистить проект?"):
            self.project_parts.clear()
            self.parts_table.delete(*self.parts_table.get_children())
            self.cost_table.delete(*self.cost_table.get_children())
            self.weight_table.delete(*self.weight_table.get_children())
            for var in self.project_stats.values():
                var.set("0.00")
            self.status_var.set("Проект очищен")

    # =============================
    # Расчеты
    # =============================
    def _price_of(self, name: str) -> float:
        """Цена из блока 'Материалы и цены'. Поддерживает русские названия и английские ключи."""
        try:
            key = str(name).strip().lower()
            alias = {
                "обрамление": "frame", "рамка": "frame", "frame": "frame",
                "перфоугол": "angle", "перфоуголок": "angle", "уголок": "angle", "angle": "angle",
                "отбойник": "bumper", "bumper": "bumper",
                "работа": "work", "work": "work",
            }.get(key, None)
            if not alias:
                return 0.0
            entry = self.material_entries.get(alias)
            if not entry:
                return 0.0
            return float((entry.get() or "0").replace(",", "."))
        except Exception:
            return 0.0



    def _weight_per_m(self, table: List[Tuple[str, float]], type_name: str) -> float:
        for t, w in table:
            if str(t) == str(type_name):
                return float(w)
        return 0.0

    def calculate_mats_count(self) -> float:
        """Сырое количество из поля (БЕЗ +15%), округление для UI/экспорта."""
        s = (self.mats_count_entry.get() or "").strip().replace(",", ".")
        try:
            return round(float(s), 1)
        except Exception:
            return 0.0

            
    def calculate_bumper_lengths(self, bumper_mm, radius, radius_part=1.0) -> Tuple[float, float]:
        import math

        # Преобразуем входные значения в числа
        try:
            bumper_mm = float(bumper_mm)
        except Exception:
            bumper_mm = 0.0

        try:
            radius = float(radius)
        except Exception:
            radius = 0.0
            
        try:
            radius_part = float(radius_part)
        except Exception:
            radius_part = 1.0


        # Прямая часть отбойника
        linear_mm = bumper_mm 

        # Длина дуги (если есть радиус и отбойник)
        arc_length = 2 * math.pi * radius * radius_part if radius > 0 else 0.0

        total_bumper_mm = linear_mm + arc_length
        return linear_mm, total_bumper_mm

    def _combo_get(self, key: str, default: str = "") -> str:
        cb = self.material_combos.get(key)
        try:
            return cb.get() if cb else default
        except Exception:
            return default

    def calculate_frame_length_m(self) -> float:
        total = 0.0
        for p in self.project_parts:
            # Пересчитываем периметр для каждой детали с учетом ее индивидуальных настроек
            linear_mm, total_bumper_mm = self.calculate_bumper_lengths(p.bumper, p.radius, p.radius_part)
            per_mm = (2 * p.width + p.length) if p.width < 1000 else (2 * p.width)
            if p.bumper and p.subtract_bumper:  # Индивидуальная настройка для каждой детали
                per_mm -= linear_mm
            perimeter_m = max(0, round(per_mm / 1000, 2))
            total += perimeter_m * p.quantity
        return round(total, 2)

    def calculate_angle_piece_count(self) -> float:
        try:
            total_length_m = self.calculate_angle_length_m()  # длина в метрах
            pieces = total_length_m / 6.0                     # базовое кол-во
            # округление вверх до ближайшей 0.1
            pieces_rounded = math.ceil(pieces * 10) / 10.0
            return pieces_rounded
        except Exception:
            return 0.0
    

    def calculate_bumper_length_m(self) -> float:
        return round(sum(p.bumper_length * p.quantity for p in self.project_parts if p.bumper_length > 0), 2)
    def _selected_mat_weight_per_m2(self) -> float:
        """Вес м² для выбранного артикула из self.mat_data."""
        art = self.article_cb.get()
        for m in self.mat_data:
            if m.get("Артикул") == art:
                try:
                    return float(m.get("Вес м²", 0.0))
                except Exception:
                    return 0.0
        return 0.0
        
    def _selected_mat_price_per_m2(self) -> float:
        """Возвращает цену м² выбранного мата ТОЛЬКО из таблицы матов.
        Никаких автопересчетов из 'Цена мата'."""
        try:
            article = (self.article_cb.get() or "").strip()
        except Exception:
            article = ""
        for row in self.mat_data:
            if str(row.get("Артикул", "")).strip() == article:
                try:
                    return float(str(row.get("Цена м²", 0)).replace(",", "."))
                except Exception:
                    return 0.0
        return 0.0
            
    
    def _get_mats_count_for_cost(self) -> float:
        """
        Возвращает количество матов для РАСЧЁТА СТОИМОСТИ:
        - ручной режим (manual_mats_var = True): ровно введённое число;
        - авто-режим: введённое/посчитанное число × 1.15 (без округления).
        """
        raw = (self.mats_count_entry.get() or "").strip().replace(",", ".")
        try:
            base = float(raw)
        except Exception:
            base = 0.0

        manual = getattr(self, "manual_mats_var", None)
        is_manual = bool(manual.get()) if manual is not None else False
        return base if is_manual else (base * 1.15)
        
    def calculate_mat_cost(self) -> float:
        """
        Себестоимость матов = (кол-во × K) × (площадь листа, м² × цена м²).
        K = 1.15 в авто-режиме (чекбокс выключен), K = 1.00 в ручном режиме.
        'Цена мата' НЕ используется.
        """
        try:
            base_count = float(self.calculate_mats_count() or 0.0)
        except Exception:
            base_count = 0.0

        # Коэффициент: авто = +15%, ручной = без надбавки
        manual = getattr(self, "manual_mats_var", None)
        k = 1.0 if (manual and manual.get()) else 1.15

        # Площадь одного листа по введённым размерам (мм -> м²)
        L_mm, W_mm = self._get_mat_dimensions_mm()
        if L_mm <= 0 or W_mm <= 0:
            return 0.0
        area_m2 = (L_mm * W_mm) / 1_000_000.0

        # Цена м² строго из таблицы матов
        price_m2 = self._selected_mat_price_per_m2()

        total = base_count * k * area_m2 * price_m2
        return round(total, 2)



    def calculate_frame_cost(self) -> float:
        try:
            type_name = self._combo_get("Обрамление")
            wpm = self._weight_per_m(self.frame_weights, type_name) or 0.0
            length_m = self.calculate_frame_length_m() or 0.0
            coef_str = self.frame_coef_entry.get().replace(",", ".") or "1.0"
            coef = float(coef_str)
            price = self._price_of("обрамление") or 0.0
            return round(length_m * wpm * price * coef, 2)
        except Exception:
            return 0.0

    def calculate_angle_cost(self) -> float:
        """Стоимость перфоуголков по количеству штук и цене за 1 шт (6 м)"""
        try:
            count = self.calculate_angle_piece_count() or 0.0
            price_per_piece = self._price_of("angle") or 0.0
            return round(count * price_per_piece, 2)
        except Exception:
            return 0.0


    

    def calculate_bumper_cost(self) -> float:
        try:
            if not any(p.bumper > 0 for p in self.project_parts):
                return 0.0
                
            type_name = self._combo_get("Отбойник")
            wpm = self._weight_per_m(self.bumper_weights, type_name) or 0.0
            length_m = self.calculate_bumper_length_m() or 0.0
            price = self._price_of("отбойник") or 0.0
            return round(length_m * wpm * price, 2)
        except Exception:
            return 0.0

    def calculate_work_cost(self) -> float:
        try:
            hours = round(sum(p.work * p.quantity for p in self.project_parts), 2) or 0.0
            price = self._price_of("работа") or 0.0
            coef_str = self.work_coeff.get().replace(",", ".") or "1.0"
            coef = float(coef_str)
            return round(hours * price * coef, 2)
        except Exception:
            return 0.0

    def calculate_zinc_cost(self) -> float:
        """Считает стоимость цинка от веса с +10%."""
        try:
            # фикс: поддержка запятой в поле ввода
            zinc_str = (self.zinc_entry.get() or "0").replace(",", ".")
            zinc_price = float(zinc_str)
            weight_with_zinc = self.calculate_total_weight_with_zinc() or 0.0
            return round(zinc_price * weight_with_zinc, 2)
        except Exception:
            return 0.0
        

    def calculate_total_weight(self) -> float:
        # Вес Решетки (маты) по площади
        mats_w = self.calculate_total_weight_mats_only()
        frame_w  = self.calculate_frame_weight_only()
        angle_w = self.calculate_angle_length_m() * self._weight_per_m(self.angle_weights, self._combo_get("Перфоуголок"))
        bumper_w = self.calculate_bumper_length_m() * self._weight_per_m(self.bumper_weights, self._combo_get("Отбойник"))

        total = mats_w + frame_w + angle_w + bumper_w
        return round(total, 2)
        
    def calculate_total_weight_with_zinc(self) -> float:
        """Возвращает вес с учетом +10% для цинкования."""
        base_weight = self.calculate_total_weight()
        return round(base_weight * 1.1, 2)

    def update_cost_table(self, total_cost: float):
        self.cost_table.delete(*self.cost_table.get_children())
        self.cost_table.insert('', 'end', values=("Маты + 15%", f"{self.calculate_mat_cost():.2f}".replace('.', ',')))
        self.cost_table.insert('', 'end', values=("Обрамление", f"{self.calculate_frame_cost():.2f}".replace('.', ',')))
        angle_pieces = self.calculate_angle_piece_count()
        angle_cost = self.calculate_angle_cost()
        self.cost_table.insert('', 'end', values=(f"Перфоуголок ({angle_pieces} шт)", f"{angle_cost:.2f}".replace('.', ',')))

        self.cost_table.insert('', 'end', values=("Отбойник", f"{self.calculate_bumper_cost():.2f}".replace('.', ',')))
        self.cost_table.insert('', 'end', values=("Цинкование", f"{self.calculate_zinc_cost():.2f}".replace('.', ',')))
        self.cost_table.insert('', 'end', values=("Работа", f"{self.calculate_work_cost():.2f}".replace('.', ',')))

    def update_weight_table(self, total_weight: float):
        self.weight_table.delete(*self.weight_table.get_children())
        self.weight_table.insert('', 'end', values=("Решетка", f"{self.calculate_total_weight_mats_only():.2f}".replace('.', ',')))
        self.weight_table.insert('', 'end', values=("Обрамление", f"{self.calculate_frame_weight_only():.2f}".replace('.', ',')))
        self.weight_table.insert('', 'end', values=("Перфоуголок", f"{self.calculate_angle_weight_only():.2f}".replace('.', ',')))
        self.weight_table.insert('', 'end', values=("Отбойник", f"{self.calculate_bumper_weight_only():.2f}".replace('.', ',')))
        self.weight_table.insert('', 'end', values=("Общий черный вес", f"{total_weight:.2f}".replace('.', ',')))
        self.weight_table.insert('', 'end', values=("Общий вес с цинком (+10%)", f"{self.calculate_total_weight_with_zinc():.2f}".replace('.', ',')))

    def calculate_total_weight_mats_only(self) -> float:
        """
        Вес Решетки = суммарная площадь всех деталей (м²) * Вес м² (для выбранного артикула).
        """
        per_m2 = self._selected_mat_weight_per_m2()
        total_area_m2 = sum(((p.length * p.width) / 1_000_000.0) * p.quantity for p in self.project_parts)
        return round(total_area_m2 * per_m2, 2)


    def calculate_frame_weight_only(self) -> float:
        """Вес обрамления с учетом коэффициента обрамления."""
        try:
            coef = float((self.frame_coef_entry.get() or "1").replace(",", "."))
        except Exception:
            coef = 1.00
        base = self.calculate_frame_length_m() * self._weight_per_m(self.frame_weights, self._combo_get("Обрамление"))
        return round(base * coef, 2)
    def calculate_angle_length_m(self) -> float:
        total = 0.0
        for p in self.project_parts:
            v = str(p.angle).strip()
            if not v or v == "0":
                continue
            try:
                length_mm = float(v)
                total += (length_mm / 1000.0) * p.quantity  # перевод в метры
            except ValueError:
                # Неправильный формат — игнорируем или логируем при желании
                continue
        return round(total, 2)

    def calculate_angle_weight_only(self) -> float:
        return round(self.calculate_angle_length_m() * self._weight_per_m(self.angle_weights, self._combo_get("Перфоуголок")), 2)


    def calculate_bumper_weight_only(self) -> float:
        # Проверяем, есть ли детали с отбойником
        if not any(p.bumper > 0 for p in self.project_parts):
            return 0.0
        return round(self.calculate_bumper_length_m() * self._weight_per_m(self.bumper_weights, self._combo_get("Отбойник")), 2)

    def recalculate_project(self):
        """
        Главный пересчет. Все вспомогательные calculate_*() сами читают
        текущие значения из UI (Entry/Combo) и возвращают 0 при пустых данных.
        """
        try:
            # Итоговая стоимость по компонентам
            total_cost = (
                getattr(self, "calculate_mat_cost",     lambda: 0.0)() +
                getattr(self, "calculate_frame_cost",   lambda: 0.0)() +
                getattr(self, "calculate_angle_cost",   lambda: 0.0)() +
                getattr(self, "calculate_bumper_cost",  lambda: 0.0)() +
                getattr(self, "calculate_zinc_cost",    lambda: 0.0)() +
                getattr(self, "calculate_work_cost",    lambda: 0.0)()
            )

            # Итоговый вес и метрики
            total_weight = getattr(self, "calculate_total_weight", lambda: 0.0)()
            mats_count   = getattr(self, "calculate_mats_count",   lambda: 0.0)()

            cost_per_kg = (total_cost / total_weight) if total_weight > 0 else 0.0

            # Карточки справа
            self.project_stats["Кол-во матов"].set(f"{mats_count:.1f}")
            self.project_stats["Себестоимость €/кг"].set(f"{cost_per_kg:.2f}")
            self.project_stats["Цена проекта €"].set(f"{total_cost:.2f}")
            self.project_stats["Общий черный вес"].set(f"{total_weight:.2f}")

            # Таблицы результатов
            if hasattr(self, "update_cost_table"):
                self.update_cost_table(total_cost)
            if hasattr(self, "update_weight_table"):
                self.update_weight_table(total_weight)

            if hasattr(self, "status_var"):
                self.status_var.set("Пересчет выполнен")
        except Exception as e:
            logging.exception("Ошибка в recalculate_project")
            messagebox.showerror("Ошибка", f"Ошибка пересчета: {e}")

            
    def _validate_data(self) -> bool:
        """Проверяет минимально необходимые данные для расчета"""
        required = [
            bool(self.project_parts),
            bool(self.article_cb.get()),
            bool(self.mats_count_entry.get()),
            all(entry.get() for entry in self.material_entries.values())
        ]
        return all(required)
        
    

    # =============================
    # Проекты (сохранение/загрузка)
    # =============================
    def save_project(self):
        """Сохранить проект в .gridmaster (JSON)."""
        project_name = self.project_name_entry.get().strip() or "МойПроект"
        default_docs = os.path.expanduser("~/Documents")
        initial_dir = self.last_export_dir if (self.last_export_dir and os.path.isdir(self.last_export_dir)) else default_docs
        file_path = filedialog.asksaveasfilename(
            title="Сохранить проект",
            initialdir=initial_dir,
            initialfile=f"{project_name}.gridmaster",
            defaultextension=".gridmaster",
            filetypes=[("Gridmaster Project", "*.gridmaster"), ("All Files", "*.*")]
        )
        if not file_path:
            return

        try:
            data = {
                "parts": [vars(p) for p in self.project_parts],
                "params": {
                    "project_name": project_name,
                    "work_coef": self.work_coeff.get(),
                    "zinc_price": self.zinc_entry.get(),
                    "frame_coef": self.frame_coef_entry.get(),
                    "mats_count": self.mats_count_entry.get(),
                    "mat_article": self.article_cb.get() or "",
                    "frame_type": self.material_combos["Обрамление"].get() if self.material_combos["Обрамление"] else "",
                    "angle_type": self.material_combos["Перфоуголок"].get() if self.material_combos["Перфоуголок"] else "",
                    "bumper_type": self.material_combos["Отбойник"].get() if self.material_combos["Отбойник"] else "",
                },
                "tables": {
                    "mat_data": self.mat_data,
                    "material_prices": self.material_prices,
                    "frame_weights": self.frame_weights,
                    "angle_weights": self.angle_weights,
                    "bumper_weights": self.bumper_weights,
                },
                "meta": {
                    "saved_at": datetime.now().isoformat(timespec="seconds")
                }
            }

            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            # Обновим «последнюю папку» и путь последнего проекта
            self.last_export_dir = os.path.dirname(file_path)
            self.current_project_path = file_path
            self.save_config()
            self.status_var.set(f"Проект сохранен: {file_path}")
            messagebox.showinfo("Успех", "Проект сохранен")
        except Exception as e:
            logging.error(f"save_project error: {e}")
            messagebox.showerror("Ошибка", f"Не удалось сохранить проект:\n{e}")
            
    def open_project_dialog(self):
        """Диалог выбора .gridmaster и загрузка проекта."""
        default_docs = os.path.expanduser("~/Documents")
        initial_dir = self.last_export_dir if (self.last_export_dir and os.path.isdir(self.last_export_dir)) else default_docs
        file_path = filedialog.askopenfilename(
            title="Открыть проект",
            initialdir=initial_dir,
            filetypes=[("gridmaster Project", "*.gridmaster"), ("All Files", "*.*")]
        )
        if not file_path:
            return
        self.load_project(file_path) 
     
    def load_config(self):
        """Загружает конфигурацию из файла"""
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                    
                    # Загрузка данных матов
                    self.mat_data = cfg.get("mat_data", [])
                    
                    # Загрузка последнего артикула
                    self._last_article = cfg.get("last_article")
                    if self._last_article and hasattr(self, 'article_cb'):
                        self.article_cb.set(self._last_article)
                    
                    # Загрузка последней директории экспорта
                    self.last_export_dir = cfg.get("last_export_dir")
                    
                    # Загрузка темы интерфейса
                    self.ui_theme = cfg.get(THEME_KEY, "dark")
                    ctk.set_appearance_mode(self.ui_theme)
                    
                    # Загрузка цен материалов
                    if 'material_prices' in cfg:
                        prices = cfg['material_prices']
                        if hasattr(self, 'material_entries'):
                            for key, entry in self.material_entries.items():
                                if key in prices:
                                    entry.delete(0, 'end')
                                    entry.insert(0, str(prices[key]))
                    
                    # Загрузка весов материалов
                    self.frame_weights = cfg.get("frame_weights", self.get_default_frame_weights())
                    self.angle_weights = cfg.get("angle_weights", self.get_default_angle_weights())
                    self.bumper_weights = cfg.get("bumper_weights", self.get_default_bumper_weights())
                    
                    # Обновление интерфейса
                    self.refresh_mat_table()
                    self.populate_material_combos()
                    
        except Exception as e:
            logging.error(f"Ошибка загрузки конфига: {e}")
            # Загрузка значений по умолчанию
            self.load_default_values()     
    def load_project(self, file_path: str):
        """Загрузить проект из .gridmaster (JSON)."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            # ---------- Справочники ----------
            tables = data.get("tables", {})
            self.mat_data = tables.get("mat_data", self.mat_data)
            self.material_prices = tables.get("material_prices", self.material_prices)
            self.frame_weights = tables.get("frame_weights", self.frame_weights)
            self.angle_weights = tables.get("angle_weights", self.angle_weights)
            self.bumper_weights = tables.get("bumper_weights", self.bumper_weights)

            # Обновим таблицы и комбобоксы размеров
            self.refresh_mat_table()
            self.refresh_material_table("Цены")
            self.refresh_material_table("Обрамление")
            self.refresh_material_table("Перфоуголок")
            self.refresh_material_table("Отбойник")
            self.populate_material_combos()

            # ---------- Параметры ----------
            params = data.get("params", {})
            # Имя проекта
            self.project_name_entry.delete(0, 'end')
            self.project_name_entry.insert(0, params.get("project_name", "МойПроект"))

            # Коэфы и цены
            self.work_coeff.delete(0, 'end');    self.work_coeff.insert(0, params.get("work_coef", "1.0"))
            self.zinc_entry.delete(0, 'end');    self.zinc_entry.insert(0, params.get("zinc_price", "0"))
            self.frame_coef_entry.delete(0, 'end'); self.frame_coef_entry.insert(0, params.get("frame_coef", "1.0"))

            # Кол-во матов
            self.mats_count_entry.delete(0, 'end'); self.mats_count_entry.insert(0, params.get("mats_count", "0"))

            # Артикул и типы материалов
            self._refresh_article_combobox(prefer=params.get("mat_article", ""))
            if self.material_combos["Обрамление"] and params.get("frame_type"):
                self.material_combos["Обрамление"].set(params["frame_type"])
            if self.material_combos["Перфоуголок"] and params.get("angle_type"):
                self.material_combos["Перфоуголок"].set(params["angle_type"])
            if self.material_combos["Отбойник"] and params.get("bumper_type"):
                self.material_combos["Отбойник"].set(params["bumper_type"])

            # ---------- Детали ----------
            self.project_parts.clear()
            for d in data.get("parts", []):
                try:
                    self.project_parts.append(
                        ProjectPart(
                            position=d.get("position", ""),
                            length=float(d.get("length", 0)),
                            width=float(d.get("width", 0)),
                            quantity=int(d.get("quantity", 0)),
                            angle=d.get("angle", ""),
                            bumper=d.get("bumper", ""),
                            work=float(d.get("work", 0)),
                            radius=float(d.get("radius", 0)),
                            perimeter=float(d.get("perimeter", 0)),
                            bumper_length=float(d.get("bumper_length", 0)),
                            radius_part=float(d.get("radius_part", 1.0)),
                        )
                    )
                except Exception:
                    continue

            self.update_parts_table()
            self.recalculate_project()

            # ---------- Запомним путь и директорию ----------
            self.current_project_path = file_path
            self.last_export_dir = os.path.dirname(file_path)
            self.save_config()

            self.status_var.set(f"Проект загружен: {file_path}")
            messagebox.showinfo("Готово", "Проект загружен")
        except Exception as e:
            logging.error(f"load_project error: {e}")
            messagebox.showerror("Ошибка", f"Не удалось загрузить проект:\n{e}")
            
    def load_last_project(self):
        """Автозагрузка последнего проекта из конфига, если путь существует."""
        try:
            if not os.path.exists(CONFIG_FILE):
                return
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            last_path = cfg.get("last_project")
            if last_path and os.path.exists(last_path):
                self.load_project(last_path)
        except Exception as e:
            logging.warning(f"load_last_project: {e}")
    
    def on_closing(self):
        answer = ask_yes_no_cancel("Выход", "Вы хотите сохранить проект перед выходом?")
        
        if answer is None:  # Отмена
            return
        elif answer:  # Да
            self.save_project()  # ← замени на актуальный метод

        try:
            self._last_article = self.article_cb.get()
            self.save_config()
        finally:
            self.root.destroy()
            
    def f2s(self, value: float, digits: int = 2) -> str:
        return f"{value:.{digits}f}".replace('.', ',')


    # Экспорт Excel
    def export_to_excel(self):
        project_name = self.project_name_entry.get().strip() or "МойПроект"

        # 1) Валидация количества матов (дробное допустимо)
        try:
            mats_count = float((self.mats_count_entry.get() or "0").replace(",", ".").strip())
            mats_count = round(mats_count, 1)
        except Exception:
            mats_count = 0.0
        if mats_count <= 0:
            messagebox.showerror(
                "Ошибка",
                "Не указано количество матов.\n"
                "Заполните поле «Кол-во матов» (число > 0, допускается дробное) и попробуйте снова."
            )
            self.mats_count_entry.focus_set()
            self.mats_count_entry.selection_range(0, 'end')
            return

        # 2) Диалог сохранения. Базовое имя БЕЗ 'Grid'
        import re, openpyxl
        default_docs = os.path.expanduser("~/Documents")
        initial_dir = self.last_export_dir if (self.last_export_dir and os.path.isdir(self.last_export_dir)) else default_docs
        base_name = f"{project_name}.xlsx"  # было f"{project_name}_Grid.xlsx" — убрали 'Grid'  ← (требование)

        def get_unique_filename(dir_path: str, base_filename: str) -> str:
            name, ext = os.path.splitext(base_filename)
            m = re.match(r"^(.*?)(?:_(\d+))?$", name)
            stem = m.group(1) if m else name
            pattern = re.compile(rf"^{re.escape(stem)}_(\d+){re.escape(ext)}$")
            max_idx = 0
            for fn in os.listdir(dir_path):
                mm = pattern.match(fn)
                if mm:
                    max_idx = max(max_idx, int(mm.group(1)))
            return os.path.join(dir_path, f"{stem}_{max_idx+1}{ext}")

        suggested_path = os.path.join(initial_dir, os.path.basename(get_unique_filename(initial_dir, base_name)))
        file_path = filedialog.asksaveasfilename(
            title="Сохранить Excel",
            initialdir=initial_dir,
            initialfile=os.path.basename(suggested_path),
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("Все файлы", "*.*")]
        )
        if not file_path:
            return

        self.last_export_dir = os.path.dirname(file_path) or initial_dir

        # 3) Открываем существующую книгу или создаем новую
        wb = None
        if os.path.exists(file_path):
            try:
                wb = openpyxl.load_workbook(file_path)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть существующий файл:\n{e}")
                return
        else:
            # Создадим новый файл, но с порядковым номером в ИМЕНИ файла <Проект>_N.xlsx
            file_path = get_unique_filename(self.last_export_dir, os.path.basename(file_path))
            wb = openpyxl.Workbook()
            # удалим дефолтный пустой лист
            if wb.active and wb.active.title == "Sheet":
                wb.remove(wb.active)

        # 4) Определяем следующий порядковый номер листов
        def next_sheet_index(book):
            max_idx = 0
            rx_parts = re.compile(r"^Детали(?:_(\d+))?$")
            rx_calc  = re.compile(r"^Реш[ее]тки(?:_(\d+))?$")  # допускаем 'е' и 'е'
            for t in book.sheetnames:
                m1 = rx_parts.match(t)
                m2 = rx_calc.match(t)
                if m1 and m1.group(1):
                    max_idx = max(max_idx, int(m1.group(1)))
                if m2 and m2.group(1):
                    max_idx = max(max_idx, int(m2.group(1)))
                if m1 and not m1.group(1):
                    max_idx = max(max_idx, 1)
                if m2 and not m2.group(1):
                    max_idx = max(max_idx, 1)
            return max(1, max_idx + 1) if (("Детали" in book.sheetnames) or any(re.match(r"^Детали_\d+$", s) for s in book.sheetnames)
                                           or ("Решетки" in book.sheetnames) or any(re.match(r"^Реш[ее]тки_\d+$", s) for s in book.sheetnames)) else 1

        idx = next_sheet_index(wb)
        ws_parts_name = "Детали" if idx == 1 and ("Детали" not in wb.sheetnames) else f"Детали_{idx}"
        ws_calc_name  = "Решетки" if idx == 1 and (("Решетки" not in wb.sheetnames) and ("Решетки" not in wb.sheetnames)) else f"Решетки_{idx}"

        # 5) Создаем лист "Детали" и заполняем (структура как в текущем экспорте)
        from openpyxl.styles import Alignment, Font, Border, Side
        ws_parts = wb.create_sheet(ws_parts_name)
        headers = [
            "Позиция", "Длина (мм)", "Ширина (мм)", "Кол-во",
            "Площадь (м²)", "Обрамление (м)", "Перфоуголок",
            "Отбойник", "Работа (ч)", "Радиус (мм)", "Длина отбойника (м)"
        ]
        ws_parts.append(headers)
        for p in self.project_parts:
            area = (p.length * p.width * p.quantity) / 1_000_000
            ws_parts.append([
                p.position, p.length, p.width, p.quantity,
                round(area, 3), round(p.perimeter, 2), p.angle, p.bumper,
                p.work, p.radius if p.radius else "",
                round(p.bumper_length, 2) if p.bumper_length > 0 else ""
            ])
        # Заголовок проекта и дата вверху таблицы
        today_str = datetime.today().strftime("%d.%m.%Y")
        ws_parts.insert_rows(1)
        ws_parts["A1"].value = f"Проект: {project_name}"
        ws_parts["A1"].font = Font(name="Calibri", size=12, bold=True)
        ws_parts["B1"].value = today_str
        ws_parts["B1"].font = Font(name="Calibri", size=12, bold=True)

        for row in ws_parts.iter_rows(min_row=1, max_row=ws_parts.max_row, min_col=1, max_col=ws_parts.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Авто-границы и подгон ширины
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws_parts.iter_rows(min_row=1, max_row=ws_parts.max_row, min_col=1, max_col=ws_parts.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin
        for col in range(1, ws_parts.max_column + 1):
            col_letter = get_column_letter(col)
            length = max(len(str(ws_parts.cell(row=r, column=col).value or "")) for r in range(1, ws_parts.max_row + 1))
            ws_parts.column_dimensions[col_letter].width = length + 4

        # 6) Создаем лист "Решетки" и переносим показатели (как у тебя сейчас, только в отдельный лист)
        ws_calc = wb.create_sheet(ws_calc_name)
        ws_calc.append(["Компонент", "Тип", "Количество", "Единицы", "Стоимость"])
        for cell in ws_calc[1]:
            cell.font = Font(bold=True)
        def _num(v):
            try:
                return float(str(v).replace(",", "."))
            except Exception:
                return 0.0

        mats_type = self.article_cb.get()
        mats_count_val = getattr(self, "calculate_mats_count", lambda: mats_count)()
        mats_cost = getattr(self, "calculate_mat_cost", lambda: 0.0)()

        frame_type = self.material_combos["Обрамление"].get()
        frame_weight = getattr(self, "calculate_frame_weight_only", lambda: 0.0)()
        frame_cost = getattr(self, "calculate_frame_cost", lambda: 0.0)()

        angle_type = self.material_combos["Перфоуголок"].get()
        angle_pieces = getattr(self, "calculate_angle_piece_count", lambda: 0.0)()
        angle_cost = getattr(self, "calculate_angle_cost", lambda: 0.0)()

        bumper_type = self.material_combos["Отбойник"].get()
        bumper_weight = getattr(self, "calculate_bumper_weight_only", lambda: 0.0)()
        bumper_cost = getattr(self, "calculate_bumper_cost", lambda: 0.0)()

        zinc_price = _num(self.zinc_entry.get() or 0)
        zinc_cost = getattr(self, "calculate_zinc_cost", lambda: 0.0)()

        total_hours = round(sum(p.work * p.quantity for p in self.project_parts), 2)
        work_cost = getattr(self, "calculate_work_cost", lambda: 0.0)()

        # агрегаты по проекту
        total_area_m2 = round(sum((p.length * p.width * p.quantity) / 1_000_000 for p in self.project_parts), 3)
        total_items = sum(p.quantity for p in self.project_parts)

        # вес и стоимость из project_stats (с запасом по ошибкам)
        try:
            total_weight = _num(self.project_stats["Общий черный вес"].get())
        except Exception:
            total_weight = _num(self.project_stats.get("Общий черный вес", 0))

        try:
            total_cost = _num(self.project_stats["Цена проекта €"].get())
        except Exception:
            total_cost = _num(self.project_stats.get("Цена проекта €", 0))

        # вес с цинком (+10%)
        total_weight_zinc = round(total_weight * 1.10, 2)

        # €/кг: если есть готовое поле — берем; иначе считаем
        try:
            cost_per_kg = _num(self.project_stats["Себестоимость €/кг"].get())
        except Exception:
            cost_per_kg = _num(self.project_stats.get("Себестоимость €/кг", 0))

        if (not cost_per_kg) and total_weight_zinc > 0:
            cost_per_kg = round(total_cost / total_weight_zinc, 2)
        
        ws_calc.append(["Количество матов", mats_type, f"{mats_count_val:.1f}", "шт", f"{mats_cost:.2f}"])
        ws_calc.append(["Перфоуголок", angle_type, round(angle_pieces, 1), "шт", angle_cost])
        ws_calc.append(["Обрамление",  frame_type, round(frame_weight, 2), "кг",  frame_cost])
        ws_calc.append(["Отбойник",    bumper_type, round(bumper_weight, 2), "кг", bumper_cost])
        ws_calc.append(["Цинк", zinc_price, "", "евро/кг", zinc_cost])
        ws_calc.append(["Работа всего часов", "", round(total_hours, 1), "часов", work_cost])
        ws_calc.append(["Общий вес черный", "", round(total_weight, 2), "кг", ""])
        ws_calc.append(["Общий вес с цинком (+10%)", "", round(total_weight_zinc, 2), "кг", ""])
        ws_calc.append(["Площадь всех решеток", "", round(total_area_m2, 3), "м²", ""])
        ws_calc.append(["Количество деталей всего", "", int(total_items), "шт", ""])  # 👈 целое число
        ws_calc.append(["Стоимость решеток всего", "", "", "EUR", round(total_cost, 2)])
        ws_calc.append(["Стоимость евро/кг", "", "", "евро/кг", round(cost_per_kg, 2)])

               # Форматирование числовых колонок (как в твоем текущем варианте) :contentReference[oaicite:2]{index=2}
        from openpyxl.styles import numbers
        for r in range(2, ws_calc.max_row + 1):  # начинаем со 2-й строки, т.к. 1-я — заголовки
            label = ws_calc[f"A{r}"].value
            # Колонка C
            c = ws_calc[f"C{r}"]
            if label == "Количество матов":
                c.number_format = "0.0"
            elif label == "Количество деталей всего":
                c.number_format = "0"
                if not isinstance(c.value, int):
                    try:
                        c.value = int(float(c.value))
                    except Exception:
                        pass
            elif isinstance(c.value, (int, float)):
                c.number_format = "0.###"
            # Колонка E
            e = ws_calc[f"E{r}"]
            if isinstance(e.value, (int, float)):
                e.number_format = "0.00"
                    

        # Колонка C: шт/кг/м² — до 3 знаков, но для "Количество матов" — 1 знак
        from openpyxl.styles import numbers
        for r in range(3, ws_calc.max_row + 1):
            if ws_calc[f"A{r}"].value == "Количество матов":
                ws_calc[f"C{r}"].number_format = "0.0"
            else:
                c = ws_calc[f"C{r}"]
                if isinstance(c.value, (int, float)):
                    c.number_format = "0.###"
            e = ws_calc[f"E{r}"]
            if isinstance(e.value, (int, float)):
                e.number_format = "0.00"
                
        # --- автоширина и рамки, как у тебя было ---
        for col in range(1, ws_calc.max_column + 1):
            col_letter = get_column_letter(col)
            length = max(len(str(ws_calc.cell(row=r, column=col).value or "")) for r in range(1, ws_calc.max_row + 1))
            ws_calc.column_dimensions[col_letter].width = length + 4

        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws_calc.iter_rows(min_row=1, max_row=ws_calc.max_row, min_col=1, max_col=ws_calc.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # --- принудительно убираем запятую у "Количество деталей всего" ---
        row_idx = None
        for r in range(1, ws_calc.max_row + 1):
            a = ws_calc[f"A{r}"].value
            if a in ("Количество деталей всего", "Количество деталей"):  # на всякий случай обе формулировки
                row_idx = r
                break

        if row_idx:
            c = ws_calc.cell(row=row_idx, column=3)  # колонка C
            try:
                c.value = int(float(c.value))  # строго целое
            except Exception:
                # если вдруг там текст вида "20," — вычистим
                v = str(c.value).strip().replace(",", "").replace(" ", "")
                c.value = int(float(v)) if v else 0
            c.number_format = "0"      # формат: целое без дробей и разделителей
        # 7) Сохраняем
        try:
            wb.save(file_path)
            self.save_config()
            messagebox.showinfo("Успех", f"Экспорт выполнен в:\n{file_path}\nДобавлены листы: «{ws_parts_name}», «{ws_calc_name}».")
        except Exception as e:
            logging.error(f"Ошибка экспорта: {e}")
            messagebox.showerror("Ошибка", f"Экспорт не удался:\n{e}")

        # =============================
        # Импорт деталей из Excel (опционально)
        # =============================
    def import_parts_from_excel(self):
        """Импорт деталей напрямую в таблицу из Excel файла (устойчивый к формату)."""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx;*.xls"), ("All Files", "*.*")],
            title="Выберите файл с данными деталей"
        )
        if not file_path:
            return

        # --- Всегда задаем название проекта из имени выбранного файла ---
        try:
            import os
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            self.project_name_entry.delete(0, 'end')
            self.project_name_entry.insert(0, base_name)
            # если нужно — сохраняем в конфиг:
            self.save_config()
        except Exception as e:
            logging.warning(f"Не удалось проставить имя проекта из файла: {e}")
        # --- конец блока автозадания имени ---

        try:
            # Читаем и нормализуем заголовки
            df = pd.read_excel(file_path).fillna('')
            df.columns = [str(c).strip() for c in df.columns]

            # Разрешенные синонимы заголовков
            rename_map = {
                "Позиция": "Позиция",
                "Длина": "Длина",
                "Длина (мм)": "Длина",
                "Ширина": "Ширина",
                "Ширина (мм)": "Ширина",
                "Количество": "Количество",
                "Кол-во": "Количество",
                "Перфоуголок": "Перфоуголок",
                "Перфоуголок": "Перфоуголок",
                "Отбойник": "Отбойник",
                "Работа": "Работа",
                "Работа (ч)": "Работа",
                "Радиус": "Радиус",
                "Радиус (мм)": "Радиус",
            }
            df = df.rename(columns={c: rename_map.get(c, c) for c in df.columns})

            required = ["Позиция", "Длина", "Ширина", "Количество"]
            if not all(c in df.columns for c in required):
                missing = [c for c in required if c not in df.columns]
                messagebox.showerror("Ошибка", f"Отсутствуют обязательные столбцы: {', '.join(missing)}")
                return

            # Утилита для перевода чисел с запятыми и единицами в float / int
            def to_float(x, default=0.0):
                if x is None:
                    return float(default)
                s = str(x).strip().replace(',', '.')
                # выкидываем все, кроме цифр, точки и минуса (оставим числа типа 12.34)
                allowed = set("0123456789.-")
                s = ''.join(ch for ch in s if ch in allowed)
                try:
                    return float(s) if s not in ("", "-", ".", "-.", ".-") else float(default)
                except Exception:
                    return float(default)

            def to_int(x, default=0):
                return int(round(to_float(x, default)))

            imported = 0
            errors = 0

            for _, row in df.iterrows():
                try:
                    position = str(row.get("Позиция", "")).strip()
                    if not position:
                        continue

                    length = to_float(row.get("Длина", 0))
                    width = to_float(row.get("Ширина", 0))
                    quantity = to_int(row.get("Количество", 1))

                    angle = str(row.get("Перфоуголок", "")).strip()
                    bumper = str(row.get("Отбойник", "")).strip()
                    work = to_float(row.get("Работа", 0))
                    radius = to_float(row.get("Радиус", 0))

                    try:
                        bumper_mm = to_float(row.get("Отбойник", 0), 0.0)
                    except Exception:
                        bumper_mm = 0.0

                    linear_mm, total_bumper_mm = self.calculate_bumper_lengths(bumper_mm, radius, 1.0)
                    bumper_len_m = round(total_bumper_mm / 1000.0, 2)

                    # Периметр: вычитаем ТОЛЬКО линейный отбойник, если чекбокс включен
                    per_mm = (2 * width + length) if width < 1000 else (2 * width)
                    if bumper_mm > 0 and self.subtract_bumper_var.get():
                        per_mm -= linear_mm
                    per_m = max(0.0, round(per_mm / 1000.0, 2))

                    self.project_parts.append(ProjectPart(
                        position=position,
                        length=length,
                        width=width,
                        quantity=quantity,
                        angle=angle,
                        bumper=bumper_mm,        # <- сохраняем ЧИСЛО мм
                        work=work,
                        radius=radius,
                        perimeter=per_m,
                        bumper_length=bumper_len_m
                    ))
                    imported += 1

                except Exception:
                    errors += 1
                    continue

            self.update_parts_table()
            self.recalculate_project()

            if imported > 0:
                msg = f"Импортировано {imported} дет."
                if errors:
                    msg += f"\nПропущено строк: {errors} (ошибки формата)."
                messagebox.showinfo("Успех", msg)
            else:
                messagebox.showwarning("Внимание", "Не удалось импортировать ни одной детали. Проверьте формат файла.")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка импорта:\n{e}")
            logging.error(f"Ошибка импорта: {e}")


    # =============================
    # Вспомогательные веса (для таблицы "Вес")
    # =============================
    def update_weight_table_snapshot(self):
        total_weight = self.calculate_total_weight()
        self.update_weight_table(total_weight)

# =============================
# Запуск приложения
# =============================
if __name__ == "__main__":
    # 1) Прочтем тему из конфига, чтобы применить ее до создания интерфейса
    def _read_theme_from_config(path="gridmaster_config.ini"):
        import json, os
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                t = cfg.get(THEME_KEY, "dark")
                return t if t in ("dark", "light") else "dark"
            except Exception:
                return "dark"
        return "dark"

    ctk.set_default_color_theme(COLOR_THEME)
    ctk.set_appearance_mode(_read_theme_from_config())

    root = ctk.CTk()
    app = GridMasterInterface(root)
    app.recalculate_project()
    root.title("Grid Master V.5.0 ©Viiptag")
    root.mainloop()
