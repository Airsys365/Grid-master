"""
manager_app.py — прога менеджера Grid Master.

Запуск: python manager_app.py
"""

import json
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import List, Dict, Optional
import customtkinter as ctk

import gridmaster_core as core
import price_manager as pm
import ai_recognizer as ai
import pdf_export as pdf_exp
import spec_reader

APP_TITLE   = "Grid Master — Менеджер"


def _bind_tooltip(widget, text: str):
    tip_win = [None]

    def show(event):
        if tip_win[0]:
            return
        x = widget.winfo_rootx() + 20
        y = widget.winfo_rooty() + 20
        w = tk.Toplevel(widget)
        w.wm_overrideredirect(True)
        w.wm_geometry(f"+{x}+{y}")
        tk.Label(w, text=text, background="#ffffe0", relief="solid",
                 borderwidth=1, font=("Segoe UI", 9), wraplength=300,
                 justify="left", padx=6, pady=4).pack()
        tip_win[0] = w

    def hide(event):
        if tip_win[0]:
            tip_win[0].destroy()
            tip_win[0] = None

    widget.bind("<Enter>", show)
    widget.bind("<Leave>", hide)

ORDERS_FILE = "orders_history.json"
API_KEY_ENV = "ANTHROPIC_API_KEY"


# ---------------------------------------------------------------------------
# Вспомогательные
# ---------------------------------------------------------------------------

def _load_orders() -> List[Dict]:
    if not os.path.exists(ORDERS_FILE):
        return []
    try:
        with open(ORDERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def _save_orders(orders: List[Dict]):
    with open(ORDERS_FILE, "w", encoding="utf-8") as f:
        json.dump(orders, f, ensure_ascii=False, indent=2)


def _float(s, default=0.0) -> float:
    try:
        return float(str(s).strip().replace(",", "."))
    except Exception:
        return default


# ---------------------------------------------------------------------------
# Главное окно
# ---------------------------------------------------------------------------

class ManagerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1100x720")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

        self.cfg = pm.load_config()
        self._check_excel_on_start()

        self.orders = _load_orders()
        self.api_key = os.environ.get(API_KEY_ENV, self.cfg.get("api_key", ""))

        self.current_parts: List[Dict] = []
        self.current_client = ""
        self._spec_parts: List[Dict] = []
        self._spec_notes: str = ""   # текстовые замечания из спеки

        self._build_ui()
        self._show_screen("history")

    # -----------------------------------------------------------------------
    # Старт: проверка Excel
    # -----------------------------------------------------------------------

    def _check_excel_on_start(self):
        if not pm.excel_files_set(self.cfg):
            self.after(300, self._ask_excel_files)
        else:
            self.cfg, reloaded = pm.reload_excel_if_needed(self.cfg)
            if reloaded:
                pm.save_config(self.cfg)
            self.after(400, self._refresh_material_combos)

    def _ask_excel_files(self):
        win = ctk.CTkToplevel(self)
        win.title("Настройка — файлы данных")
        win.geometry("560x380")
        win.grab_set()

        ctk.CTkLabel(win, text="Укажите Excel-файлы с данными", font=("Segoe UI", 14, "bold")).pack(pady=14)

        fields = {
            "mats":          "Файл матов (Артикул, Ячейка, Цена м², Вес м²):",
            "frame_weight":  "Веса обрамления (Размер, Вес):",
            "angle_weight":  "Веса перфоуголка (Размер, Вес):",
            "bumper_weight": "Веса отбойника (Размер, Вес):",
            "strip_types":   "Типы полос/Название, Ширина, Толщина (необязательно):",
        }
        entries = {}
        for key, label in fields.items():
            row = ctk.CTkFrame(win)
            row.pack(fill="x", padx=16, pady=3)
            ctk.CTkLabel(row, text=label, anchor="w", width=340).pack(side="left")
            e = ctk.CTkEntry(row, width=100)
            e.insert(0, self.cfg["excel_files"].get(key, ""))
            e.pack(side="left", padx=4)
            entries[key] = e
            ctk.CTkButton(row, text="...", width=32,
                          command=lambda k=key, entry=e: self._browse_excel(k, entry)).pack(side="left")

        def save():
            for k, e in entries.items():
                self.cfg["excel_files"][k] = e.get().strip()
            self.cfg = pm.reload_excel(self.cfg)
            pm.save_config(self.cfg)
            win.destroy()
            self._refresh_material_combos()
            messagebox.showinfo("Готово", "Данные загружены успешно!")

        ctk.CTkButton(win, text="Сохранить и загрузить", command=save).pack(pady=12)

    def _browse_excel(self, key: str, entry: ctk.CTkEntry):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            entry.delete(0, "end")
            entry.insert(0, path)

    def _open_settings(self):
        win = ctk.CTkToplevel(self)
        win.title("Настройки")
        win.geometry("520x580")
        win.grab_set()

        ctk.CTkLabel(win, text="Настройки", font=("Segoe UI", 14, "bold")).pack(pady=14)

        # API ключ
        api_frame = ctk.CTkFrame(win, fg_color="transparent")
        api_frame.pack(fill="x", padx=20, pady=6)
        ctk.CTkLabel(api_frame, text="API ключ Claude:", anchor="w").pack(fill="x")
        key_row = ctk.CTkFrame(api_frame, fg_color="transparent")
        key_row.pack(fill="x", pady=(2, 0))
        api_entry = ctk.CTkEntry(key_row, width=360, show="*")
        api_entry.insert(0, self.api_key)
        api_entry.pack(side="left")
        show_var = ctk.BooleanVar(value=False)
        def toggle_show():
            api_entry.configure(show="" if show_var.get() else "*")
        ctk.CTkCheckBox(key_row, text="👁", variable=show_var, width=40,
                        command=toggle_show).pack(side="left", padx=6)

        # Excel файлы
        ctk.CTkButton(win, text="📂 Excel-файлы данных...", width=220,
                      fg_color="transparent", border_width=1,
                      command=lambda: [win.destroy(), self._ask_excel_files()]).pack(pady=8)

        # ── Нормативы трудозатрат ──────────────────────────────────────────
        norms_frame = ctk.CTkFrame(win, border_width=1)
        norms_frame.pack(fill="x", padx=20, pady=(8, 4))
        ctk.CTkLabel(norms_frame, text="Нормативы трудозатрат",
                     font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(8, 4))

        norms_cfg = self.cfg.get("labour_norms", {})
        norm_fields = [
            ("t_straight_h_per_m",   "Прямой рез (ч/м):",         "0.02"),
            ("t_diagonal_h_per_cut", "Косой рез (ч/штуку):",       "0.05"),
            ("t_arc_h_per_m",        "Радиус-дуга (ч/м):",         "0.15"),
            ("complexity_k",         "Коэфф. сложности k:",        "0.30"),
            ("min_hours_per_part",   "Минимум часов на деталь:",    "0.05"),
        ]
        norm_entries = {}
        for key, label, default in norm_fields:
            row = ctk.CTkFrame(norms_frame, fg_color="transparent")
            row.pack(fill="x", padx=12, pady=2)
            ctk.CTkLabel(row, text=label, width=210, anchor="w").pack(side="left")
            e = ctk.CTkEntry(row, width=90)
            e.insert(0, str(norms_cfg.get(key, default)))
            e.pack(side="left")
            norm_entries[key] = e

        ctk.CTkLabel(norms_frame,
                     text="Значения сохраняются и не требуют повторного ввода.",
                     font=("Segoe UI", 10), text_color="gray").pack(anchor="w", padx=12, pady=(2, 8))

        def save():
            self.api_key = api_entry.get().strip()
            self.cfg["api_key"] = self.api_key
            # Сохраняем нормативы
            norms = {}
            for key, e in norm_entries.items():
                try:
                    norms[key] = float(e.get().strip().replace(",", "."))
                except ValueError:
                    pass  # оставляем прежнее значение если введено некорректно
            existing = self.cfg.get("labour_norms", {})
            existing.update(norms)
            self.cfg["labour_norms"] = existing
            pm.save_config(self.cfg)
            win.destroy()
            messagebox.showinfo("Сохранено", "Настройки сохранены.")

        btns = ctk.CTkFrame(win, fg_color="transparent")
        btns.pack(pady=16)
        ctk.CTkButton(btns, text="Сохранить", width=120, command=save).pack(side="left", padx=8)
        ctk.CTkButton(btns, text="Отмена", width=100, fg_color="transparent",
                      border_width=1, command=win.destroy).pack(side="left")

    # -----------------------------------------------------------------------
    # Структура UI
    # -----------------------------------------------------------------------

    def _build_ui(self):
        self.container = ctk.CTkFrame(self)
        self.container.pack(fill="both", expand=True)
        self.screens: Dict[str, ctk.CTkFrame] = {}
        for name in ("history", "upload", "review", "result"):
            frame = ctk.CTkFrame(self.container)
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)
            self.screens[name] = frame

        self._build_history_screen()
        self._build_upload_screen()
        self._build_review_screen()
        self._build_result_screen()

    def _show_screen(self, name: str):
        self.screens[name].lift()

    # -----------------------------------------------------------------------
    # Экран 1 — История
    # -----------------------------------------------------------------------

    def _build_history_screen(self):
        s = self.screens["history"]

        top = ctk.CTkFrame(s, fg_color="transparent")
        top.pack(fill="x", padx=20, pady=16)
        ctk.CTkLabel(top, text="Grid Master — Менеджер", font=("Segoe UI", 20, "bold")).pack(side="left")
        ctk.CTkButton(top, text="+ Новый заказ", width=140, command=self._new_order).pack(side="right")
        ctk.CTkButton(top, text="⚙ Настройки", width=110,
                      fg_color="transparent", border_width=1,
                      command=self._open_settings).pack(side="right", padx=8)

        list_frame = ctk.CTkScrollableFrame(s, label_text="История заказов")
        list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        self._history_frame = list_frame
        self._refresh_history()

    def _refresh_history(self):
        for w in self._history_frame.winfo_children():
            w.destroy()
        if not self.orders:
            ctk.CTkLabel(self._history_frame, text="Заказов пока нет. Нажмите «+ Новый заказ».",
                         text_color="grey").pack(pady=40)
            return
        for order in reversed(self.orders):
            self._history_row(order)

    def _history_row(self, order: Dict):
        row = ctk.CTkFrame(self._history_frame, fg_color=("grey90", "grey20"), corner_radius=8)
        row.pack(fill="x", pady=3, padx=4)
        ctk.CTkLabel(row, text=order.get("date", ""), width=90).pack(side="left", padx=10)
        ctk.CTkLabel(row, text=order.get("client", ""), width=220, anchor="w").pack(side="left")
        price = order.get("client_price", 0)
        ctk.CTkLabel(row, text=f"{price:,.2f} €".replace(",", " "),
                     font=("Segoe UI", 12, "bold"), width=120).pack(side="left")
        ctk.CTkButton(row, text="Открыть ▸", width=90,
                      command=lambda o=order: self._open_order(o)).pack(side="right", padx=8, pady=6)

    def _new_order(self):
        self.current_parts = []
        self.current_client = ""
        self._spec_parts = []
        self._spec_notes = ""
        self._client_entry.delete(0, "end")
        self._files_label.configure(text="Файлы не выбраны")
        self._pdf_paths = []
        self._spec_path_str.set("")
        self._spec_label.configure(text="не выбрана", text_color="grey")
        self._update_to_review_btn()
        self._show_screen("upload")

    def _open_order(self, order: Dict):
        self.current_parts  = order.get("parts", [])
        self.current_client = order.get("client", "")
        self._populate_review()
        self._show_screen("review")

    # -----------------------------------------------------------------------
    # Экран 2 — Загрузка чертежей
    # -----------------------------------------------------------------------

    def _build_upload_screen(self):
        s = self.screens["upload"]
        self._pdf_paths: List[str] = []
        self._spec_path_str = ctk.StringVar(value="")

        top = ctk.CTkFrame(s, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=12)
        ctk.CTkButton(top, text="← Назад", width=80, fg_color="transparent",
                      command=lambda: self._show_screen("history")).pack(side="left")
        ctk.CTkButton(top, text="⚙ Настройки", width=110, fg_color="transparent",
                      border_width=1, command=self._open_settings).pack(side="right")

        ctk.CTkLabel(s, text="Новый заказ", font=("Segoe UI", 18, "bold")).pack(pady=(0, 12))

        client_row = ctk.CTkFrame(s, fg_color="transparent")
        client_row.pack(pady=6)
        ctk.CTkLabel(client_row, text="Клиент:", width=70).pack(side="left")
        self._client_entry = ctk.CTkEntry(client_row, width=300, placeholder_text="Название компании / имя")
        self._client_entry.pack(side="left", padx=8)

        drop = ctk.CTkFrame(s, width=500, height=140, fg_color=("grey85", "grey25"),
                            corner_radius=16, border_width=2, border_color=("grey70", "grey40"))
        drop.pack(pady=12)
        drop.pack_propagate(False)
        ctk.CTkLabel(drop, text="📄  Перетащи PDF сюда\nили нажми для выбора",
                     font=("Segoe UI", 13), text_color="grey").pack(expand=True)
        drop.bind("<Button-1>", lambda e: self._choose_files())

        self._files_label = ctk.CTkLabel(s, text="Файлы не выбраны", text_color="grey")
        self._files_label.pack()

        spec_row = ctk.CTkFrame(s, fg_color="transparent")
        spec_row.pack(pady=(8, 0))
        ctk.CTkLabel(spec_row, text="Спека заказчика:", width=130, anchor="w").pack(side="left")
        self._spec_label = ctk.CTkLabel(spec_row, text="не выбрана", text_color="grey", width=220, anchor="w")
        self._spec_label.pack(side="left")
        ctk.CTkButton(spec_row, text="📋 Выбрать", width=110,
                      fg_color="transparent", border_width=1,
                      command=self._choose_spec).pack(side="left", padx=6)
        ctk.CTkButton(spec_row, text="✕", width=28, fg_color="transparent",
                      text_color="grey", command=self._clear_spec).pack(side="left")

        btn_row = ctk.CTkFrame(s, fg_color="transparent")
        btn_row.pack(pady=14)
        ctk.CTkButton(btn_row, text="Выбрать файлы", width=160,
                      fg_color="transparent", border_width=1,
                      command=self._choose_files).pack(side="left", padx=8)
        self._recog_btn = ctk.CTkButton(btn_row, text="Распознать →", width=160,
                                         command=self._start_recognition)
        self._recog_btn.pack(side="left", padx=8)

        self._recog_status = ctk.CTkLabel(s, text="", text_color="grey")
        self._recog_status.pack()

        # Кнопка «К проверке» — видна только если уже есть распознанные детали
        self._to_review_btn = ctk.CTkButton(
            s, text="→ К проверке позиций", width=220,
            fg_color=("#2E6B9E", "#1a4a70"),
            command=lambda: self._show_screen("review"))
        # показывается/скрывается через _update_to_review_btn()

    def _choose_files(self):
        paths = filedialog.askopenfilenames(
            title="Выберите чертежи",
            filetypes=[("PDF", "*.pdf"), ("Все файлы", "*.*")]
        )
        if paths:
            self._pdf_paths = list(paths)
            names = ", ".join(os.path.basename(p) for p in paths)
            self._files_label.configure(text=names, text_color="white")

    def _choose_spec(self):
        path = filedialog.askopenfilename(
            title="Спецификация заказчика",
            filetypes=[("Word/Excel/PDF", "*.docx *.doc *.xlsx *.xls *.pdf"), ("Все файлы", "*.*")]
        )
        if path:
            self._spec_path_str.set(path)
            self._spec_label.configure(text=os.path.basename(path), text_color="white")

    def _clear_spec(self):
        self._spec_path_str.set("")
        self._spec_label.configure(text="не выбрана", text_color="grey")
        self._spec_parts = []

    def _start_recognition(self):
        if not self._pdf_paths:
            messagebox.showwarning("Нет файлов", "Сначала выберите PDF-чертежи.")
            return
        if not self.api_key:
            messagebox.showwarning("Нет API ключа", "Введите API ключ в Настройках (⚙).")
            return
        self.cfg["api_key"] = self.api_key
        pm.save_config(self.cfg)
        self.current_client = self._client_entry.get().strip()
        self._recog_btn.configure(state="disabled", text="Распознаю...")
        self._recog_status.configure(text="")

        def work():
            from concurrent.futures import ThreadPoolExecutor, as_completed
            import traceback

            results = {}
            errors = []
            total = len(self._pdf_paths)
            done_count = [0]

            def process_one(path):
                return path, ai.recognize_pdf(path, self.api_key)

            with ThreadPoolExecutor(max_workers=min(total, 8)) as executor:
                futures = {executor.submit(process_one, p): p for p in self._pdf_paths}
                for future in as_completed(futures):
                    path = futures[future]
                    try:
                        _, parts = future.result()
                        results[path] = parts
                    except Exception:
                        errors.append(f"{os.path.basename(path)}: {traceback.format_exc()}")
                        results[path] = []
                    done_count[0] += 1
                    n_done = done_count[0]
                    self.after(0, lambda d=n_done: self._recog_status.configure(
                        text=f"Обрабатываю: {d}/{total}...", text_color="grey"))

            all_parts = []
            for path in self._pdf_paths:
                all_parts.extend(results.get(path, []))

            self.current_parts = all_parts
            if errors:
                err_text = "\n\n".join(errors)
                self.after(0, lambda t=err_text: messagebox.showerror("Ошибка распознавания", t))
            self.after(0, self._on_recognition_done)

        threading.Thread(target=work, daemon=True).start()

    def _update_to_review_btn(self):
        """Показывает кнопку «→ К проверке» если есть распознанные позиции."""
        if self.current_parts:
            self._to_review_btn.pack(pady=(4, 8))
        else:
            self._to_review_btn.pack_forget()

    def _recalc_all_work(self):
        """Пересчитывает нормо-часы для ВСЕХ текущих деталей по актуальным нормативам."""
        if not self.current_parts:
            return
        norms = self.cfg.get("labour_norms", {})
        for p in self.current_parts:
            p["work"] = 0.0  # сбросить, чтобы _auto_fill_work пересчитал
        self._auto_fill_work(self.current_parts)
        self._populate_review()
        messagebox.showinfo("Готово", f"Трудозатраты пересчитаны для {len(self.current_parts)} позиций.")

    def _auto_fill_work(self, parts: list) -> None:
        """Заполняет поле work для деталей, у которых оно равно 0 или не задано."""
        norms = self.cfg.get("labour_norms", {})
        for p in parts:
            if not p.get("work"):
                tmp = core.Part(
                    position=str(p.get("position", "")),
                    length=float(p.get("length", 0) or 0),
                    width=float(p.get("width", 0) or 0),
                    quantity=int(p.get("quantity", 1) or 1),
                    angle=str(p.get("angle", "0")),
                    bumper=float(p.get("bumper", 0) or 0),
                    work=0.0,
                    radius=float(p.get("radius", 0) or 0),
                    radius_part=float(p.get("radius_part", 1.0) or 1.0),
                    length2=float(p.get("length2", 0) or 0),
                )
                p["work"] = core.calc_work_hours(tmp, norms)

    def _on_recognition_done(self):
        self._recog_btn.configure(state="normal", text="Распознать →")
        if not self.current_parts:
            self._recog_status.configure(
                text="Не удалось распознать позиции. Добавьте вручную.", text_color="orange")
        else:
            n = len(self.current_parts)
            self._recog_status.configure(text=f"Распознано {n} поз.", text_color="green")

        spec_path = self._spec_path_str.get()
        if spec_path and os.path.exists(spec_path):
            try:
                raw = spec_reader.read_spec(spec_path, self.api_key)
                # Отделяем текстовые замечания от позиций
                self._spec_parts = [p for p in raw if p.get("position")]
                notes_list = [p.get("notes", "") for p in raw if p.get("notes")]
                self._spec_notes = "\n".join(notes_list)
                if self._spec_parts and self.current_parts:
                    self.current_parts = spec_reader.compare(self.current_parts, self._spec_parts)
            except Exception as e:
                messagebox.showwarning("Спека", f"Не удалось прочитать спеку:\n{e}")

        self._auto_fill_work(self.current_parts)
        self._update_to_review_btn()
        self._populate_review()
        self._show_screen("review")

    # -----------------------------------------------------------------------
    # Экран 3 — Проверка позиций и параметры
    # -----------------------------------------------------------------------

    def _build_review_screen(self):
        s = self.screens["review"]

        top = ctk.CTkFrame(s, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=10)
        ctk.CTkButton(top, text="← Назад", width=80, fg_color="transparent",
                      command=lambda: self._show_screen("upload")).pack(side="left")
        ctk.CTkLabel(top, text="Проверка позиций", font=("Segoe UI", 16, "bold")).pack(side="left", padx=12)
        ctk.CTkButton(top, text="Рассчитать →", width=140,
                      fg_color=("#2E8B57", "#2E8B50"),
                      command=self._calculate).pack(side="right")
        ctk.CTkButton(top, text="↻ Трудозатраты", width=130, fg_color="transparent",
                      border_width=1,
                      command=self._recalc_all_work).pack(side="right", padx=6)
        ctk.CTkButton(top, text="⚙ Настройки", width=110, fg_color="transparent",
                      border_width=1,
                      command=self._open_settings).pack(side="right", padx=4)

        main = ctk.CTkFrame(s, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=16, pady=(0, 8))
        main.columnconfigure(0, weight=3)
        main.columnconfigure(1, weight=2)
        main.rowconfigure(0, weight=1)

        # --- Левая колонка: позиции ---
        left = ctk.CTkFrame(main)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        parts_top = ctk.CTkFrame(left, fg_color="transparent")
        parts_top.pack(fill="x", padx=8, pady=6)
        ctk.CTkLabel(parts_top, text="Позиции", font=("Segoe UI", 12, "bold")).pack(side="left")
        ctk.CTkButton(parts_top, text="+ Добавить", width=90,
                      command=self._add_part_manual).pack(side="right")
        self._apply_spec_btn = ctk.CTkButton(
            parts_top, text="✓ Принять все из спеки", width=190,
            fg_color=("grey70", "grey30"), command=self._apply_spec_quantities)
        self._apply_spec_btn.pack(side="right", padx=6)

        self._parts_scroll = ctk.CTkScrollableFrame(left)
        self._parts_scroll.pack(fill="both", expand=True, padx=6, pady=(0, 6))

        # Панель замечаний из спеки (скрыта если пусто)
        self._notes_frame = ctk.CTkFrame(left, fg_color=("grey85", "grey20"))
        self._notes_header = ctk.CTkFrame(self._notes_frame, fg_color="transparent")
        self._notes_header.pack(fill="x", padx=8, pady=(6, 0))
        ctk.CTkLabel(self._notes_header, text="📝 Замечания из спеки:", font=("Segoe UI", 10, "bold")).pack(side="left")
        ctk.CTkButton(self._notes_header, text="✕", width=28, fg_color="transparent",
                      text_color="grey", command=lambda: self._notes_frame.pack_forget()).pack(side="right")
        self._notes_text = ctk.CTkTextbox(self._notes_frame, height=70, state="disabled",
                                           font=("Segoe UI", 10))
        self._notes_text.pack(fill="x", padx=8, pady=6)

        # --- Правая колонка: материалы и цены ---
        right = ctk.CTkScrollableFrame(main, label_text="Материалы и цены")
        right.grid(row=0, column=1, sticky="nsew")
        self._params_frame = right
        self._build_params_panel(right)

    def _build_params_panel(self, parent):
        def combo_row(label, combo):
            r = ctk.CTkFrame(parent, fg_color="transparent")
            r.pack(fill="x", padx=8, pady=(4, 0))
            ctk.CTkLabel(r, text=label, anchor="w", font=("Segoe UI", 10)).pack(fill="x")
            combo.pack(fill="x", pady=(0, 2))

        ctk.CTkLabel(parent, text="Материалы", font=("Segoe UI", 11, "bold")).pack(
            anchor="w", padx=8, pady=(10, 2))

        mat_articles = [m["article"] for m in self.cfg.get("mat_data", [])]
        self._mat_cb = ctk.CTkComboBox(parent, values=mat_articles or ["—"],
                                        command=self._on_mat_changed)
        last = self.cfg.get("last_mat_article", "")
        if last in mat_articles:   self._mat_cb.set(last)
        elif mat_articles:         self._mat_cb.set(mat_articles[0])
        combo_row("Мат (артикул):", self._mat_cb)

        frame_types = [t for t, _ in self.cfg.get("frame_weights", [])]
        self._frame_cb = ctk.CTkComboBox(parent, values=frame_types or ["—"],
                                          command=lambda v: self._on_material_type_changed("frame", v))
        last_f = self.cfg.get("last_frame_type", "")
        if last_f in frame_types:  self._frame_cb.set(last_f)
        elif frame_types:          self._frame_cb.set(frame_types[0])
        combo_row("Обрамление:", self._frame_cb)

        angle_types = [t for t, _ in self.cfg.get("angle_weights", [])]
        self._angle_cb = ctk.CTkComboBox(parent, values=angle_types or ["—"],
                                          command=lambda v: self._on_material_type_changed("angle", v))
        last_a = self.cfg.get("last_angle_type", "")
        if last_a in angle_types:  self._angle_cb.set(last_a)
        elif angle_types:          self._angle_cb.set(angle_types[0])
        combo_row("Перфоуголок:", self._angle_cb)

        bumper_types = [t for t, _ in self.cfg.get("bumper_weights", [])]
        self._bumper_cb = ctk.CTkComboBox(parent, values=bumper_types or ["—"],
                                           command=lambda v: self._on_material_type_changed("bumper", v))
        last_b = self.cfg.get("last_bumper_type", "")
        if last_b in bumper_types: self._bumper_cb.set(last_b)
        elif bumper_types:         self._bumper_cb.set(bumper_types[0])
        combo_row("Отбойник:", self._bumper_cb)

        ctk.CTkLabel(parent, text="Цены (редактируемые)", font=("Segoe UI", 11, "bold")).pack(
            anchor="w", padx=8, pady=(12, 2))

        self._price_entries: Dict[str, ctk.CTkEntry] = {}
        price_labels = {
            "mat":    "Мат €/м²:",
            "frame":  "Обрамление €/кг:",
            "angle":  "Перфоуголок €/шт:",
            "bumper": "Отбойник €/кг:",
            "work":   "Работа €/час:",
            "zinc":   "Цинк €/кг:",
        }
        for key, label in price_labels.items():
            r = ctk.CTkFrame(parent, fg_color="transparent")
            r.pack(fill="x", padx=8, pady=2)
            ctk.CTkLabel(r, text=label, width=140, anchor="w").pack(side="left")
            e = ctk.CTkEntry(r, width=90)
            val = pm.get_price(self.cfg, key)
            e.insert(0, f"{val:.2f}" if val else "")
            e.pack(side="left")
            e.bind("<FocusOut>", lambda ev, k=key: self._on_price_edited(k))
            self._price_entries[key] = e

        ctk.CTkLabel(parent, text="Прочее", font=("Segoe UI", 11, "bold")).pack(
            anchor="w", padx=8, pady=(12, 2))

        r = ctk.CTkFrame(parent, fg_color="transparent")
        r.pack(fill="x", padx=8, pady=2)
        ctk.CTkLabel(r, text="Коэф. обрамления:", width=140, anchor="w").pack(side="left")
        self._frame_coef_entry = ctk.CTkEntry(r, width=60)
        self._frame_coef_entry.insert(0, "1.0")
        self._frame_coef_entry.pack(side="left")

        r2 = ctk.CTkFrame(parent, fg_color="transparent")
        r2.pack(fill="x", padx=8, pady=2)
        ctk.CTkLabel(r2, text="Коэф. работы:", width=140, anchor="w").pack(side="left")
        self._work_coef_entry = ctk.CTkEntry(r2, width=60)
        self._work_coef_entry.insert(0, "1.0")
        self._work_coef_entry.pack(side="left")

    def _refresh_material_combos(self):
        """Обновить выпадающие списки материалов после загрузки Excel."""
        mat_articles = [m["article"] for m in self.cfg.get("mat_data", [])]
        if mat_articles:
            self._mat_cb.configure(values=mat_articles)
            last = self.cfg.get("last_mat_article", "")
            self._mat_cb.set(last if last in mat_articles else mat_articles[0])

        frame_types = [t for t, _ in self.cfg.get("frame_weights", [])]
        if frame_types:
            self._frame_cb.configure(values=frame_types)
            last_f = self.cfg.get("last_frame_type", "")
            self._frame_cb.set(last_f if last_f in frame_types else frame_types[0])

        angle_types = [t for t, _ in self.cfg.get("angle_weights", [])]
        if angle_types:
            self._angle_cb.configure(values=angle_types)
            last_a = self.cfg.get("last_angle_type", "")
            self._angle_cb.set(last_a if last_a in angle_types else angle_types[0])

        bumper_types = [t for t, _ in self.cfg.get("bumper_weights", [])]
        if bumper_types:
            self._bumper_cb.configure(values=bumper_types)
            last_b = self.cfg.get("last_bumper_type", "")
            self._bumper_cb.set(last_b if last_b in bumper_types else bumper_types[0])

        # Подставить цену мата из данных
        if mat_articles:
            self._on_mat_changed(self._mat_cb.get())

        # Обновить поля цен из конфига
        for key, e in self._price_entries.items():
            val = pm.get_price(self.cfg, key)
            e.delete(0, "end")
            e.insert(0, f"{val:.2f}" if val else "")

        # Подставить сохранённые цены для выбранных типов
        for key, cb_attr in [("frame", "_frame_cb"), ("angle", "_angle_cb"), ("bumper", "_bumper_cb")]:
            cb = getattr(self, cb_attr, None)
            if cb:
                self._on_material_type_changed(key, cb.get())

    def _on_mat_changed(self, article: str):
        spec = pm.get_mat_spec(self.cfg, article)
        if spec:
            e = self._price_entries.get("mat")
            if e:
                e.delete(0, "end")
                e.insert(0, f"{spec['price_per_m2']:.2f}")
            # Сбрасываем manual_override — артикул сменился, цена из таблицы актуальна
            self.cfg["prices"]["mat"]["manual_override"] = False
            self.cfg["last_mat_article"] = article

    def _on_material_type_changed(self, key: str, type_name: str):
        """Смена типа обрамления/перфоуголка/отбойника — подставить сохранённую цену."""
        saved = self.cfg.get("type_prices", {}).get(key, {})
        if type_name in saved:
            e = self._price_entries.get(key)
            if e:
                e.delete(0, "end")
                e.insert(0, f"{saved[type_name]:.2f}")

    def _on_price_edited(self, key: str):
        e = self._price_entries.get(key)
        if not e:
            return
        val = _float(e.get())
        pm.set_price(self.cfg, key, val, manual=True)
        # Запомнить цену для текущего типа (для frame/angle/bumper)
        cb_map = {"frame": "_frame_cb", "angle": "_angle_cb", "bumper": "_bumper_cb"}
        if key in cb_map:
            cb = getattr(self, cb_map[key], None)
            if cb:
                type_name = cb.get()
                if "type_prices" not in self.cfg:
                    self.cfg["type_prices"] = {}
                if key not in self.cfg["type_prices"]:
                    self.cfg["type_prices"][key] = {}
                self.cfg["type_prices"][key][type_name] = val
                pm.save_config(self.cfg)

    def _populate_review(self):
        for w in self._parts_scroll.winfo_children():
            w.destroy()
        for i, part in enumerate(self.current_parts):
            self._part_row(i, part)

        # Показать/скрыть панель замечаний
        notes = getattr(self, "_spec_notes", "")
        if notes:
            self._notes_text.configure(state="normal")
            self._notes_text.delete("1.0", "end")
            self._notes_text.insert("1.0", notes)
            self._notes_text.configure(state="disabled")
            self._notes_frame.pack(fill="x", padx=6, pady=(0, 4))
        else:
            self._notes_frame.pack_forget()

    def _part_row(self, idx: int, part: Dict):
        conf = part.get("confidence", "high")
        has_conflict = bool(part.get("quantity_conflict"))
        only_spec = part.get("only_in_spec", False)

        if only_spec:
            bg = ("lightcyan", "#003344")
        elif has_conflict:
            bg = ("#fff3cd", "#4a3800")
        elif conf == "medium":
            bg = ("#fff8dc", "#4a4000")
        elif conf == "low":
            bg = ("#ffe0cc", "#5a2000")
        else:
            bg = ("grey90", "grey22")

        row = ctk.CTkFrame(self._parts_scroll, fg_color=bg, corner_radius=6)
        row.pack(fill="x", pady=2, padx=2)

        icon = "🔵" if only_spec else ("⚠" if (has_conflict or conf != "high") else "✓")
        ctk.CTkLabel(row, text=icon, width=24).pack(side="left", padx=4)
        ctk.CTkLabel(row, text=str(part.get("position", "")), width=160, anchor="w").pack(side="left")
        l2 = part.get("length2", 0) or 0
        size_str = (f"{part.get('length',0)}/{l2} × {part.get('width',0)} мм"
                    if l2 else f"{part.get('length',0)} × {part.get('width',0)} мм")
        ctk.CTkLabel(row, text=size_str, width=140).pack(side="left")

        # Количество + кнопка принять из спеки
        if has_conflict:
            qc = part["quantity_conflict"]
            qty_text = f"× {part.get('quantity',1)}"
            qty_lbl = ctk.CTkLabel(row, text=qty_text, width=60)
            qty_lbl.configure(text_color=("#b85c00", "#ffaa44"))
            qty_lbl.pack(side="left")
            # Кнопка "принять из спеки" прямо в строке
            spec_qty = qc["spec"]
            ctk.CTkButton(row, text=f"✓ {spec_qty} из спеки", width=120, height=24,
                          fg_color=("#2E6B9E", "#1a4a70"),
                          command=lambda i=idx, q=spec_qty: self._accept_spec_qty(i, q)
                          ).pack(side="left", padx=4)
        else:
            qty_text = f"× {part.get('quantity', 1)}"
            ctk.CTkLabel(row, text=qty_text, width=160).pack(side="left")

        ctk.CTkButton(row, text="✏", width=30, fg_color="transparent",
                      command=lambda i=idx: self._edit_part(i)).pack(side="right", padx=4)
        ctk.CTkButton(row, text="✕", width=30, fg_color="transparent", text_color="red",
                      command=lambda i=idx: self._delete_part(i)).pack(side="right")

        notes = part.get("notes", "")
        diffs = part.get("diff", [])
        tooltip_parts = []
        if notes:
            tooltip_parts.append(notes)
        if diffs:
            tooltip_parts.append("Расхождения: " + "; ".join(diffs))
        if tooltip_parts:
            tip_text = "\n".join(tooltip_parts)
            icon_text = "⚠💬" if diffs else "💬"
            note_lbl = ctk.CTkLabel(row, text=icon_text, width=30, cursor="hand2")
            note_lbl.pack(side="left", padx=(4, 0))
            _bind_tooltip(note_lbl, tip_text)

    def _accept_spec_qty(self, idx: int, qty: int):
        """Принять количество из спеки для одной позиции."""
        part = self.current_parts[idx]
        part["quantity"] = qty
        part.pop("quantity_conflict", None)
        part.pop("diff", None)
        self._populate_review()

    def _add_part_manual(self):
        part = {"position": "", "length": 0, "width": 0, "quantity": 1,
                "work": 0.0, "notes": "", "confidence": "high"}
        self.current_parts.append(part)
        self._edit_part(len(self.current_parts) - 1)

    def _edit_part(self, idx: int):
        part = self.current_parts[idx]
        win = ctk.CTkToplevel(self)
        win.title("Редактировать позицию")
        win.geometry("440x420")
        win.grab_set()
        win.resizable(False, False)

        # Поля без "Работа" — они рендерятся отдельно с кнопкой пересчёта
        fields = [
            ("Позиция/обозначение",    "position",    str(part.get("position", ""))),
            ("Длина (мм)",             "length",       str(part.get("length", ""))),
            ("Длина 2 (мм, трапеция)", "length2",      str(part.get("length2", "") or "")),
            ("Ширина (мм)",            "width",        str(part.get("width", ""))),
            ("Количество",             "quantity",     str(part.get("quantity", 1))),
            ("Радиус (мм)",            "radius",       str(part.get("radius", ""))),
            ("Доля окружности",        "radius_part",  str(part.get("radius_part", "1.0"))),
            ("Отбойник (мм)",          "bumper",       str(part.get("bumper", ""))),
            ("Примечание",             "notes",        str(part.get("notes", ""))),
        ]
        entries = {}
        first_entry = [None]
        for label, key, val in fields:
            r = ctk.CTkFrame(win, fg_color="transparent")
            r.pack(fill="x", padx=16, pady=3)
            ctk.CTkLabel(r, text=label, width=170, anchor="w").pack(side="left")
            e = ctk.CTkEntry(r, width=180)
            e.insert(0, val)
            e.pack(side="left")
            entries[key] = e
            if first_entry[0] is None:
                first_entry[0] = e

        # Строка "Работа" — поле + кнопка пересчёта
        work_row = ctk.CTkFrame(win, fg_color="transparent")
        work_row.pack(fill="x", padx=16, pady=3)
        ctk.CTkLabel(work_row, text="Работа (часов):", width=170, anchor="w").pack(side="left")
        work_entry = ctk.CTkEntry(work_row, width=110)
        work_entry.insert(0, str(part.get("work", "")))
        work_entry.pack(side="left")
        entries["work"] = work_entry

        def _recalc_work():
            """Пересчитать нормо-часы по текущим значениям полей."""
            try:
                tmp = core.Part(
                    position=entries["position"].get(),
                    length=float(entries["length"].get() or 0),
                    width=float(entries["width"].get() or 0),
                    quantity=int(float(entries["quantity"].get() or 1)),
                    angle=str(part.get("angle", "0")),
                    bumper=float(entries["bumper"].get() or 0),
                    work=0.0,
                    radius=float(entries["radius"].get() or 0),
                    radius_part=float(entries["radius_part"].get() or 1.0),
                    length2=float(entries["length2"].get() or 0),
                )
                h = core.calc_work_hours(tmp, self.cfg.get("labour_norms", {}))
                work_entry.delete(0, "end")
                work_entry.insert(0, str(h))
            except Exception:
                pass

        ctk.CTkButton(work_row, text="↻", width=36, height=28,
                      fg_color="transparent", border_width=1,
                      command=_recalc_work).pack(side="left", padx=(6, 0))

        def save():
            for key, e in entries.items():
                v = e.get().strip()
                if key in ("length", "length2", "width", "quantity"):
                    part[key] = int(_float(v) or 0)
                elif key in ("radius", "bumper", "work"):
                    part[key] = _float(v)
                elif key == "radius_part":
                    part[key] = _float(v) or 1.0
                else:
                    part[key] = v
            part["confidence"] = "high"
            part.pop("quantity_conflict", None)
            part.pop("diff", None)
            self.current_parts[idx] = part
            self._populate_review()
            win.destroy()

        def cancel():
            # Если позиция новая (пустая) — удалить её
            if not part.get("position") and part.get("length", 0) == 0:
                if idx < len(self.current_parts) and self.current_parts[idx] is part:
                    self.current_parts.pop(idx)
                    self._populate_review()
            win.destroy()

        btn_row = ctk.CTkFrame(win, fg_color="transparent")
        btn_row.pack(pady=12)
        save_btn = ctk.CTkButton(btn_row, text="Сохранить", width=130, command=save)
        save_btn.pack(side="left", padx=8)
        ctk.CTkButton(btn_row, text="Отмена", width=100, fg_color="transparent",
                      border_width=1, command=cancel).pack(side="left")

        # Enter = сохранить, Escape = отмена
        win.bind("<Return>", lambda e: save())
        win.bind("<Escape>", lambda e: cancel())

        if first_entry[0]:
            first_entry[0].focus_set()

    def _delete_part(self, idx: int):
        self.current_parts.pop(idx)
        self._populate_review()

    def _apply_spec_quantities(self):
        """Применяет количества из спеки ко всем позициям с конфликтом."""
        self.current_parts = spec_reader.apply_spec_quantities(self.current_parts)
        for p in self.current_parts:
            p.pop("quantity_conflict", None)
            p.pop("diff", None)
        self._populate_review()

    # -----------------------------------------------------------------------
    # Расчёт
    # -----------------------------------------------------------------------

    def _build_order(self) -> core.Order:
        parts = []
        for p in self.current_parts:
            parts.append(core.Part(
                position=str(p.get("position", "")),
                length=float(p.get("length", 0)),
                length2=float(p.get("length2", 0) or 0),
                width=float(p.get("width", 0)),
                quantity=int(p.get("quantity", 1)),
                angle=str(p.get("angle", "0")),
                bumper=float(p.get("bumper", 0)),
                work=float(p.get("work", 0)),
                radius=float(p.get("radius", 0)),
                radius_part=float(p.get("radius_part", 1.0)),
                subtract_bumper=bool(p.get("subtract_bumper", True)),
                bumper_length=float(p.get("bumper_length", 0)),
            ))

        article = self._mat_cb.get()
        spec = pm.get_mat_spec(self.cfg, article)
        mat = core.MatSpec(
            article=article,
            price_per_m2=_float(self._price_entries["mat"].get()),
            weight_per_m2=spec["weight_per_m2"] if spec else 0.0,
        )

        return core.Order(
            parts=parts,
            mat=mat,
            frame_type=self._frame_cb.get(),
            frame_price=_float(self._price_entries["frame"].get()),
            frame_coef=_float(self._frame_coef_entry.get()) or 1.0,
            frame_weights=self.cfg.get("frame_weights", []),
            angle_type=self._angle_cb.get(),
            angle_price=_float(self._price_entries["angle"].get()),
            angle_weights=self.cfg.get("angle_weights", []),
            bumper_type=self._bumper_cb.get(),
            bumper_price=_float(self._price_entries["bumper"].get()),
            bumper_weights=self.cfg.get("bumper_weights", []),
            work_price=_float(self._price_entries["work"].get()),
            work_coef=_float(self._work_coef_entry.get()) or 1.0,
            zinc_price=_float(self._price_entries["zinc"].get()),
        )

    def _calculate(self):
        if not self.current_parts:
            messagebox.showwarning("Нет позиций", "Добавьте хотя бы одну позицию.")
            return
        try:
            order = self._build_order()
            self._last_result = core.calculate(order)
            self._show_result()
        except Exception as e:
            messagebox.showerror("Ошибка расчёта", str(e))

    # -----------------------------------------------------------------------
    # Экран 4 — Результат
    # -----------------------------------------------------------------------

    def _build_result_screen(self):
        s = self.screens["result"]

        top = ctk.CTkFrame(s, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=10)
        ctk.CTkButton(top, text="← Назад", width=80, fg_color="transparent",
                      command=lambda: self._show_screen("review")).pack(side="left")
        ctk.CTkLabel(top, text="Результат", font=("Segoe UI", 16, "bold")).pack(side="left", padx=12)

        body = ctk.CTkFrame(s, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=16)
        body.columnconfigure(0, weight=2)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        left = ctk.CTkFrame(body)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        ctk.CTkLabel(left, text="Себестоимость", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=8)

        self._cost_labels: Dict[str, ctk.CTkLabel] = {}
        rows_cfg = [
            ("mat",    "Маты (+15%):"),
            ("frame",  "Обрамление:"),
            ("angle",  "Перфоуголок:"),
            ("bumper", "Отбойник:"),
            ("work",   "Работа:"),
            ("zinc",   "Цинкование:"),
            ("total",  "ИТОГО себест.:"),
        ]
        for key, label in rows_cfg:
            r = ctk.CTkFrame(left, fg_color="transparent")
            r.pack(fill="x", padx=12, pady=2)
            font = ("Segoe UI", 11, "bold") if key == "total" else ("Segoe UI", 11)
            ctk.CTkLabel(r, text=label, width=160, anchor="w", font=font).pack(side="left")
            lbl = ctk.CTkLabel(r, text="—", width=100, anchor="e", font=font)
            lbl.pack(side="right")
            self._cost_labels[key] = lbl

        ctk.CTkLabel(left, text="Состав заказа", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(14,4))
        r0 = ctk.CTkFrame(left, fg_color="transparent")
        r0.pack(fill="x", padx=12, pady=2)
        ctk.CTkLabel(r0, text="Позиций (строк):", width=160, anchor="w").pack(side="left")
        self._lbl_positions_count = ctk.CTkLabel(r0, text="—", width=100, anchor="e")
        self._lbl_positions_count.pack(side="right")
        r0b = ctk.CTkFrame(left, fg_color="transparent")
        r0b.pack(fill="x", padx=12, pady=2)
        ctk.CTkLabel(r0b, text="Деталей (штук):", width=160, anchor="w").pack(side="left")
        self._lbl_parts_total = ctk.CTkLabel(r0b, text="—", width=100, anchor="e")
        self._lbl_parts_total.pack(side="right")

        r0c = ctk.CTkFrame(left, fg_color="transparent")
        r0c.pack(fill="x", padx=12, pady=2)
        ctk.CTkLabel(r0c, text="Матов (+15%), листов:", width=160, anchor="w").pack(side="left")
        self._lbl_mats_count = ctk.CTkLabel(r0c, text="—", width=100, anchor="e")
        self._lbl_mats_count.pack(side="right")

        r0d = ctk.CTkFrame(left, fg_color="transparent")
        r0d.pack(fill="x", padx=12, pady=2)
        ctk.CTkLabel(r0d, text="Перфоуголок (3м), шт:", width=160, anchor="w").pack(side="left")
        self._lbl_angle_pcs = ctk.CTkLabel(r0d, text="—", width=100, anchor="e")
        self._lbl_angle_pcs.pack(side="right")

        ctk.CTkLabel(left, text="Веса", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(14,4))
        self._weight_labels: Dict[str, ctk.CTkLabel] = {}
        for key, label in [("w_mats","Решётка, кг:"), ("w_frame","Обрамление, кг:"),
                            ("w_total","Общий вес, кг:"), ("w_zinc","С цинком (+10%), кг:")]:
            r = ctk.CTkFrame(left, fg_color="transparent")
            r.pack(fill="x", padx=12, pady=2)
            ctk.CTkLabel(r, text=label, width=160, anchor="w").pack(side="left")
            lbl = ctk.CTkLabel(r, text="—", width=100, anchor="e")
            lbl.pack(side="right")
            self._weight_labels[key] = lbl

        right = ctk.CTkFrame(body)
        right.grid(row=0, column=1, sticky="nsew")

        ctk.CTkLabel(right, text="Цена клиенту", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=8)

        ctk.CTkLabel(right, text="Маржа %:", anchor="w").pack(anchor="w", padx=12)
        self._margin_var = ctk.StringVar(value=str(self.cfg.get("margin_pct", 35)))
        self._margin_entry = ctk.CTkEntry(right, width=100, textvariable=self._margin_var)
        self._margin_entry.pack(anchor="w", padx=12, pady=4)
        self._margin_var.trace_add("write", lambda *_: self._update_client_price())

        self._client_price_label = ctk.CTkLabel(right, text="— €",
            font=("Segoe UI", 28, "bold"), text_color=("#2E8B57", "#3CB371"))
        self._client_price_label.pack(pady=20)

        self._cost_per_kg_label = ctk.CTkLabel(right, text="", text_color="grey")
        self._cost_per_kg_label.pack()

        btns = ctk.CTkFrame(right, fg_color="transparent")
        btns.pack(pady=20)
        ctk.CTkButton(btns, text="📄 PDF оффер", width=140,
                      fg_color=("#2E8B57", "#2E8B50"),
                      command=self._export_pdf).pack(pady=6)
        ctk.CTkButton(btns, text="📊 Экспорт в Excel", width=140,
                      command=self._export_xlsx).pack(pady=6)
        ctk.CTkButton(btns, text="✂️ Раскрой матов", width=140,
                      command=self._cutting_dialog).pack(pady=6)
        ctk.CTkButton(btns, text="💾 Сохранить заказ", width=140,
                      command=self._save_order).pack(pady=6)

    def _show_result(self):
        r = self._last_result
        positions = len(self.current_parts)
        total_units = sum(int(p.get("quantity", 1)) for p in self.current_parts)
        self._lbl_positions_count.configure(text=str(positions))
        self._lbl_parts_total.configure(text=str(total_units))
        self._lbl_mats_count.configure(text=str(r.mats_count_with_k))
        import math as _math
        angle_pcs = _math.ceil(r.angle_length_m / 3.0) if r.angle_length_m > 0 else 0
        self._lbl_angle_pcs.configure(text=str(angle_pcs))
        self._cost_labels["mat"].configure(text=f"{r.mat_cost:.2f} €")
        self._cost_labels["frame"].configure(text=f"{r.frame_cost:.2f} €")
        self._cost_labels["angle"].configure(text=f"{r.angle_cost:.2f} €")
        self._cost_labels["bumper"].configure(text=f"{r.bumper_cost:.2f} €")
        self._cost_labels["work"].configure(text=f"{r.work_cost:.2f} €")
        self._cost_labels["zinc"].configure(text=f"{r.zinc_cost:.2f} €")
        self._cost_labels["total"].configure(text=f"{r.total_cost:.2f} €")
        self._weight_labels["w_mats"].configure(text=f"{r.weight_mats:.2f}")
        self._weight_labels["w_frame"].configure(text=f"{r.weight_frame:.2f}")
        self._weight_labels["w_total"].configure(text=f"{r.total_weight:.2f}")
        self._weight_labels["w_zinc"].configure(text=f"{r.total_weight_with_zinc:.2f}")
        self._cost_per_kg_label.configure(text=f"Себест. €/кг: {r.cost_per_kg:.2f}")
        self._update_client_price()
        self._show_screen("result")

    def _update_client_price(self):
        if not hasattr(self, "_last_result"):
            return
        margin = _float(self._margin_var.get()) if self._margin_var.get() else 0.0
        client_price = round(self._last_result.total_cost * (1 + margin / 100), 2)
        self._client_price_label.configure(text=f"{client_price:,.2f} €".replace(",", " "))

    def _export_pdf(self):
        if not hasattr(self, "_last_result"):
            return
        margin = _float(self._margin_var.get()) or 0.0

        # Диалог выбора языка PDF
        lang_win = ctk.CTkToplevel(self)
        lang_win.title("Язык PDF")
        lang_win.geometry("260x130")
        lang_win.grab_set()
        lang_win.resizable(False, False)

        ctk.CTkLabel(lang_win, text="Выберите язык оффера:", font=("Segoe UI", 11)).pack(pady=(16, 8))
        lang_var = ctk.StringVar(value=self.cfg.get("pdf_language", "en"))
        btn_row = ctk.CTkFrame(lang_win, fg_color="transparent")
        btn_row.pack()

        def pick(lang):
            lang_var.set(lang)
            lang_win.destroy()
            self._do_export_pdf(margin, lang)

        ctk.CTkButton(btn_row, text="🇬🇧  English", width=110,
                      command=lambda: pick("en")).pack(side="left", padx=8)
        ctk.CTkButton(btn_row, text="🇪🇪  Eesti", width=110,
                      command=lambda: pick("et")).pack(side="left", padx=8)

        lang_win.bind("<Escape>", lambda e: lang_win.destroy())

    def _do_export_pdf(self, margin: float, language: str):
        self.cfg["pdf_language"] = language
        pm.save_config(self.cfg)
        client = self.current_client or "client"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"Offer_{client}.pdf",
        )
        if not path:
            return
        try:
            pdf_exp.export_offer_pdf(
                output_path=path,
                client_name=self.current_client,
                parts=self.current_parts,
                total_cost=self._last_result.total_cost,
                margin_pct=margin,
                company_name=self.cfg.get("company_name", "Grid Master"),
                company_contact=self.cfg.get("company_contact", ""),
                language=language,
            )
            messagebox.showinfo("Done", f"PDF saved:\n{path}")
        except Exception as e:
            messagebox.showerror("PDF Error", str(e))

    # ------------------------------------------------------------------
    # Cutting / nesting
    # ------------------------------------------------------------------

    def _cutting_dialog(self):
        """Run nesting and show result summary + visual preview."""
        if not self.current_parts:
            messagebox.showwarning("Нет деталей", "Список позиций пуст.")
            return

        mat_l = int(self.cfg.get("mat_length", 6100))
        mat_w = int(self.cfg.get("mat_width", 1000))

        try:
            order = self._build_order()
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
            return
        result = core.cutting_run(order.parts, mat_l, mat_w, kerf=5)

        mats = result["mats"]
        mat_count = result["mat_count"]
        skipped = result["skipped"]

        if not mats:
            messagebox.showwarning("Раскрой", "Нет деталей подходящего размера.")
            return

        msg = (f"Листов (раскладок): {len(mats)}\n"
               f"Матов к списанию:    {mat_count:.1f}\n")
        if skipped:
            msg += f"\nПредупреждение: {skipped} шт. не вписываются в мат и исключены."

        dlg = ctk.CTkToplevel(self)
        dlg.title("Раскрой матов")
        dlg.geometry("320x200")
        dlg.grab_set()
        dlg.resizable(False, False)

        ctk.CTkLabel(dlg, text="Результат раскроя", font=("Segoe UI", 13, "bold")).pack(pady=(16, 6))
        ctk.CTkLabel(dlg, text=msg, justify="left").pack(padx=20, anchor="w")

        btns = ctk.CTkFrame(dlg, fg_color="transparent")
        btns.pack(pady=12)
        ctk.CTkButton(btns, text="👁  Схема", width=120,
                      command=lambda: (dlg.destroy(), self._cutting_preview(mats, mat_l, mat_w))
                      ).pack(side="left", padx=8)
        ctk.CTkButton(btns, text="Закрыть", width=100,
                      command=dlg.destroy).pack(side="left", padx=8)

    def _cutting_preview(self, mats: list, mat_length: int, mat_width: int):
        """Visual mat-by-mat nesting preview."""
        try:
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            from matplotlib.patches import Rectangle as MRect
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        except ImportError:
            messagebox.showerror("matplotlib", "Установите: pip install matplotlib")
            return

        win = ctk.CTkToplevel(self)
        win.title("Схема раскроя")
        win.geometry("900x620")
        win.grab_set()

        state = {"idx": 0}

        canvas_frame = ctk.CTkFrame(win)
        canvas_frame.pack(fill="both", expand=True, padx=8, pady=8)

        nav = ctk.CTkFrame(win)
        nav.pack(fill="x", padx=8, pady=(0, 8))

        fig, ax = plt.subplots(figsize=(10, 3.5))
        mpl_canvas = FigureCanvasTkAgg(fig, master=canvas_frame)
        mpl_canvas.get_tk_widget().pack(fill="both", expand=True)

        idx_label = ctk.CTkLabel(nav, text="")
        idx_label.pack(side="left", padx=12)

        def draw(idx):
            ax.clear()
            ax.set_xlim(0, mat_length)
            ax.set_ylim(0, mat_width)
            ax.set_aspect("equal", adjustable="box")
            ax.set_xlabel("длина (мм)")
            ax.set_ylabel("ширина (мм)")
            ax.set_title(f"Мат #{idx + 1} / {len(mats)}", fontsize=11)
            ax.add_patch(MRect((0, 0), mat_length, mat_width,
                               linewidth=2, edgecolor="black", facecolor="#f5f5f5"))
            colors = plt.cm.tab20.colors
            for i, d in enumerate(mats[idx]):
                color = colors[i % len(colors)]
                ax.add_patch(MRect((d["x"], d["y"]), d["length"], d["width"],
                                   linewidth=1, edgecolor="black", facecolor=color, alpha=0.7))
                ax.text(d["x"] + d["length"] / 2, d["y"] + d["width"] / 2,
                        f'{d["position"]}\n{d["length"]}×{d["width"]}',
                        ha="center", va="center", fontsize=7, clip_on=True)
            fig.tight_layout(pad=0.5)
            mpl_canvas.draw_idle()
            idx_label.configure(text=f"Мат {idx + 1} из {len(mats)}")

        def prev_():
            state["idx"] = (state["idx"] - 1) % len(mats)
            draw(state["idx"])

        def next_():
            state["idx"] = (state["idx"] + 1) % len(mats)
            draw(state["idx"])

        ctk.CTkButton(nav, text="←", width=70, command=prev_).pack(side="left", padx=4)
        ctk.CTkButton(nav, text="→", width=70, command=next_).pack(side="left", padx=4)
        ctk.CTkButton(nav, text="Закрыть", width=90,
                      command=win.destroy).pack(side="right", padx=8)

        draw(0)

    def _export_xlsx(self):
        if not hasattr(self, "_last_result"):
            return
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        import math as _math

        r = self._last_result
        client = self.current_client or "project"
        date_str = datetime.now().strftime("%Y-%m-%d")
        margin = _float(self._margin_var.get()) or 0.0
        client_price = round(r.total_cost * (1 + margin / 100), 2)
        angle_pcs = _math.ceil(r.angle_length_m / 3.0) if r.angle_length_m > 0 else 0

        wb = openpyxl.Workbook()

        # ── Лист 1: Позиции (для склада) ────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Позиции"

        hdr_fill = PatternFill("solid", fgColor="2E6B9E")
        hdr_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["№", "Позиция", "Длина (мм)", "Длина 2 (мм)", "Ширина (мм)",
                   "Кол-во", "Радиус (мм)", "Доля окружн.", "Отбойник (мм)",
                   "Норм.-часов/шт", "Норм.-часов итого", "Примечание"]
        col_widths = [5, 16, 12, 12, 12, 8, 12, 12, 13, 16, 17, 24]

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
            ws1.column_dimensions[cell.column_letter].width = w

        ws1.row_dimensions[1].height = 30

        for i, p in enumerate(self.current_parts, 1):
            qty = int(p.get("quantity", 1))
            work_per = float(p.get("work", 0))
            row_data = [
                i,
                str(p.get("position", "")),
                int(p.get("length", 0)),
                int(p.get("length2", 0)) or "",
                int(p.get("width", 0)),
                qty,
                float(p.get("radius", 0)) or "",
                float(p.get("radius_part", 1.0)) if p.get("radius") else "",
                float(p.get("bumper", 0)) or "",
                round(work_per, 3),
                round(work_per * qty, 3),
                str(p.get("notes", "")),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws1.cell(row=i + 1, column=col, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

        # Итоговая строка трудозатрат
        total_work = sum(float(p.get("work", 0)) * int(p.get("quantity", 1))
                         for p in self.current_parts)
        last_row = len(self.current_parts) + 2
        ws1.cell(row=last_row, column=10, value="ИТОГО часов:")
        ws1.cell(row=last_row, column=10).font = Font(bold=True)
        ws1.cell(row=last_row, column=11, value=round(total_work, 2)).font = Font(bold=True)

        # ── Лист 2: Сводка (себестоимость, материалы) ───────────────────────
        ws2 = wb.create_sheet("Сводка")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 18

        def _row(label, value, bold=False):
            r_idx = ws2.max_row + 1
            c1 = ws2.cell(row=r_idx, column=1, value=label)
            c2 = ws2.cell(row=r_idx, column=2, value=value)
            if bold:
                c1.font = Font(bold=True)
                c2.font = Font(bold=True)
            c2.alignment = Alignment(horizontal="right")

        ws2.cell(row=1, column=1, value="Проект").font = Font(bold=True, size=13)
        ws2.cell(row=1, column=2, value=client)
        ws2.cell(row=2, column=1, value="Дата")
        ws2.cell(row=2, column=2, value=date_str)
        ws2.append([])

        ws2.cell(row=ws2.max_row + 1, column=1,
                  value="СЕБЕСТОИМОСТЬ").font = Font(bold=True, color="2E6B9E")
        _row("Маты (+15%)", f"{r.mat_cost:.2f} €")
        _row("Обрамление", f"{r.frame_cost:.2f} €")
        _row("Перфоуголок", f"{r.angle_cost:.2f} €")
        _row("Отбойник", f"{r.bumper_cost:.2f} €")
        _row("Работа", f"{r.work_cost:.2f} €")
        _row("Цинкование", f"{r.zinc_cost:.2f} €")
        _row("ИТОГО себест.", f"{r.total_cost:.2f} €", bold=True)
        ws2.append([])

        ws2.cell(row=ws2.max_row + 1, column=1,
                  value="МАТЕРИАЛЫ").font = Font(bold=True, color="2E6B9E")
        _row("Матов листов (+15%)", r.mats_count_with_k)
        _row("Обрамление, м", f"{r.frame_length_m:.2f}")
        _row("Перфоуголок, м", f"{r.angle_length_m:.2f}")
        _row("Перфоуголок (3м), шт", angle_pcs)
        _row("Отбойник, м", f"{r.bumper_length_m:.2f}")
        ws2.append([])

        ws2.cell(row=ws2.max_row + 1, column=1,
                  value="ВЕСА").font = Font(bold=True, color="2E6B9E")
        _row("Решётка, кг", f"{r.weight_mats:.2f}")
        _row("Обрамление, кг", f"{r.weight_frame:.2f}")
        _row("Общий вес, кг", f"{r.total_weight:.2f}")
        _row("С цинком (+10%), кг", f"{r.total_weight_with_zinc:.2f}")
        ws2.append([])

        ws2.cell(row=ws2.max_row + 1, column=1,
                  value="ЦЕНА КЛИЕНТУ").font = Font(bold=True, color="2E6B9E")
        _row(f"Маржа {margin:.0f}%", "")
        _row("Цена клиенту", f"{client_price:.2f} €", bold=True)

        # ── Сохранить ────────────────────────────────────────────────────────
        safe_client = "".join(c for c in client if c.isalnum() or c in " _-").strip()
        default_name = f"{safe_client}_{date_str}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name,
        )
        if not path:
            return
        try:
            wb.save(path)
            messagebox.showinfo("Готово", f"Файл сохранён:\n{path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{e}")

    def _save_order(self):
        if not hasattr(self, "_last_result"):
            return
        from datetime import datetime
        margin = _float(self._margin_var.get()) or 0.0
        client_price = round(self._last_result.total_cost * (1 + margin / 100), 2)
        order = {
            "date":         datetime.now().strftime("%d.%m.%Y"),
            "client":       self.current_client,
            "total_cost":   self._last_result.total_cost,
            "margin_pct":   margin,
            "client_price": client_price,
            "parts":        self.current_parts,
        }
        self.orders.append(order)
        _save_orders(self.orders)
        self.cfg["margin_pct"] = margin
        pm.save_config(self.cfg)
        self._refresh_history()
        messagebox.showinfo("Сохранено", "Заказ добавлен в историю.")


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = ManagerApp()
    app.mainloop()
