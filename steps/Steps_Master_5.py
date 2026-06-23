
# Полная версия на customtkinter

import os
import json
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import customtkinter as ctk

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import math
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle

CONFIG_FILE = "config.json"


# ---------------------------- Утилиты и конфиг ----------------------------

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_config(config):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)


def _autofit_columns(ws, padding=2, max_width=60):
    """Подгоняет ширину всех колонок по максимальной длине значения в столбце."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max_len + padding, max_width)
def _select_row(df: pd.DataFrame, column: str, value) -> pd.Series:
    """Безопасно выбирает первую строку по значению, приводя обе стороны к str."""
    mask = df[column].astype(str) == str(value)
    if not mask.any():
        raise ValueError(f"В таблице нет значения {value!r} в колонке {column!r}")
    return df.loc[mask].iloc[0]

def _to_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace("\u00A0", "").replace(",", ".")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize_numeric_columns(df: pd.DataFrame,
                              skip_cols=None,
                              round_map=None,
                              default_round=3) -> pd.DataFrame:
    """
    Преобразует числовые столбцы к float и округляет.
    skip_cols — колонки, которые не трогаем (текстовые идентификаторы).
    round_map — словарь {колонка: количество_знаков}, иначе default_round.
    """
    if skip_cols is None:
        skip_cols = []
    if round_map is None:
        round_map = {}

    out = df.copy()
    for col in out.columns:
        if col in skip_cols:
            continue
        # Пытаемся привести к числу
        coerced = out[col].apply(_to_float)
        if coerced.notna().sum() > 0 and coerced.isna().sum() < len(coerced):
            # смешанные типы — если большинство как число, приводим
            out[col] = coerced
        elif coerced.notna().all():
            out[col] = coerced

        # Округление числовых
        if pd.api.types.is_numeric_dtype(out[col]):
            digits = round_map.get(col, default_round)
            out[col] = out[col].round(digits)

    return out


# ---------------------------- Класс форматированного ввода ----------------------------

class FormattedEntry(ctk.CTkEntry):
    def __init__(self, master, decimals: int = 2, **kwargs):
        # ВАЖНО: вытащить decimals из kwargs, чтобы он НЕ ушёл в CTkEntry (иначе ValueError)
        self.decimals = int(kwargs.pop("decimals", decimals))
        self._var = tk.StringVar()
        super().__init__(master, textvariable=self._var, **kwargs)

        # автоформат при выходе из поля и по Enter
        self.bind("<FocusOut>", lambda e: self._on_focus_out())
        self.bind("<Return>",   lambda e: self._on_focus_out())

    # нормализация: убираем пробелы, запятую превращаем в точку
    def _normalize_text(self, text: str) -> str:
        return str(text).strip().replace(" ", "").replace(",", ".")

    # чтение как число (принимает и "1,23", и "1.23")
    def get_float(self) -> float:
        s = self._normalize_text(self._var.get())
        if s == "":
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0

    # задать текст с форматированием (разделитель тысяч — пробел, дробная часть — ЗАПЯТАЯ)
    def set_text(self, value) -> None:
        if value is None:
            self._var.set("")
            return
        try:
            # Преобразуем в число и форматируем
            if isinstance(value, str):
                # Заменяем запятую на точку для преобразования в float
                num = float(value.replace(",", "."))
            else:
                num = float(value)
            self._var.set(self._format(num))
        except Exception:
            # Если не число, оставляем как есть
            self._var.set(str(value))
            
    # форматирование в строку: "1 234,50" при decimals=2, "1 234" при decimals=0
    def _format(self, num: float) -> str:
        fmt = f"{{:,.{self.decimals}f}}"   # 1,234.50 / 1,234
        s = fmt.format(num)
        return s.replace(",", " ").replace(".", ",")  # 1 234,50

    # автоформат текущего содержимого
    def _on_focus_out(self) -> None:
        self.set_text(self._var.get())
            

# ---------------------------- Доменная логика ----------------------------

class StepCalculator:
    def __init__(self):
        self.angle_weights: pd.DataFrame | None = None
        self.force_weights: pd.DataFrame | None = None
        self.mat_data: pd.DataFrame | None = None
        self.side_data: pd.DataFrame | None = None

        self.current_project = None
        self.zinc_price_per_kg = 0.0
        self.hourly_rate = 0.0
        self.angle_price_per_kg = 0.0
        self.kickplate_price_per_kg = 0.0
        

        self.config = {}
        self.load_config_and_tables()
        
        self.zinc_price_per_kg = self.config.get('last_zinc_price', 0.0)
        self.angle_price_per_kg = self.config.get('last_angle_price', 0.0)
        self.kickplate_price_per_kg = self.config.get('last_kickplate_price', 0.0)
        self.hourly_rate = self.config.get('last_hourly_rate', 0.0)

    def load_config_and_tables(self):
        """Загружает конфиг и все таблицы (если пути сохранены)."""
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    self.config = json.load(f)
                    print("Конфиг загружен:", self.config)

            if "Боковины" in self.config:
                self._load_table_data("Боковины", self.config["Боковины"])
            if "Типы матов" in self.config:
                self._load_table_data("Типы матов", self.config["Типы матов"])
            if "Перфоуголок" in self.config:
                self._load_table_data("Перфоуголок", self.config["Перфоуголок"])
            if "Кикплейт" in self.config:
                self._load_table_data("Кикплейт", self.config["Кикплейт"])
        except Exception as e:
            print(f"Ошибка при загрузке конфига: {e}")

    def _load_table_data(self, table_type, filepath):
        """Загружает данные таблицы, сохраняя цены для боковин и матов."""
        try:
            if not os.path.exists(filepath):
                print(f"Файл {table_type} не найден по пути: {filepath}")
                return False

            # Читаем Excel, удаляем пустые строки/столбцы
            df = pd.read_excel(filepath, header=0).dropna(how='all').dropna(axis=1, how='all')
            df.columns = [str(c).strip() for c in df.columns]

            if table_type == "Перфоуголок":
                # Теперь содержит только размер и вес (без цены)
                self.angle_weights = df
                self.angle_weights.loc[len(self.angle_weights)] = ['Нет', 0]  # Без цены!

            elif table_type == "Кикплейт":
                # Только размер и вес (цена берётся из полей ввода)
                self.force_weights = df
                self.force_weights.loc[len(self.force_weights)] = ['Нет', 0]

            elif table_type == "Типы матов":
                # Оставляем цены в таблице (как раньше)
                required = ["Артикул", "Цена м²", "Вес м²"]
                missing = [c for c in required if c not in df.columns]
                if missing:
                    raise ValueError(f"В таблице 'Типы матов' нет колонок: {', '.join(missing)}")
                
                # Конвертируем числовые колонки
                for num_col in ["Цена м²", "Вес м²", "Цена мата", "Вес мата"]:
                    if num_col in df.columns:
                        df[num_col] = df[num_col].apply(_to_float)
                self.mat_data = df

            elif table_type == "Боковины":
                # Оставляем цены в таблице (как раньше)
                self.side_data = df

            print(f"Таблица {table_type} успешно загружена")
            return True

        except Exception as e:
            print(f"Ошибка загрузки таблицы {table_type}: {e}")
            return False
    
    def load_table(self, table_type, filepath):
        """Публичный метод для загрузки таблицы с сохранением в конфиг."""
        if self._load_table_data(table_type, filepath):
            self.config[table_type] = filepath
            self._save_config()
            return True
        return False

    def _save_config(self):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=4)
            print("Конфиг сохранен")
        except Exception as e:
            print(f"Ошибка сохранения конфига: {e}")
            
    def _save_last_prices(self):
        # сохранить последние введённые цены в config.json
        self.config.update({
            'last_zinc_price': self.zinc_price_per_kg,
            'last_angle_price': self.angle_price_per_kg,
            'last_kickplate_price': self.kickplate_price_per_kg,
            'last_hourly_rate': self.hourly_rate
        })
        self._save_config()

    def set_zinc_price(self, price):
        self.zinc_price_per_kg = price

    def set_angle_price(self, price):
        self.angle_price_per_kg = price

    def set_kickplate_price(self, price):
        self.kickplate_price_per_kg = price

    def set_hourly_rate(self, rate):
        self.hourly_rate = rate
        
    

    def create_project(self, length, quantity, side_choice, grid_choice,
                       angle_choice=None, profile_choice=None,
                       work_hours=0, complexity=1.0):
        try:
            side_info = _select_row(self.side_data, 'Размер', side_choice)
            width = _to_float(side_info['Ширина']) or 0.0

            self.current_project = {
                'length': float(length),
                'width': float(width),
                'quantity': int(quantity),
                'side_choice': side_choice,
                'grid_choice': grid_choice,
                'angle_choice': (angle_choice if angle_choice and angle_choice != 'Нет' else None),
                'profile_choice': (profile_choice if profile_choice and profile_choice != 'Нет' else None),
                'work_hours': float(work_hours),
                'complexity': float(complexity),
                'hourly_rate': self.hourly_rate,
                'has_zinc': self.zinc_price_per_kg > 0,
                'results': None,
                'intermediate_results': None
            }
            return self.calculate()
        except Exception as e:
            print(f"Ошибка создания проекта: {e}")
            return False


    def calculate(self):
        if not self.current_project:
            return False
        try:
            cp = self.current_project
            intermediate = {}

            # ==== Боковины ====
            side_info = _select_row(self.side_data, 'Размер', cp['side_choice'])
            side_w_1pc = _to_float(side_info.get('Вес')) or 0.0
            side_p_1pc = _to_float(side_info.get('Цена')) or 0.0

            qty = int(cp['quantity'])
            side_pcs_per_unit = 2

            side_weight_unit = side_pcs_per_unit * side_w_1pc
            side_price_unit  = side_pcs_per_unit * side_p_1pc

            side_pcs_total    = side_pcs_per_unit * qty
            side_weight_total = side_weight_unit * qty
            side_price_total  = side_price_unit  * qty

            intermediate['Боковины'] = {
                'Штук (всего)': side_pcs_total,
                'Вес (кг, всего)': round(side_weight_total, 3),
                'Цена (€)': round(side_price_total, 2)
            }

            # ==== Решетка ====
            grid_row = _select_row(self.mat_data, 'Артикул', cp['grid_choice'])
            grid_w_m2 = _to_float(grid_row.get('Вес м²')) or 0.0
            grid_p_m2 = _to_float(grid_row.get('Цена м²')) or 0.0

            grid_area = (cp['length'] * cp['width']) / 1_000_000.0
            grid_weight = grid_area * grid_w_m2
            grid_price  = grid_area * grid_p_m2

            intermediate['Решетка'] = {
                'Вес (кг)': grid_weight * qty,
                'Цена (€)': grid_price  * qty
            }

            # ==== Перфоуголок ====
            angle_weight_unit = 0.0
            angle_price_unit  = 0.0
            angle_len_total_m = 0.0
            angle_pieces_count = 0.0

            if cp['angle_choice']:
                angle_row = _select_row(self.angle_weights, 'Размер', cp['angle_choice'])
                angle_len_unit_m = cp['length'] / 1000.0
                angle_len_total_m = angle_len_unit_m * qty

                # вес можно оставить для статистики
                angle_w_per_m = _to_float(angle_row.get('Вес')) or 0.0
                angle_weight_unit = angle_w_per_m * angle_len_unit_m

                # теперь стоимость считается по длине
                angle_price_m = self.angle_price_per_kg   # теперь это €/м
                angle_price_unit = angle_len_unit_m * angle_price_m
                angle_price_total = angle_price_unit * qty

                angle_weight_total = angle_weight_unit * qty
                angle_price_total  = angle_price_unit  * qty
                angle_pieces_count = angle_len_total_m / 6.0

                intermediate['Перфоуголок'] = {
                    'Вес (кг)': angle_weight_unit * qty,
                    'Длина (м)': angle_len_total_m,
                    'Цена (€)': angle_price_total,
                    'Штук по 6 м': angle_len_total_m / 6.0
                }

            # ==== Кикплейт ====
            profile_weight = 0.0
            profile_price  = 0.0
            if cp['profile_choice']:
                profile_row = _select_row(self.force_weights, 'Размер', cp['profile_choice'])
                profile_w_per_m  = _to_float(profile_row.get('Вес')) or 0.0
                profile_price_kg = self.kickplate_price_per_kg

                profile_weight = profile_w_per_m * (cp['length'] / 1000.0)
                profile_price  = profile_weight * profile_price_kg

                intermediate['Кикплэйт'] = {
                    'Вес (кг)': profile_weight * qty,
                    'Цена (€)': profile_price  * qty
                }

            # ==== Работа ====
            work_cost = cp['work_hours'] * cp['hourly_rate']
            intermediate['Работа'] = {
                'Часы': cp['work_hours'],
                'Стоимость (€)': work_cost * qty
            }

            # ==== Итоги ====
            unit_weight = side_weight_unit + grid_weight + angle_weight_unit + profile_weight
            unit_price  = (side_price_unit + grid_price + angle_price_unit + profile_price) * (cp['complexity'] or 1.0)
            unit_price_with_work = unit_price + work_cost

            total_weight = unit_weight * qty
            total_price  = unit_price_with_work * qty
            total_weight_with_zinc = total_weight * 1.1

            if cp['has_zinc']:
                zinc_cost = total_weight_with_zinc * self.zinc_price_per_kg
                total_price += zinc_cost
                intermediate['Цинкование'] = {
                    'Цена (€/кг)': self.zinc_price_per_kg,
                    'Вес с цинком (кг)': total_weight_with_zinc,
                    'Стоимость цинка (€)': zinc_cost
                }

            cp['results'] = {
                'unit_weight': round(unit_weight, 3),
                'unit_price': round(unit_price, 2),
                'unit_price_with_work': round(unit_price_with_work, 2),
                'total_weight': round(total_weight, 3),
                'total_weight_with_zinc': round(total_weight_with_zinc, 3),
                'total_price': round(total_price, 2),
                'work_cost': round(work_cost, 2)
            }
            cp['intermediate_results'] = intermediate
            self._save_last_prices()
            return True
            
            self.config.update({
                'last_zinc_price': self.zinc_price_per_kg,
                'last_angle_price': self.angle_price_per_kg,
                'last_kickplate_price': self.kickplate_price_per_kg,
                'last_hourly_rate': self.hourly_rate
            })
            self._save_config()
            return True

        except Exception as e:
            print(f"Ошибка расчета: {e}")
            return False

    def export_to_excel(self, filename):
        """Экспорт проекта в Excel (добавляет новый лист, если файл существует)."""
        if not self.current_project or not self.current_project.get('results'):
            return False
        try:
            if os.path.exists(filename):
                wb = openpyxl.load_workbook(filename)
            else:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                    del wb["Sheet"]

            base_title = "Расчет ступеней"
            title = base_title
            n = 1
            while title in wb.sheetnames:
                n += 1
                title = f"{base_title} ({n})"
            ws = wb.create_sheet(title)

            bold = openpyxl.styles.Font(bold=True)
            header_font = openpyxl.styles.Font(bold=True, size=12)
            center = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            right = openpyxl.styles.Alignment(horizontal="right", vertical="center")
            num2 = '#,##0.00'

            ws["A1"] = "Расчет ступеней"
            ws["A1"].font = header_font
            ws["B1"] = datetime.now().strftime('%d.%m.%Y')
            ws.merge_cells("B1:E1")

            ws.append(["Компонент", "Тип", "Количество", "Единицы", "Стоимость"])
            for c in "ABCDE":
                ws[f"{c}3"].font = bold
                ws[f"{c}3"].alignment = center

            r = 4
            cp = self.current_project
            res = cp["results"]
            inter = cp["intermediate_results"]
            
            # Количество матов (если считали раскрой)
            mats_count = int(self.current_project.get("mats_count", 0) or 0)
            if mats_count > 0:
                ws.cell(row=r, column=1, value="Количество матов")
                ws.cell(row=r, column=3, value=mats_count).number_format = '#,##0'
                ws.cell(row=r, column=4, value="шт")
                r += 1

            # Ступени - ИЗМЕНЯЕМ РАСЧЕТ СТОИМОСТИ ЕДИНИЦЫ
            ws.cell(row=r, column=1, value="Ступени")
            c = ws.cell(row=r, column=2, value=str(int(cp["length"])))
            c.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
            ws.cell(row=r, column=3, value=cp["quantity"]).number_format = num2
            ws.cell(row=r, column=4, value="шт")
            
            # Рассчитываем стоимость единицы с учетом цинкования
            unit_price_with_zinc = res["total_price"] / cp["quantity"]
            c = ws.cell(row=r, column=5, value=unit_price_with_zinc)
            c.number_format = num2
            r += 1

            # Кикплейт
            side = inter.get("Боковины", {})
            ws.cell(row=r, column=1, value="Боковины")
            ws.cell(row=r, column=2, value=cp["side_choice"])
            sides_count = int(round(cp.get("quantity", 0))) * 2
            ws.cell(row=r, column=3, value=sides_count).number_format = '#,##0'
            ws.cell(row=r, column=4, value="шт")
            ws.cell(row=r, column=5, value=side.get("Цена (€)", 0)).number_format = num2
            r += 1

            # Решетка
            grid = inter.get("Решетка", {})
            ws.cell(row=r, column=1, value="Решетка")
            ws.cell(row=r, column=2, value=cp["grid_choice"])
            c = ws.cell(row=r, column=3, value=grid.get("Вес (кг)", 0)); c.number_format = num2
            ws.cell(row=r, column=4, value="кг")
            c = ws.cell(row=r, column=5, value=grid.get("Цена (€)", 0)); c.number_format = num2
            r += 1

            # Перфоуголок (если есть)
            if "Перфоуголок" in inter:
                ang = inter["Перфоуголок"]
                ws.cell(row=r, column=1, value="Перфоуголок")
                ws.cell(row=r, column=2, value=cp.get("angle_choice", ""))
                qty = ang.get("Штук по 6 м", 0)
                if qty:
                    ws.cell(row=r, column=3, value=qty).number_format = num2
                    ws.cell(row=r, column=4, value="шт")
                else:
                    ws.cell(row=r, column=3, value=ang.get("Вес (кг)", 0)).number_format = num2
                    ws.cell(row=r, column=4, value="кг")
                ws.cell(row=r, column=5, value=ang.get("Цена (€)", 0)).number_format = num2
                r += 1

            # Работа
            work = inter.get("Работа", {})
            ws.cell(row=r, column=1, value="Работа")
            ws.cell(row=r, column=2, value="")
            ws.cell(row=r, column=3, value=cp["work_hours"] * cp["quantity"]).number_format = num2
            ws.cell(row=r, column=4, value="часов")
            ws.cell(row=r, column=5, value=work.get("Стоимость (€)", 0)).number_format = num2
            r += 1

            # Общий черный вес
            ws.cell(row=r, column=1, value="Общий черный вес (кг)")
            ws.cell(row=r, column=3, value=res["total_weight"]).number_format = num2
            ws.cell(row=r, column=4, value="кг"); r += 1

            # Общий вес с цинком +10%
            ws.cell(row=r, column=1, value="Общий вес с цинком, +10% (кг)")
            ws.cell(row=r, column=3, value=res["total_weight_with_zinc"]).number_format = num2
            r += 1
            
            

            # Кикплэйт (если есть)
            if "Кикплэйт" in inter:
                prof = inter["Кикплэйт"]
                ws.cell(row=r, column=1, value="Кикплэйт")
                ws.cell(row=r, column=2, value=cp.get("profile_choice") or "Нет")
                ws.cell(row=r, column=3, value=prof.get("Вес (кг)", 0)).number_format = num2
                ws.cell(row=r, column=4, value="кг")
                ws.cell(row=r, column=5, value=prof.get("Цена (€)", 0)).number_format = num2
                r += 1
            else:
                ws.cell(row=r, column=1, value="Кикплэйт")
                ws.cell(row=r, column=2, value="Нет")
                ws.cell(row=r, column=4, value="кг"); r += 1

            # Цинкование
            if "Цинкование" in inter:
                zinc = inter["Цинкование"]
                ws.cell(row=r, column=1, value="Цинкование (€/кг)")
                ws.cell(row=r, column=3, value=zinc.get("Цена (€/кг)", 0)).number_format = num2
                ws.cell(row=r, column=5, value=zinc.get("Стоимость цинка (€)", 0)).number_format = num2
                r += 1
            else:
                ws.cell(row=r, column=1, value="Цинкование (€/кг)"); r += 1

            # ИТОГО
            ws.cell(row=r, column=1, value="Общая стоимость (€)")
            ws.cell(row=r, column=5, value=res["total_price"]).number_format = num2
            r += 1           

            if '_autofit_columns' in globals():
                _autofit_columns(ws)

            wb.save(filename)
            return True
        except Exception as e:
            print(f"Ошибка экспорта в Excel: {e}")
            return False


# ---------------------------- GUI на customtkinter ----------------------------

class StepCalculatorApp:
    def __init__(self, root: ctk.CTk):
        self.root = root
        self.root.title("Steps Master calculator ©Viiptag")
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        # --- Новые переменные для раскроя матов ---
        self.mat_length_var = tk.IntVar(value=6100)
        self.mat_width_var = tk.IntVar(value=1000)
        self._cutting_results = []  # список матов, каждый мат — список прямоугольников

        # Домен
        self.calculator = StepCalculator()

        # Переменные
        self.project_name_var = tk.StringVar()

        # Поля ввода
        self.length_entry: FormattedEntry | None = None
        self.quantity_entry: FormattedEntry | None = None
        self.zinc_price_entry: FormattedEntry | None = None
        self.work_hours_entry: FormattedEntry | None = None
        self.hourly_rate_entry: FormattedEntry | None = None
        self.complexity_entry: FormattedEntry | None = None
        self.mats_count_entry: FormattedEntry | None = None
        self.angle_price_entry: FormattedEntry | None = None
        self.kickplate_price_entry: FormattedEntry | None = None

        # Комбобоксы
        self.side_combobox: ctk.CTkComboBox | None = None
        self.angle_combobox: ctk.CTkComboBox | None = None
        self.grid_combobox: ctk.CTkComboBox | None = None
        self.profile_combobox: ctk.CTkComboBox | None = None

        # Результаты
        self.results_vars = {
            'unit_price': tk.StringVar(),
            'unit_weight': tk.StringVar(),
            'total_weight': tk.StringVar(),
            'total_price': tk.StringVar()
        }

        # Таблицы в редакторе (вкладки)
        self.calculator_tables: dict[str, pd.DataFrame] = {}
        self.tables_tabview: ctk.CTkTabview | None = None
        self.table_trees: dict[str, ttk.Treeview] = {}  # имя таблицы -> Treeview

        # Notebook → CTkTabview
        self.tabview = ctk.CTkTabview(self.root)
        self.tabview.grid(row=0, column=0, sticky="nsew")
        self.tabview.add("Проект")
        self.tabview.add("Таблицы")

        self.project_tab = self.tabview.tab("Проект")
        self.tables_tab = self.tabview.tab("Таблицы")

        self.create_project_tab()
        self.create_tables_tab()
        
        # ТЕПЕРЬ УСТАНАВИВАЕМ ЗНАЧЕНИЯ ПОСЛЕ СОЗДАНИЯ ВСЕХ ЭЛЕМЕНТОВ ИНТЕРФЕЙСА
        self.zinc_price_entry.set_text(self.calculator.zinc_price_per_kg)
        self.angle_price_entry.set_text(self.calculator.angle_price_per_kg)
        self.kickplate_price_entry.set_text(self.calculator.kickplate_price_per_kg)
        self.hourly_rate_entry.set_text(self.calculator.hourly_rate)

        self.check_loaded_tables()
        
        
        
        # Настройки стиля
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.root.geometry("900x700")
        
    def _create_label_entry(self, parent, label_text, width=120):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.pack(fill="x", pady=3)
        ctk.CTkLabel(frame, text=label_text, width=140).pack(side="left")
        entry = FormattedEntry(frame, width=width)
        entry.pack(side="right")
        return entry

    def _create_combobox(self, parent, label_text, values):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.pack(fill="x", pady=3)
        ctk.CTkLabel(frame, text=label_text, width=140).pack(side="left")
        cb = ctk.CTkComboBox(frame, values=values, width=120)
        cb.pack(side="right")
        return cb
        
    
    def create_project_tab(self):
        import customtkinter as ctk
        import tkinter as tk

        # ====== Контейнер вкладки ======
        main_frame = ctk.CTkFrame(self.project_tab)
        main_frame.pack(fill="both", expand=True, padx=16, pady=16)

        # ====== Верхняя строка: имя проекта + кнопки справа ======
        top_row = ctk.CTkFrame(main_frame, fg_color="transparent")
        top_row.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(top_row, text="Имя проекта:").pack(side="left", padx=(0, 8))
        # поле имени проекта (растягивается)
        self.project_name_entry = ctk.CTkEntry(top_row, textvariable=self.project_name_var)
        self.project_name_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        # три кнопки справа
        ctk.CTkButton(top_row, text="Открыть", command=self.open_project).pack(side="left", padx=6)        
        ctk.CTkButton(top_row, text="Сохранить", command=self.save_project).pack(side="right", padx=6)
        ctk.CTkButton(top_row, text="Очистить", command=self.clear_project).pack(side="right", padx=6)
        

        # ====== Ряд из двух колонок ======
        cols_row = ctk.CTkFrame(main_frame, fg_color="transparent")
        cols_row.pack(fill="x", pady=(0, 8))

        # ----- Левая колонка -----
        left_col = ctk.CTkFrame(cols_row, fg_color="transparent")
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 8))

        # Группа: Основные параметры
        group_main = ctk.CTkFrame(left_col, corner_radius=8)
        group_main.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(group_main, text="Основные параметры", font=ctk.CTkFont(weight="bold"))\
            .grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(8, 6))

        # поля (создаём здесь, чтобы точно существовали)
        self.length_entry = FormattedEntry(group_main)
        self.quantity_entry = FormattedEntry(group_main)
        self.side_combobox  = ctk.CTkComboBox(group_main, values=['Нет'])
        self.angle_combobox = ctk.CTkComboBox(group_main, values=['Нет'])
        self.grid_combobox  = ctk.CTkComboBox(group_main, values=['Нет'])

        row = 1
        ctk.CTkLabel(group_main, text="Длина ступени (мм):").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.length_entry.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(group_main, text="Количество:").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.quantity_entry.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(group_main, text="Боковина:").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.side_combobox.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(group_main, text="Перфоуголок:").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.angle_combobox.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(group_main, text="Выбрать мат:").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.grid_combobox.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        group_main.grid_columnconfigure(1, weight=1)

        # Светофор под комбобоксами
        self.tables_status_frame = ctk.CTkFrame(group_main, fg_color="transparent")
        self.tables_status_frame.grid(row=row, column=0, columnspan=2, sticky="w", padx=8, pady=(2, 6))
        self._status_labels = {
            "Боковины":     ctk.CTkLabel(self.tables_status_frame, text="Боковины: —", text_color="#9CA3AF"),
            "Типы матов":   ctk.CTkLabel(self.tables_status_frame, text="Типы матов: —", text_color="#9CA3AF"),
            "Кикплейт":     ctk.CTkLabel(self.tables_status_frame, text="Кикплейт: —", text_color="#9CA3AF"),
            "Перфоуголок":  ctk.CTkLabel(self.tables_status_frame, text="Перфоуголок: —", text_color="#9CA3AF"),
        }
        for lbl in self._status_labels.values():
            lbl.pack(side="left", padx=(0, 10))
        self.update_table_status_lights()

        # Группа: Дополнительно
        group_extra = ctk.CTkFrame(left_col, corner_radius=8)
        group_extra.pack(fill="x")
        ctk.CTkLabel(group_extra, text="Дополнительно", font=ctk.CTkFont(weight="bold"))\
            .grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(8, 6))

        self.profile_combobox = ctk.CTkComboBox(group_extra, values=['Нет'])  # Кикплейт
        self.work_hours_entry = FormattedEntry(group_extra)  # часы
        self.complexity_entry = FormattedEntry(group_extra)  # коэффициент
        self.complexity_entry.set_text("1,00")

        row = 1
        ctk.CTkLabel(group_extra, text="Кикплейт:").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.profile_combobox.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(group_extra, text="Часы работы на 1 шт.:").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.work_hours_entry.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(group_extra, text="Сложность:").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.complexity_entry.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        group_extra.grid_columnconfigure(1, weight=1)

        # ----- Правая колонка -----
        right_col = ctk.CTkFrame(cols_row, fg_color="transparent")
        right_col.pack(side="left", fill="both", expand=True, padx=(8, 0))

        # Группа: Цены
        prices_group = ctk.CTkFrame(right_col, corner_radius=8)
        prices_group.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(prices_group, text="Цены", font=ctk.CTkFont(weight="bold"))\
            .grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(8, 6))

        row = 1
        ctk.CTkLabel(prices_group, text="Цинкование (€/кг):").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.zinc_price_entry = FormattedEntry(prices_group)
        self.zinc_price_entry.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(prices_group, text="Перфоуголок (€/м):").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.angle_price_entry = FormattedEntry(prices_group)
        self.angle_price_entry.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(prices_group, text="Кикплейт (€/кг):").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.kickplate_price_entry = FormattedEntry(prices_group)
        self.kickplate_price_entry.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1

        ctk.CTkLabel(prices_group, text="Работа (€/час):").grid(row=row, column=0, sticky="e", padx=8, pady=4)
        self.hourly_rate_entry = FormattedEntry(prices_group)
        self.hourly_rate_entry.grid(row=row, column=1, sticky="we", padx=8, pady=4); row += 1
        
        

        prices_group.grid_columnconfigure(1, weight=1)

        mats_group = ctk.CTkFrame(right_col, corner_radius=8)
        mats_group.pack(fill="x")
        ctk.CTkLabel(mats_group, text="Параметры матов", font=ctk.CTkFont(weight="bold"))\
            .grid(row=0, column=0, columnspan=3, sticky="w", padx=10, pady=(8, 6))

        # --- ПАРАМЕТРЫ МАТОВ (внутри mats_group) ---
        r = 1

        ctk.CTkLabel(mats_group, text="Длина мата (мм):").grid(row=r, column=0, sticky="e", padx=8, pady=4)
        self.mat_length_entry = FormattedEntry(mats_group, decimals=0)
        self.mat_length_entry.set_text("6100")
        self.mat_length_entry.grid(row=r, column=1, sticky="we", padx=8, pady=4)
        r += 1

        ctk.CTkLabel(mats_group, text="Ширина мата (мм):").grid(row=r, column=0, sticky="e", padx=8, pady=4)
        self.mat_width_entry = FormattedEntry(mats_group, decimals=0)
        self.mat_width_entry.set_text("1000")
        self.mat_width_entry.grid(row=r, column=1, sticky="we", padx=8, pady=4)
        r += 1

        ctk.CTkLabel(mats_group, text="Количество матов:").grid(row=r, column=0, sticky="e", padx=8, pady=4)
        self.mats_count_entry = FormattedEntry(mats_group, decimals=0)
        self.mats_count_entry.set_text("0")
        self.mats_count_entry.configure(state="disabled")
        self.mats_count_entry.grid(row=r, column=1, sticky="we", padx=8, pady=4)
        r += 1  # <-- кнопки идут на следующей строке

        # Кнопки под количеством матов
        buttons_row = ctk.CTkFrame(mats_group, fg_color="transparent")
        buttons_row.grid(row=r, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 6))

        ctk.CTkButton(buttons_row, text="Рассчитать количество",
                      command=self.calculate_mats).pack(side="left", padx=(0, 8))
        ctk.CTkButton(buttons_row, text="Предпросмотр",
                      command=self.preview_cutting).pack(side="left")
        r += 1

        mats_group.grid_columnconfigure(1, weight=1)


        # ====== Панель действий под колонками ======
        actions = ctk.CTkFrame(main_frame, fg_color="transparent")
        actions.pack(fill="x", pady=(0, 8))
        ctk.CTkButton(actions, text="Рассчитать",        command=self.calculate).pack(side="left", padx=6)
        ctk.CTkButton(actions, text="Экспорт в Excel", command=self.export_to_excel).pack(side="left", padx=6)
        
        

        # ====== Результаты (внизу!) ======
        results_frame = ctk.CTkFrame(main_frame, corner_radius=8)
        results_frame.pack(fill="both", expand=True)

        # строка-«сводка»
        res_grid = ctk.CTkFrame(results_frame, fg_color="transparent")
        res_grid.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(res_grid, text="Стоимость за единицу:").grid(row=0, column=0, sticky="e", padx=5)
        ctk.CTkLabel(res_grid, textvariable=self.results_vars['unit_price']).grid(row=0, column=1, sticky="w", padx=5)

        ctk.CTkLabel(res_grid, text="Вес единицы:").grid(row=1, column=0, sticky="e", padx=5)
        ctk.CTkLabel(res_grid, textvariable=self.results_vars['unit_weight']).grid(row=1, column=1, sticky="w", padx=5)

        ctk.CTkLabel(res_grid, text="Черный вес заказа:").grid(row=0, column=2, sticky="e", padx=5)
        ctk.CTkLabel(res_grid, textvariable=self.results_vars['total_weight']).grid(row=0, column=3, sticky="w", padx=5)

        ctk.CTkLabel(res_grid, text="Общая стоимость:").grid(row=1, column=2, sticky="e", padx=5)
        ctk.CTkLabel(res_grid, textvariable=self.results_vars['total_price']).grid(row=1, column=3, sticky="w", padx=5)

        # большой лог/промежуточные вычисления
        self.intermediate_text = ctk.CTkTextbox(results_frame, wrap="none")
        self.intermediate_text.pack(fill="both", expand=True, padx=10, pady=(0, 10))
    def update_table_status_lights(self):
        """Обновляет индикаторы ✓/— для таблиц под блоком компонентов."""
        ok_color = "#16A34A"   # зелёный
        off_color = "#9CA3AF"  # серый

        # что загружено в модели
        states = {
            "Боковины":    getattr(self.calculator, "side_data", None) is not None,
            "Типы матов":  getattr(self.calculator, "mat_data", None) is not None,
            "Кикплейт":    getattr(self.calculator, "force_weights", None) is not None,
            "Перфоуголок": getattr(self.calculator, "angle_weights", None) is not None,
        }

        # если панель ещё не создана — просто выйти
        if not hasattr(self, "_status_labels"):
            return

        for name, lbl in self._status_labels.items():
            if states.get(name):
                lbl.configure(text=f"{name}: ✓", text_color=ok_color)
            else:
                lbl.configure(text=f"{name}: —", text_color=off_color)

    def clear_project(self):
        """Очистка всех полей проекта"""
        self.project_name_var.set("")
        self.length_entry.set_text("")
        self.quantity_entry.set_text("")
        # ... очистка всех остальных полей ...
        messagebox.showinfo("Очистка", "Все поля проекта очищены")

    def open_project(self):
        """Загрузка проекта из файла"""
        filepath = filedialog.askopenfilename(filetypes=[("Steps Project", "*.hgp")])
        if filepath:
            try:
                # Логика загрузки проекта
                messagebox.showinfo("Успех", f"Проект загружен из {filepath}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось загрузить проект: {str(e)}")

    def save_project(self):
        """Сохранение проекта в файл"""
        if not self.project_name_var.get():
            messagebox.showerror("Ошибка", "Введите имя проекта")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".hgp",
            filetypes=[("Steps Project", "*.hgp")],
            initialfile=f"{self.project_name_var.get()}.hgp"
        )
        if filepath:
            # Логика сохранения проекта
            messagebox.showinfo("Успех", f"Проект сохранен в {filepath}")

    

    def preview_project(self):
        """Предварительный просмотр проекта"""
        if not self.validate_inputs():
            return
        
        # Создаем окно предпросмотра
        preview_window = ctk.CTkToplevel(self.root)
        preview_window.title("Предпросмотр проекта")
        preview_window.geometry("800x600")
        
        # Добавляем содержимое предпросмотра
        # ... 

    def validate_inputs(self):
        """Проверка заполнения обязательных полей"""
        if not self.length_entry.get_text():
            messagebox.showerror("Ошибка", "Введите длину ступени")
            return False
        if not self.quantity_entry.get_text():
            messagebox.showerror("Ошибка", "Введите количество")
            return False
        # Другие проверки...
        return True        
            
    def create_tables_tab(self):
        ctk.CTkLabel(self.tables_tab, text="Импорт таблиц", font=("Segoe UI", 16, "bold")).pack(anchor="w", padx=12, pady=(12, 0))

        # Панель импорта
        import_group = ctk.CTkFrame(self.tables_tab)
        import_group.pack(fill="x", padx=12, pady=(0, 6))
        for text, name in [
            ("Боковины", "Боковины"),
            ("Кикплейт", "Кикплейт"),
            ("Типы матов", "Типы матов"),
            ("Перфоуголок", "Перфоуголок"),
        ]:
            ctk.CTkButton(import_group, text=text,
                          command=lambda n=name: self.import_single_table(n)).pack(side="left", padx=6, pady=6)

        # Вкладки с таблицами
        self.tables_tabview = ctk.CTkTabview(self.tables_tab)
        self.tables_tabview.pack(fill="both", expand=True, padx=12, pady=(6, 6))

        # Стиль ttk для деревьев
        import tkinter.ttk as ttk
        style = ttk.Style()
        style.configure("Treeview", font=("Segoe UI", 18), rowheight=36)
        style.configure("Treeview.Heading", font=("Segoe UI", 18))

        # Создаём вкладки и деревья
        for name in ["Боковины", "Кикплейт", "Типы матов", "Перфоуголок"]:
            tab = self.tables_tabview.add(name)

            frame = ctk.CTkFrame(tab)
            frame.pack(fill="both", expand=True, padx=6, pady=6)

            tree = ttk.Treeview(frame, show="headings")
            tree.pack(fill="both", expand=True, side="left")

            scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            scroll.pack(side="right", fill="y")
            tree.configure(yscrollcommand=scroll.set)

            

            self.table_trees[name] = tree

        # Первичная инициализация содержимого вкладок
        self.load_tables_to_editor()

    # ---------- Служебные методы ----------

    def check_loaded_tables(self):
        loaded = []
        if self.calculator.side_data is not None:
            loaded.append("Боковины")
        if self.calculator.mat_data is not None:
            loaded.append("Типы матов")
        if self.calculator.angle_weights is not None:
            loaded.append("Перфоуголок")
        if self.calculator.force_weights is not None:
            loaded.append("Кикплейт")

        if loaded:
            print(f"Загружены таблицы: {', '.join(loaded)}")
            self.update_project_comboboxes()
            self.load_tables_to_editor()
        else:
            print("Таблицы не загружены")
            messagebox.showinfo("Информация",
                                "Таблицы не загружены. Пожалуйста, импортируйте их на вкладке 'Таблицы'")

    def update_project_comboboxes(self):
        if self.calculator.side_data is not None:
            values = list(self.calculator.side_data['Размер'].astype(str).unique())
            self.side_combobox.configure(values=values)
        if self.calculator.angle_weights is not None:
            vals = ['Нет'] + list(self.calculator.angle_weights['Размер'].astype(str).unique())
            self.angle_combobox.configure(values=vals)
        if self.calculator.mat_data is not None:
            values = list(self.calculator.mat_data['Артикул'].astype(str).unique())
            self.grid_combobox.configure(values=values)
            if values:
                self.grid_combobox.set(values[0])
        if self.calculator.force_weights is not None:
            vals = ['Нет'] + list(self.calculator.force_weights['Размер'].astype(str).unique())
            self.profile_combobox.configure(values=vals)

    def load_tables_to_editor(self):
        self.calculator_tables = {
            "Перфоуголок": self.calculator.angle_weights,
            "Кикплейт":    self.calculator.force_weights,
            "Типы матов":  self.calculator.mat_data,
            "Боковины":    self.calculator.side_data,
        }
        # Отрисовать каждую таблицу в её дереве (если дерево уже создано)
        for name, df in self.calculator_tables.items():
            tree = self.table_trees.get(name)
            if tree is not None:
                self.display_table_in_tree(name, df, tree)
            
    def display_table_in_tree(self, name: str, df: pd.DataFrame | None, tree):
        """Отобразить DataFrame в указанном Treeview."""
        # Очистка
        tree.delete(*tree.get_children())
        for col in tree["columns"]:
            tree.heading(col, text="")
            tree.column(col, width=0)
        tree["columns"] = []

        if df is None or df.empty:
            return

        # Заголовки
        cols = [str(c) for c in df.columns]
        # Защита от дублей названий столбцов
        seen = {}
        for i, c in enumerate(cols):
            if c in seen:
                cols[i] = f"{c}_{seen[c]}"
                seen[c] += 1
            else:
                seen[c] = 1

        tree["columns"] = cols
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor="center", stretch=False)

        # Данные
        for _, row in df.iterrows():
            values = []
            for v in row:
                if pd.isna(v):
                    values.append("")
                elif isinstance(v, (int, float)):
                    values.append(f"{float(v):.2f}".replace(",", " ").replace(".", ","))  # как было
                else:
                    values.append(str(v))
            tree.insert("", "end", values=values)


    
    

    # ---------- Импорт/сохранение таблиц ----------

    def import_single_table(self, table_type):
        path = filedialog.askopenfilename(
            title=f"Выберите файл для таблицы {table_type}",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not path:
            return

        if self.calculator.load_table(table_type, path):
            messagebox.showinfo("Успех", f"Таблица {table_type} успешно загружена")
            self.update_project_comboboxes()
            self.load_tables_to_editor()
        else:
            messagebox.showerror("Ошибка", f"Не удалось загрузить таблицу {table_type}")

    

    # ---------- Расчет/экспорт ----------

    def calculate(self):
        # Проверки выбора
        side_val = (self.side_combobox.get() or "").strip()
        grid_val = (self.grid_combobox.get() or "").strip()

        if not side_val or side_val.lower() == "нет":
            messagebox.showerror("Ошибка", "Выберите Размер боковины")
            return
        if not grid_val or grid_val.lower() == "нет":
            messagebox.showerror("Ошибка", "Выберите Тип решетки")
            return

        try:
            length = self.length_entry.get_float()
            quantity = self.quantity_entry.get_float()
            work_hours = self.work_hours_entry.get_float()
            complexity = self.complexity_entry.get_float()
            zinc_price = self.zinc_price_entry.get_float()
            angle_price = self.angle_price_entry.get_float()
            kickplate_price = self.kickplate_price_entry.get_float()
            hourly_rate = self.hourly_rate_entry.get_float()

            if length <= 0:
                messagebox.showerror("Ошибка", "Длина должна быть положительной")
                return
            if quantity <= 0:
                messagebox.showerror("Ошибка", "Количество должно быть положительным")
                return
            if work_hours < 0:
                messagebox.showerror("Ошибка", "Часы работы не могут быть отрицательными")
                return
            if complexity <= 0:
                messagebox.showerror("Ошибка", "Коэффициент сложности должен быть положительным")
                return
            if zinc_price < 0:
                messagebox.showerror("Ошибка", "Цена цинкования не может быть отрицательной")
                return
            if angle_price < 0:
                messagebox.showerror("Ошибка", "Цена перфоуголка не может быть отрицательной")
                return
            if kickplate_price < 0:
                messagebox.showerror("Ошибка", "Цена кикплейта не может быть отрицательной")
                return
            if hourly_rate < 0:
                messagebox.showerror("Ошибка", "Ставка работы не может быть отрицательной")
                return

            self.calculator.set_zinc_price(zinc_price)
            self.calculator.set_angle_price(angle_price)
            self.calculator.set_kickplate_price(kickplate_price)
            self.calculator.set_hourly_rate(hourly_rate)

            if self.calculator.side_data is None:
                messagebox.showerror("Ошибка", "Не загружена таблица Боковины")
                return
            if self.calculator.mat_data is None:
                messagebox.showerror("Ошибка", "Не загружена таблица Типы матов")
                return
                
            mats_count = 0
            if self.calculator.current_project and 'mats_count' in self.calculator.current_project:
                mats_count = self.calculator.current_project['mats_count']

            success = self.calculator.create_project(
                length=length,
                quantity=quantity,
                side_choice=self.side_combobox.get(),
                grid_choice=self.grid_combobox.get(),
                angle_choice=self.angle_combobox.get(),
                profile_choice=self.profile_combobox.get(),
                work_hours=work_hours,
                complexity=complexity
            )
            if not success or not self.calculator.current_project or not self.calculator.current_project.get('results'):
                messagebox.showerror("Ошибка", "Не удалось выполнить расчет. Проверьте данные.")
                return
                
            #mats_count = self.mats_count_entry.get_float() if self.mats_count_entry else 0
            #self.calculator.current_project['mats_count'] = mats_count

            self.show_results()

        except ValueError as e:
            messagebox.showerror("Ошибка", f"Неверный числовой формат: {e}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла непредвиденная ошибка: {str(e)}")
            print(f"Подробности ошибки: {e}")
            
    # ---------- Локальный алгоритм раскроя (без сохранения файлов) ----------

    def _build_dense_columns(self, items, mat_width, kerf=5):
        remaining = [dict(d) for d in items]
        remaining.sort(key=lambda d: (d["length"], d["width"]), reverse=True)
        columns = []
        for part in remaining:
            best_col = None
            best_delta = None
            for col in columns:
                need_h = (col["height_used"] == 0) and part["width"] or (kerf + part["width"])
                if col["height_used"] + need_h > mat_width:
                    continue
                new_wx = max(col["width_x"], part["length"])
                delta = new_wx - col["width_x"]
                if (best_delta is None) or (delta < best_delta):
                    best_delta = delta
                    best_col = col
            if best_col is None:
                columns.append({"width_x": part["length"], "height_used": part["width"], "items": [part]})
            else:
                best_col["width_x"] = max(best_col["width_x"], part["length"])
                best_col["items"].append(part)
                best_col["height_used"] += (kerf + part["width"])
        for col in columns:
            col["items"].sort(key=lambda d: d["width"], reverse=True)
        return columns

    def _place_columns_into_mats(self, columns, mat_length, mat_width, kerf=5):
        cols = sorted(columns, key=lambda c: c["width_x"], reverse=True)
        mats = []
        while cols:
            x = 0
            bin_res = []
            i = 0
            while i < len(cols):
                col = cols[i]
                need_x = col["width_x"] if x == 0 else (kerf + col["width_x"])
                if x + need_x <= mat_length:
                    y = 0
                    for j, d in enumerate(col["items"]):
                        if j > 0:
                            y += kerf
                        bin_res.append({
                            "position": d["position"],
                            "length": d["length"],
                            "width": d["width"],
                            "x": x if need_x == col["width_x"] else x + kerf,
                            "y": y
                        })
                        y += d["width"]
                    x += need_x
                    cols.pop(i)
                    continue
                i += 1
            if not bin_res and cols:
                col = cols.pop(0)
                y = 0
                for j, d in enumerate(col["items"]):
                    if j > 0:
                        y += kerf
                    bin_res.append({
                        "position": d["position"],
                        "length": d["length"],
                        "width": d["width"],
                        "x": 0,
                        "y": y
                    })
                    y += d["width"]
            mats.append(bin_res)
        return mats

    def _make_parts_for_cutting(self):
        """
        Формирует список прямоугольников для раскроя:
        позиция — просто порядковый номер, 
        длина = длина ступени, 
        ширина = ширина выбранной боковины.
        """
        if not self.side_combobox.get():
            raise ValueError("Выберите Размер боковины")
        if not self.grid_combobox.get():
            raise ValueError("Выберите Тип решетки")
        length = self.length_entry.get_float()
        qty = int(self.quantity_entry.get_float())
        if length <= 0 or qty <= 0:
            raise ValueError("Длина и Количество должны быть положительными")

        # берём ширину по выбранной боковине из загруженной таблицы
        side_row = _select_row(self.calculator.side_data, 'Размер', self.side_combobox.get())
        width = _to_float(side_row.get('Ширина')) or 0.0
        if width <= 0:
            raise ValueError("В таблице 'Боковины' некорректная ширина")

        parts = []
        for i in range(qty):
            parts.append({"position": str(i + 1), "length": int(length), "width": int(width)})
        return parts

    def calculate_mats(self):
        """Основной метод расчета количества матов с учетом раскроя"""
        try:
            # Получаем размеры мата
            mat_L = int(round(self.mat_length_entry.get_float() or 0)) if hasattr(self, "mat_length_entry") else int(self.mat_length_var.get() or 0)
            mat_W = int(round(self.mat_width_entry.get_float()  or 0)) if hasattr(self, "mat_width_entry")  else int(self.mat_width_var.get()  or 0)

            
            # Проверка ввода
            if mat_L <= 0 or mat_W <= 0:
                messagebox.showerror("Ошибка", "Размеры мата должны быть положительными")
                return

            # Получаем детали для раскроя
            parts = self._make_parts_for_cutting()

            # Фильтрация деталей, которые не помещаются в мат
            filtered = [d for d in parts if d["length"] <= mat_L and d["width"] <= mat_W]
            if len(filtered) < len(parts):
                messagebox.showwarning("Внимание", 
                                     "Некоторые детали больше мата и не будут учтены в раскрое")

            # Оптимизация раскроя
            columns = self._build_dense_columns(filtered, mat_W, kerf=5)
            self._cutting_results = self._place_columns_into_mats(columns, mat_L, mat_W, kerf=5)

            # Получаем итоговое количество матов
            mats_count = len(self._cutting_results)

            # Обновляем интерфейс
            self.mats_count_entry.configure(state="normal")
            self.mats_count_entry.set_text(str(mats_count))
            self.mats_count_entry.configure(state="disabled")

            # Сохраняем в проект
            if self.calculator.current_project is None:
                self.calculator.current_project = {
                    'results': {
                        'unit_price': 0, 
                        'unit_weight': 0,
                        'total_weight': 0, 
                        'total_price': 0
                    }
                }
            self.calculator.current_project['mats_count'] = mats_count

            messagebox.showinfo("Готово", f"Количество матов: {mats_count}")
            
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректные числовые значения")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
            print(f"Ошибка расчета матов: {traceback.format_exc()}")

    def preview_cutting(self):
        if not self._cutting_results:
            # если ещё не считали — посчитаем
            self.calculate_mats()
            if not self._cutting_results:
                return

        mat_L = int(round(self.mat_length_entry.get_float())) if hasattr(self, "mat_length_entry") else int(self.mat_length_var.get())
        mat_W = int(round(self.mat_width_entry.get_float()))  if hasattr(self, "mat_width_entry")  else int(self.mat_width_var.get())

        # Показываем предпросмотр без сохранения файлов: строим фигуры в памяти
        win = ctk.CTkToplevel(self.root)
        win.title("Предпросмотр раскроя")
        win.geometry("1100x800")
        win.grab_set()

        # Индекс текущего изображения
        idx = {"i": 0}

        img_label = ctk.CTkLabel(win, text="")
        img_label.pack(padx=12, pady=12, fill="both", expand=True)

        ctrl = ctk.CTkFrame(win); ctrl.pack(fill="x", padx=12, pady=(0,12))
        def btn(parent, text, cmd, w=120):
            return ctk.CTkButton(parent, text=text, command=cmd, width=w)

        def draw_to_label(i):
            mats = self._cutting_results
            if not mats: return
            fig, ax = plt.subplots(figsize=(11.7, 8.27))
            ax.set_xlim(0, mat_L); ax.set_ylim(0, mat_W)
            ax.set_aspect('equal', adjustable='box'); ax.axis('off')
            ax.add_patch(Rectangle((0,0), mat_L, mat_W, linewidth=2, edgecolor='black', facecolor='none'))

            for d in mats[i]:
                ax.add_patch(Rectangle((d["x"], d["y"]), d["length"], d["width"],
                                       linewidth=1, edgecolor='black', facecolor='none'))
                ax.text(d["x"] + d["length"]/2, d["y"] + d["width"]/2,
                        f'{d["position"]}\n{d["length"]}×{d["width"]}',
                        fontsize=8, ha='center', va='center')
            fig.tight_layout(pad=0); fig.subplots_adjust(left=0.01, right=0.99, top=0.98, bottom=0.01)

            # Рендерим в буфер и показываем в Label (без сохранения на диск)
            from io import BytesIO
            from PIL import Image, ImageTk
            buf = BytesIO()
            fig.savefig(buf, format="png", dpi=150, bbox_inches='tight', pad_inches=0.01)
            plt.close(fig)
            buf.seek(0)
            im = Image.open(buf)
            photo = ImageTk.PhotoImage(im)
            img_label.configure(image=photo, text="")
            img_label.image = photo
            win.title(f"Предпросмотр — Мат {i+1}/{len(mats)}")

        def prev_():
            idx["i"] = (idx["i"] - 1) % len(self._cutting_results)
            draw_to_label(idx["i"])

        def next_():
            idx["i"] = (idx["i"] + 1) % len(self._cutting_results)
            draw_to_label(idx["i"])

        btn(ctrl, "←", prev_, 80).pack(side="left", padx=6, pady=6)
        btn(ctrl, "→", next_, 80).pack(side="left", padx=6, pady=6)

        draw_to_label(idx["i"])
            

    def show_results(self):
        if not self.calculator.current_project:
            return

        results = self.calculator.current_project.get('results')
        intermediate = self.calculator.current_project.get('intermediate_results')
        if not results or not intermediate:
            return

        def fmt(x):
            return f"{x:,.2f}".replace(",", " ").replace(".", ",")

        unit_price_final = results['total_price'] / self.calculator.current_project['quantity']
        quantity = self.calculator.current_project['quantity']
        if quantity > 0:
            unit_price_final = results['total_price'] / quantity
            self.results_vars['unit_price'].set(f"{fmt(unit_price_final)} €")
        else:
            self.results_vars['unit_price'].set("0,00 €")
            
        self.results_vars['unit_weight'].set(f"{fmt(results['unit_weight'])} кг")
        self.results_vars['total_weight'].set(f"{fmt(results['total_weight'])} кг")
        self.results_vars['total_price'].set(f"{fmt(results['total_price'])} €")

        self.intermediate_text.configure(state="normal")
        self.intermediate_text.delete("1.0", "end")

        for component, data in intermediate.items():
            self.intermediate_text.insert("end", f"{component}:\n")
            for key, value in data.items():
                if isinstance(value, (int, float)):
                    value = fmt(value)
                self.intermediate_text.insert("end", f"  {key}: {value}\n")
            self.intermediate_text.insert("end", "\n")

        self.intermediate_text.configure(state="disabled")

    def export_to_excel(self):
        if not self.calculator.current_project or not self.calculator.current_project.get('results'):
            messagebox.showerror("Ошибка", "Сначала выполните расчет")
            return

        raw_name = self.project_name_var.get().strip()
        if not raw_name:
            raw_name = datetime.now().strftime("Проект_%Y-%m-%d_%H-%M")
            self.project_name_var.set(raw_name)

        safe_name = "".join(c for c in raw_name if c.isalnum() or c in " _-").strip()
        if not safe_name:
            safe_name = datetime.now().strftime("Проект_%Y-%m-%d_%H-%M")
            self.project_name_var.set(safe_name)

        default_filename = f"{safe_name}_Steps.xlsx"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Сохранить расчет как...",
            initialfile=default_filename
        )
        if not filepath:
            return

        try:
            if self.calculator.export_to_excel(filepath):
                messagebox.showinfo("Успех", f"Файл успешно сохранен:\n{filepath}")
                try:
                    os.startfile(os.path.dirname(filepath))  # Windows
                except Exception:
                    pass
            else:
                messagebox.showerror("Ошибка", "Не удалось сохранить файл")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при сохранении файла: {str(e)}")


# ---------------------------- Запуск ----------------------------

if __name__ == "__main__":
    # Настройки customtkinter
    ctk.set_appearance_mode("light")              # "light" | "dark" | "system"
    ctk.set_default_color_theme("blue")           # "blue" | "green" | "dark-blue"

    root = ctk.CTk()
    root.geometry("1200x800")
    app = StepCalculatorApp(root)
    root.title("Steps Master v.5.0 ©Viiptag")
    root.mainloop()