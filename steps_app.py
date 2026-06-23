"""
steps_app.py — Steps calculator window (CTkToplevel) integrated into Grid Master.
Uses shared manager_config.json. Only adds side_data (Боковины) specific to stairs.
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
from datetime import datetime
from typing import Callable, Dict, Optional
import openpyxl
from openpyxl.utils import get_column_letter

import price_manager as pm
from steps_core import calc_steps


def _fmt(v: float, decimals: int = 2) -> str:
    try:
        return f"{v:,.{decimals}f}".replace(",", " ").replace(".", ",")
    except Exception:
        return str(v)


def open_steps_window(parent: ctk.CTk, cfg: Dict, save_cfg: Callable):
    """Create and show the Steps calculator window."""
    win = StepsWindow(parent, cfg, save_cfg)
    win.focus()
    return win


class StepsWindow(ctk.CTkToplevel):
    def __init__(self, parent, cfg: Dict, save_cfg: Callable):
        super().__init__(parent)
        self.cfg = cfg
        self.save_cfg = save_cfg
        self._result: Optional[Dict] = None

        self.title("Расчёт ступеней — Steps Master")
        self.geometry("820x700")
        self.resizable(True, True)

        self._build_ui()
        self._refresh_combos()
        self._load_prices()

    # ──────────────────────────────────────────────────────────────────────────
    # UI build
    # ──────────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        main = ctk.CTkFrame(self)
        main.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        main.grid_columnconfigure(0, weight=1)
        main.grid_columnconfigure(1, weight=1)
        main.grid_rowconfigure(3, weight=1)

        # ── Row 0: top bar ────────────────────────────────────────────────────
        top = ctk.CTkFrame(main, fg_color="transparent")
        top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        ctk.CTkLabel(top, text="Расчёт ступеней",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(side="left")
        ctk.CTkButton(top, text="Боковины (Excel)",
                      command=self._import_sides).pack(side="right", padx=4)

        # ── Row 1: left column — params ───────────────────────────────────────
        left = ctk.CTkFrame(main, corner_radius=8)
        left.grid(row=1, column=0, sticky="new", padx=(0, 6), pady=(0, 8))
        left.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(left, text="Параметры",
                     font=ctk.CTkFont(weight="bold")).grid(
                         row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(8, 4))

        fields = [
            ("Длина ступени (мм):", "_e_length"),
            ("Количество (шт):",    "_e_qty"),
        ]
        for r, (lbl, attr) in enumerate(fields, start=1):
            ctk.CTkLabel(left, text=lbl).grid(row=r, column=0, sticky="e", padx=8, pady=4)
            e = ctk.CTkEntry(left)
            e.grid(row=r, column=1, sticky="ew", padx=8, pady=4)
            setattr(self, attr, e)

        combos = [
            ("Боковина:",     "_cb_side",      []),
            ("Мат (решётка):", "_cb_mat",      []),
            ("Перфоуголок:",  "_cb_angle",     ["Нет"]),
            ("Кикплейт:",     "_cb_kickplate", ["Нет"]),
        ]
        base_row = len(fields) + 1
        for r, (lbl, attr, vals) in enumerate(combos, start=base_row):
            ctk.CTkLabel(left, text=lbl).grid(row=r, column=0, sticky="e", padx=8, pady=4)
            cb = ctk.CTkComboBox(left, values=vals or ["—"])
            cb.grid(row=r, column=1, sticky="ew", padx=8, pady=4)
            setattr(self, attr, cb)

        ctk.CTkLabel(left, text="Часы работы (на 1 шт):").grid(
            row=base_row + len(combos), column=0, sticky="e", padx=8, pady=4)
        self._e_hours = ctk.CTkEntry(left)
        self._e_hours.grid(row=base_row + len(combos), column=1, sticky="ew", padx=8, pady=4)
        self._e_hours.insert(0, "0")

        # ── Row 1: right column — prices ──────────────────────────────────────
        right = ctk.CTkFrame(main, corner_radius=8)
        right.grid(row=1, column=1, sticky="new", padx=(6, 0), pady=(0, 8))
        right.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(right, text="Цены",
                     font=ctk.CTkFont(weight="bold")).grid(
                         row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(8, 4))

        price_fields = [
            ("Цинкование (€/кг):",   "_ep_zinc"),
            ("Перфоуголок (€/м):",   "_ep_angle"),
            ("Кикплейт (€/кг):",     "_ep_kick"),
            ("Работа (€/час):",      "_ep_work"),
        ]
        for r, (lbl, attr) in enumerate(price_fields, start=1):
            ctk.CTkLabel(right, text=lbl).grid(row=r, column=0, sticky="e", padx=8, pady=4)
            e = ctk.CTkEntry(right)
            e.grid(row=r, column=1, sticky="ew", padx=8, pady=4)
            e.bind("<FocusOut>", self._on_price_edited)
            setattr(self, attr, e)

        # ── Row 2: action buttons ─────────────────────────────────────────────
        btns = ctk.CTkFrame(main, fg_color="transparent")
        btns.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        ctk.CTkButton(btns, text="Рассчитать", command=self._calculate).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="Экспорт в Excel", command=self._export_xlsx).pack(side="left", padx=4)

        # ── Row 3: results ────────────────────────────────────────────────────
        res_frame = ctk.CTkFrame(main, corner_radius=8)
        res_frame.grid(row=3, column=0, columnspan=2, sticky="nsew")
        res_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        ctk.CTkLabel(res_frame, text="Результаты",
                     font=ctk.CTkFont(weight="bold")).grid(
                         row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(8, 4))

        summary_labels = [
            ("Стоимость ед. (€):", "_lbl_unit_cost"),
            ("Вес ед. (кг):",      "_lbl_unit_weight"),
            ("Общий вес (кг):",    "_lbl_total_weight"),
            ("Итого (€):",         "_lbl_total_cost"),
        ]
        for c, (lbl_text, attr) in enumerate(summary_labels):
            ctk.CTkLabel(res_frame, text=lbl_text).grid(row=1, column=c, sticky="e", padx=(8, 2))
            lbl = ctk.CTkLabel(res_frame, text="—", font=ctk.CTkFont(weight="bold"))
            lbl.grid(row=2, column=c, sticky="w", padx=(2, 8))
            setattr(self, attr, lbl)

        self._txt_details = ctk.CTkTextbox(res_frame, wrap="none", state="disabled")
        self._txt_details.grid(row=3, column=0, columnspan=4, sticky="nsew",
                                padx=8, pady=(4, 8))
        res_frame.grid_rowconfigure(3, weight=1)

    # ──────────────────────────────────────────────────────────────────────────
    # Data helpers
    # ──────────────────────────────────────────────────────────────────────────

    def _refresh_combos(self):
        # Side (Боковины)
        sides = [s["name"] for s in self.cfg.get("side_data", [])]
        self._cb_side.configure(values=sides or ["—"])
        if sides:
            self._cb_side.set(sides[0])

        # Mat articles
        mats = [m["article"] for m in self.cfg.get("mat_data", [])]
        self._cb_mat.configure(values=mats or ["—"])
        if mats:
            self._cb_mat.set(mats[0])

        # Angle (Перфоуголок)
        angles = ["Нет"] + [item[0] for item in self.cfg.get("angle_weights", [])]
        self._cb_angle.configure(values=angles)
        self._cb_angle.set("Нет")

        # Kickplate (Кикплейт / Отбойник)
        bumpers = ["Нет"] + [item[0] for item in self.cfg.get("bumper_weights", [])]
        self._cb_kickplate.configure(values=bumpers)
        self._cb_kickplate.set("Нет")

    def _load_prices(self):
        prices = self.cfg.get("steps_prices", {})
        for attr, key in [
            ("_ep_zinc",  "zinc_eur_kg"),
            ("_ep_angle", "angle_eur_m"),
            ("_ep_kick",  "kickplate_eur_kg"),
            ("_ep_work",  "work_eur_h"),
        ]:
            entry: ctk.CTkEntry = getattr(self, attr)
            entry.delete(0, "end")
            entry.insert(0, str(prices.get(key, "0")))

    def _save_prices(self):
        prices = self.cfg.setdefault("steps_prices", {})
        for attr, key in [
            ("_ep_zinc",  "zinc_eur_kg"),
            ("_ep_angle", "angle_eur_m"),
            ("_ep_kick",  "kickplate_eur_kg"),
            ("_ep_work",  "work_eur_h"),
        ]:
            try:
                val = float(getattr(self, attr).get().replace(",", "."))
            except ValueError:
                val = 0.0
            prices[key] = val
        self.save_cfg(self.cfg)

    def _on_price_edited(self, _event=None):
        self._save_prices()

    # ──────────────────────────────────────────────────────────────────────────
    # Import Боковины Excel
    # ──────────────────────────────────────────────────────────────────────────

    def _import_sides(self):
        path = filedialog.askopenfilename(
            title="Выберите Excel с Боковинами",
            filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")],
        )
        if not path:
            return
        try:
            side_data = pm.load_side_data(path)
            self.cfg["side_data"] = side_data
            self.cfg.setdefault("excel_files", {})["sides"] = path
            self.save_cfg(self.cfg)
            self._refresh_combos()
            messagebox.showinfo("Боковины", f"Загружено записей: {len(side_data)}")
        except Exception as ex:
            messagebox.showerror("Ошибка", str(ex))

    # ──────────────────────────────────────────────────────────────────────────
    # Calculate
    # ──────────────────────────────────────────────────────────────────────────

    def _get_float(self, entry: ctk.CTkEntry, name: str) -> Optional[float]:
        try:
            return float(entry.get().replace(",", "."))
        except ValueError:
            messagebox.showerror("Ввод", f"Неверное значение: {name}")
            return None

    def _calculate(self):
        self._save_prices()

        length = self._get_float(self._e_length, "Длина ступени")
        qty_s  = self._get_float(self._e_qty,    "Количество")
        hours  = self._get_float(self._e_hours,  "Часы работы")
        if length is None or qty_s is None or hours is None:
            return
        qty = max(1, int(qty_s))

        side_name  = self._cb_side.get()
        mat_art    = self._cb_mat.get()
        angle_name = self._cb_angle.get()
        kick_name  = self._cb_kickplate.get()

        if not self.cfg.get("side_data"):
            messagebox.showwarning("Данные", "Загрузите таблицу Боковин (кнопка сверху)")
            return
        if not self.cfg.get("mat_data"):
            messagebox.showwarning("Данные", "Таблица матов не загружена")
            return

        result = calc_steps(
            length_mm=length,
            quantity=qty,
            side_name=side_name,
            mat_article=mat_art,
            angle_name=angle_name if angle_name != "Нет" else None,
            kickplate_name=kick_name if kick_name != "Нет" else None,
            work_hours=hours,
            cfg=self.cfg,
        )

        if not result["ok"]:
            messagebox.showerror("Расчёт", result.get("error", "Ошибка"))
            return

        self._result = {"params": {
            "length_mm": length, "qty": qty,
            "side": side_name, "mat": mat_art,
            "angle": angle_name, "kick": kick_name,
            "hours": hours,
        }, **result}

        self._show_result(result)

    def _show_result(self, result: Dict):
        t = result["totals"]
        b = result["breakdown"]

        self._lbl_unit_cost.configure(   text=f"{_fmt(t['unit_cost_eur'])} €")
        self._lbl_unit_weight.configure( text=f"{_fmt(t['total_weight_kg'] / max(1, self._result['params']['qty']), 3)} кг")
        self._lbl_total_weight.configure(text=f"{_fmt(t['total_weight_kg'], 3)} кг")
        self._lbl_total_cost.configure(  text=f"{_fmt(t['total_cost_eur'])} €")

        lines = []
        if "sides" in b:
            s = b["sides"]
            lines.append(f"Боковины:    {s['pcs']} шт  |  вес {_fmt(s['weight_kg'], 3)} кг  |  {_fmt(s['cost_eur'])} €")
        if "grid" in b:
            g = b["grid"]
            lines.append(f"Решётка:     площадь {_fmt(g['area_m2_per_unit'], 4)} м²/шт  |  вес {_fmt(g['weight_kg'], 3)} кг  |  {_fmt(g['cost_eur'])} €")
        if "angle" in b:
            a = b["angle"]
            lines.append(f"Перфоуголок: {_fmt(a['len_m'], 3)} м  |  вес {_fmt(a['weight_kg'], 3)} кг  |  {_fmt(a['cost_eur'])} €")
        if "kickplate" in b:
            k = b["kickplate"]
            lines.append(f"Кикплейт:    вес {_fmt(k['weight_kg'], 3)} кг  |  {_fmt(k['cost_eur'])} €")
        if "work" in b:
            w = b["work"]
            lines.append(f"Работа:      {_fmt(w['hours_total'], 1)} ч  |  {_fmt(w['cost_eur'])} €")
        if "zinc" in b:
            z = b["zinc"]
            lines.append(f"Цинкование:  {_fmt(z['weight_with_zinc_kg'], 3)} кг (+10%)  |  {_fmt(z['cost_eur'])} €")
        lines.append("─" * 60)
        lines.append(f"Материалы:   {_fmt(t['material_cost_eur'])} €")
        lines.append(f"Работа:      {_fmt(t['work_cost_eur'])} €")
        lines.append(f"Цинк:        {_fmt(t['zinc_cost_eur'])} €")
        lines.append(f"ИТОГО:       {_fmt(t['total_cost_eur'])} €")

        if result.get("warnings"):
            lines.append("")
            for w in result["warnings"]:
                lines.append(f"⚠ {w}")

        self._txt_details.configure(state="normal")
        self._txt_details.delete("1.0", "end")
        self._txt_details.insert("end", "\n".join(lines))
        self._txt_details.configure(state="disabled")

    # ──────────────────────────────────────────────────────────────────────────
    # Excel export
    # ──────────────────────────────────────────────────────────────────────────

    def _export_xlsx(self):
        if not self._result:
            messagebox.showwarning("Экспорт", "Сначала выполните расчёт")
            return

        date_str = datetime.now().strftime("%Y-%m-%d")
        default_name = f"Ступени_{date_str}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name,
        )
        if not path:
            return

        try:
            _write_xlsx(path, self._result)
            messagebox.showinfo("Экспорт", f"Сохранено: {path}")
        except Exception as ex:
            messagebox.showerror("Ошибка", str(ex))


def _write_xlsx(path: str, result: Dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ступени"

    bold = openpyxl.styles.Font(bold=True)
    center = openpyxl.styles.Alignment(horizontal="center")
    num2 = "#,##0.00"

    p = result["params"]
    t = result["totals"]
    b = result["breakdown"]

    ws.append(["Расчёт ступеней", "", datetime.now().strftime("%d.%m.%Y")])
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=13)
    ws.append([])
    ws.append(["Параметры"])
    ws["A3"].font = bold
    ws.append(["Длина ступени (мм)", p["length_mm"]])
    ws.append(["Количество (шт)",    p["qty"]])
    ws.append(["Боковина",           p["side"]])
    ws.append(["Мат",                p["mat"]])
    ws.append(["Перфоуголок",        p["angle"] or "—"])
    ws.append(["Кикплейт",           p["kick"] or "—"])
    ws.append(["Часы/шт",            p["hours"]])
    ws.append([])

    ws.append(["Компонент", "Кол-во/Объём", "Ед.", "Стоимость (€)"])
    header_row = ws.max_row
    for c in range(1, 5):
        ws.cell(header_row, c).font = bold
        ws.cell(header_row, c).alignment = center

    def row(name, qty_val, unit, cost):
        ws.append([name, qty_val, unit, cost])
        ws.cell(ws.max_row, 4).number_format = num2

    if "sides" in b:
        s = b["sides"]
        row("Боковины", s["pcs"], "шт", s["cost_eur"])
    if "grid" in b:
        g = b["grid"]
        row("Решётка", round(g["area_m2_per_unit"] * p["qty"], 4), "м²", g["cost_eur"])
    if "angle" in b:
        a = b["angle"]
        row("Перфоуголок", a["len_m"], "м", a["cost_eur"])
    if "kickplate" in b:
        k = b["kickplate"]
        row("Кикплейт", k["weight_kg"], "кг", k["cost_eur"])
    if "work" in b:
        w = b["work"]
        row("Работа", w["hours_total"], "ч", w["cost_eur"])
    if "zinc" in b:
        z = b["zinc"]
        row("Цинкование", z["weight_with_zinc_kg"], "кг", z["cost_eur"])

    ws.append([])
    ws.append(["Общий вес (кг)",    t["total_weight_kg"]])
    ws.append(["Вес с цинком (кг)", t["total_weight_zinc_kg"]])
    ws.append(["ИТОГО (€)", "", "", t["total_cost_eur"]])
    total_row = ws.max_row
    for c in (1, 4):
        ws.cell(total_row, c).font = bold
    ws.cell(total_row, 4).number_format = num2

    # autofit
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            (len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    wb.save(path)
